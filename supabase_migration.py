import os
import pandas as pd
from openpyxl import load_workbook
import supabase
import uuid
import json
from dotenv import load_dotenv
from openpyxl.utils import get_column_letter

# Load environment variables
load_dotenv()

# Initialize Supabase client
supabase_url = os.getenv("SUPABASE_URL")
supabase_key = os.getenv("SUPABASE_KEY")

if not supabase_url or not supabase_key:
    print("Error: Supabase credentials not found. Please set SUPABASE_URL and SUPABASE_KEY in .env file")
    exit(1)

client = supabase.create_client(supabase_url, supabase_key)

# Excel file path
EXCEL_PATH = "APPEARANCE.xlsx"

def create_tables():
    """
    Create necessary tables in Supabase.
    
    Note: Supabase doesn't support direct table creation via the Python client.
    These SQL commands need to be executed in the Supabase SQL editor.
    This function serves as documentation for the required schema.
    """
    print("To create the required tables, run the following SQL in the Supabase SQL Editor:")
    
    # SQL for teams table
    teams_sql = """
    CREATE TABLE IF NOT EXISTS teams (
        id UUID PRIMARY KEY DEFAULT uuid_generate_v4(),
        name TEXT NOT NULL UNIQUE,
        created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP
    );
    """
    
    # SQL for players table
    players_sql = """
    CREATE TABLE IF NOT EXISTS players (
        id UUID PRIMARY KEY DEFAULT uuid_generate_v4(),
        name TEXT NOT NULL,
        team_id UUID REFERENCES teams(id) ON DELETE CASCADE,
        value TEXT,
        salary TEXT,
        total_appearances INTEGER DEFAULT 0,
        created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(name, team_id)
    );
    """
    
    # SQL for appearances table
    appearances_sql = """
    CREATE TABLE IF NOT EXISTS appearances (
        id UUID PRIMARY KEY DEFAULT uuid_generate_v4(),
        player_id UUID REFERENCES players(id) ON DELETE CASCADE,
        match_day TEXT NOT NULL,
        date DATE,
        appeared BOOLEAN DEFAULT TRUE,
        match_id UUID REFERENCES matches(id) ON DELETE SET NULL,
        created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP
    );
    """
    
    # SQL for matches table
    matches_sql = """
    CREATE TABLE IF NOT EXISTS matches (
        id UUID PRIMARY KEY DEFAULT uuid_generate_v4(),
        home_team_id UUID REFERENCES teams(id) ON DELETE CASCADE,
        away_team_id UUID REFERENCES teams(id) ON DELETE CASCADE,
        match_day TEXT NOT NULL,
        date DATE DEFAULT CURRENT_DATE,
        created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP
    );
    """
    
    print(teams_sql)
    print(players_sql)
    print(matches_sql)
    print(appearances_sql)
    
    print("\nCreate the tables in this order: teams, players, matches, appearances")
    print("After creating the tables, return to this script to migrate your Excel data.")

def upsert_teams():
    """
    Extract teams from Excel sheets and insert them into Supabase
    """
    print(f"Loading Excel workbook from {EXCEL_PATH}...")
    
    if not os.path.exists(EXCEL_PATH):
        print(f"Error: Excel file not found at {EXCEL_PATH}")
        return None
    
    # Open the workbook
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    
    # Extract team names (sheet names)
    teams = []
    for sheet_name in wb.sheetnames:
        # Skip common non-team sheets
        if sheet_name.lower() not in ['summary', 'stats', 'instructions', 'data', 'settings', 'config', 'history']:
            teams.append(sheet_name)
    
    wb.close()
    
    print(f"Found {len(teams)} teams: {teams}")
    
    # Create team objects for database
    team_objects = [{"name": team} for team in teams]
    
    # Insert teams into Supabase
    result = []
    for team in team_objects:
        # Check if team already exists
        existing = client.table("teams").select("id, name").eq("name", team["name"]).execute()
        
        if len(existing.data) == 0:
            # Team doesn't exist, insert it
            response = client.table("teams").insert(team).execute()
            result.append(response.data[0])
            print(f"Added new team: {team['name']}")
        else:
            # Team exists, return existing data
            result.append(existing.data[0])
            print(f"Team already exists: {team['name']}")
    
    print(f"Processed {len(result)} teams")
    return result

def extract_players_from_excel(team_name, team_id):
    """
    Extract players from an Excel sheet for the given team
    """
    try:
        print(f"\nExtracting players from sheet: {team_name}")
        
        # Read the Excel sheet for the team
        excel_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), EXCEL_PATH)
        print(f"Looking for Excel file at: {excel_path}")
        
        if not os.path.exists(excel_path):
            print(f"ERROR: Excel file not found at {excel_path}")
            return []
            
        df = pd.read_excel(excel_path, sheet_name=team_name)
        
        # Print column information for debugging
        print(f"Sheet '{team_name}' columns: {df.columns.tolist()}")
        print(f"Sheet dimensions: {df.shape[0]} rows x {df.shape[1]} columns")
        
        # Initialize player list
        players = []
        
        # Assume first column is player names (usually Column A)
        player_column = df.iloc[:, 0]  # Column A
        
        # Get other columns if they exist
        value_column = df.iloc[:, 1] if df.shape[1] > 1 else None  # Column B if exists
        salary_column = df.iloc[:, 2] if df.shape[1] > 2 else None  # Column C if exists
        appearances_column = df.iloc[:, 3] if df.shape[1] > 3 else None  # Column D if exists
        
        # Determine if first row is header
        is_header_row = False
        if isinstance(player_column.iloc[0], str) and player_column.iloc[0].lower() in ["player", "player name", "name"]:
            is_header_row = True
            print(f"Detected header row: {player_column.iloc[0]}")
            
        start_row = 1 if is_header_row else 0
        
        # Debug output for the first few rows
        print("\nFirst 5 rows of data:")
        for i in range(start_row, min(start_row + 5, len(player_column))):
            player_name = player_column.iloc[i] if i < len(player_column) else None
            value = value_column.iloc[i] if value_column is not None and i < len(value_column) else None
            salary = salary_column.iloc[i] if salary_column is not None and i < len(salary_column) else None
            print(f"Row {i+1}: Name: {player_name}, Value: {value}, Salary: {salary}")
        
        # Extract player data - iterate through all rows
        skipped_rows = 0
        for i in range(start_row, len(player_column)):
            name = player_column.iloc[i]
            
            # Skip empty rows
            if pd.isna(name) or not str(name).strip():
                skipped_rows += 1
                continue
            
            # Create player object
            value = "Unknown"
            salary = "Unknown"
            appearances = 0
            
            # Get value if available
            if value_column is not None and i < len(value_column) and not pd.isna(value_column.iloc[i]):
                value = str(value_column.iloc[i]).strip()
            
            # Get salary if available
            if salary_column is not None and i < len(salary_column) and not pd.isna(salary_column.iloc[i]):
                salary = str(salary_column.iloc[i]).strip()
            
            # Get appearances if available
            if appearances_column is not None and i < len(appearances_column) and not pd.isna(appearances_column.iloc[i]):
                try:
                    appearances = int(appearances_column.iloc[i])
                except (ValueError, TypeError):
                    appearances = 0
            
            player = {
                "name": str(name).strip(),
                "team_id": team_id,
                "value": value,
                "salary": salary,
                "total_appearances": appearances,
            }
            
            players.append(player)
        
        total_rows = len(player_column) - start_row
        print(f"\nExtracted {len(players)} players from {team_name} sheet")
        print(f"Skipped {skipped_rows} empty rows")
        print(f"Processed {total_rows} total rows")
        
        # Print the first few extracted players for verification
        if players:
            print("\nSample of extracted players:")
            for i, player in enumerate(players[:5]):
                print(f"Player {i+1}: {player['name']} - {player['value']} - {player['salary']} - {player['total_appearances']} appearances")
        else:
            print("WARNING: No players were extracted from this sheet")
        
        return players
    
    except Exception as e:
        print(f"ERROR extracting players from {team_name}: {str(e)}")
        import traceback
        traceback.print_exc()
        return []

def upsert_players(team_data):
    """
    Extract players from Excel and insert them into Supabase
    """
    all_players = []
    
    for team in team_data:
        team_id = team["id"]
        team_name = team["name"]
        
        print(f"\n=== Processing team: {team_name} ===")
        
        # Extract players for this team
        players = extract_players_from_excel(team_name, team_id)
        
        if not players:
            print(f"No players found for team {team_name}")
            continue
            
        print(f"Upserting {len(players)} players to Supabase for team {team_name}")
        
        # Insert players into Supabase
        success_count = 0
        for idx, player in enumerate(players):
            try:
                # Check if player already exists
                existing = client.table("players") \
                    .select("id, name, team_id") \
                    .eq("name", player["name"]) \
                    .eq("team_id", player["team_id"]) \
                    .execute()
                
                if len(existing.data) == 0:
                    # Player doesn't exist, insert it
                    response = client.table("players").insert(player).execute()
                    if response.data:
                        all_players.append(response.data[0])
                        success_count += 1
                        if idx % 10 == 0:  # Print progress every 10 players
                            print(f"Added new player: {player['name']} to team {team_name} ({idx+1}/{len(players)})")
                else:
                    # Player exists, update the data
                    player_id = existing.data[0]["id"]
                    response = client.table("players").update(player).eq("id", player_id).execute()
                    if response.data:
                        all_players.append(response.data[0])
                        success_count += 1
                        if idx % 10 == 0:  # Print progress every 10 players
                            print(f"Updated player: {player['name']} in team {team_name} ({idx+1}/{len(players)})")
            except Exception as e:
                print(f"Error processing player {player['name']}: {str(e)}")
        
        print(f"Successfully processed {success_count}/{len(players)} players for team {team_name}")
    
    print(f"\nTotal processed players: {len(all_players)}")
    return all_players

def extract_appearances(team_id, team_name, player_mapping):
    """
    Extract player appearances from Excel sheets
    """
    print(f"\n=== Extracting appearances for team: {team_name} ===")
    appearances = []
    
    try:
        # Load Excel file
        excel_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), EXCEL_PATH)
        wb = load_workbook(excel_path, data_only=True)
        
        # Find the sheet for this team
        sheet_name = None
        for name in wb.sheetnames:
            if name.upper() == team_name.upper():
                sheet_name = name
                break
        
        if not sheet_name:
            print(f"Sheet not found for team {team_name}")
            return appearances
        
        sheet = wb[sheet_name]
        
        # Get column headers
        headers = [cell.value for cell in sheet[1]] if sheet.max_row > 0 else []
        print(f"Headers found: {headers}")
        
        # Identify match day columns
        match_day_cols = []
        for col_idx, header in enumerate(headers, 1):
            if header and isinstance(header, str):
                # Check for MD in the header name
                if 'MD' in header.upper():
                    match_name = header
                    col_letter = get_column_letter(col_idx)
                    match_day_cols.append((col_letter, match_name))
                    print(f"Found match day column: {match_name} at column {col_letter}")
            # Also check for numeric columns that might be appearance data
            elif header and isinstance(header, (int, float)):
                match_name = f"MD{int(header)}"
                col_letter = get_column_letter(col_idx)
                match_day_cols.append((col_letter, match_name))
                print(f"Found numeric match day column: {match_name} at column {col_letter}")
        
        print(f"Found {len(match_day_cols)} match day columns for team {team_name}")
        
        if not match_day_cols:
            print(f"No match day columns found for team {team_name}")
            return appearances
        
        # Find the player name column
        player_col = None
        for col_idx, header in enumerate(headers, 1):
            if header and isinstance(header, str) and header.upper() in ['PLAYER', 'NAME', 'PLAYER NAME']:
                player_col = get_column_letter(col_idx)
                break
        
        if not player_col:
            player_col = 'A'  # Default to first column
            print(f"Player name column not found for team {team_name}, using column A")
        else:
            print(f"Using player name column: {player_col}")
        
        # Process each row and extract match day appearances
        for row_idx in range(2, sheet.max_row + 1):  # Start from row 2 (skip header)
            player_name_cell = sheet[f"{player_col}{row_idx}"]
            player_name = player_name_cell.value
            
            # Skip empty rows
            if not player_name:
                continue
                
            player_name = str(player_name).strip().upper()
            
            # Find player_id from mapping
            player_id = None
            for db_player in player_mapping:
                if db_player["team_id"] == team_id and db_player["name"].upper() == player_name:
                    player_id = db_player["id"]
                    break
            
            if not player_id:
                print(f"Player {player_name} not found in database for team {team_name}")
                continue
                
            # Process each match day column
            for col_letter, match_name in match_day_cols:
                cell_value = sheet[f"{col_letter}{row_idx}"].value
                
                # Skip if appearance value is not 1 (player didn't play)
                if cell_value not in [1, '1', 1.0]:
                    continue
                    
                # Extract match day number from match name
                match_day = None
                if 'MD' in match_name.upper():
                    # Try to extract numeric part from MD string
                    md_parts = match_name.upper().split('MD')
                    if len(md_parts) > 1:
                        try:
                            # Extract just the numeric part
                            md_numeric = ''.join(c for c in md_parts[1] if c.isdigit())
                            if md_numeric:
                                match_day = int(md_numeric)
                        except ValueError:
                            pass
                
                if match_day is None:
                    try:
                        # Try parsing as a direct number
                        if isinstance(match_name, (int, float)):
                            match_day = int(match_name)
                        else:
                            # Try extracting numbers from string
                            match_day = int(''.join(c for c in match_name if c.isdigit()))
                    except (ValueError, TypeError):
                        # If we can't parse it, use a sequential number
                        match_day = len(appearances) + 1
                
                appearance = {
                    "player_id": player_id,
                    "team_id": team_id,
                    "match_day": match_day,
                    "match_name": match_name
                }
                appearances.append(appearance)
                print(f"Added appearance for {player_name} on {match_name} (MD{match_day})")
        
        print(f"Extracted {len(appearances)} total appearances for team {team_name}")
                
    except Exception as e:
        print(f"Error extracting appearances for team {team_name}: {str(e)}")
        traceback.print_exc()
    
    return appearances

def upsert_appearances(team_data, player_data):
    """
    Extract appearances from Excel and insert them into Supabase
    """
    all_appearances = []
    team_appearance_counts = {}
    
    for team in team_data:
        team_id = team["id"]
        team_name = team["name"]
        
        # Get players for this team
        team_players = [p for p in player_data if p["team_id"] == team_id]
        print(f"\n=== Processing appearances for team: {team_name} ({len(team_players)} players) ===")
        
        if not team_players:
            print(f"No players found for team {team_name}, skipping appearances")
            continue
        
        # Extract appearances for this team
        appearances = extract_appearances(team_id, team_name, team_players)
        
        # Keep track of appearances per team
        team_appearance_counts[team_name] = len(appearances)
        
        # Insert appearances into Supabase
        if appearances:
            try:
                response = client.table("appearances").insert(appearances).execute()
                inserted_count = len(response.data)
                all_appearances.extend(response.data)
                print(f"Successfully added {inserted_count} appearances for team {team_name}")
                
                # Player appearance summary
                player_appearances = {}
                for appearance in appearances:
                    player_id = appearance["player_id"]
                    player_name = next((p["name"] for p in team_players if p["id"] == player_id), "Unknown")
                    if player_name not in player_appearances:
                        player_appearances[player_name] = 0
                    player_appearances[player_name] += 1
                
                # Print players with most appearances
                print("\nPlayers with most appearances:")
                sorted_appearances = sorted(player_appearances.items(), key=lambda x: x[1], reverse=True)
                for i, (player_name, count) in enumerate(sorted_appearances[:5]):
                    print(f"{i+1}. {player_name}: {count} matches")
                    
            except Exception as e:
                print(f"Error inserting appearances for team {team_name}: {str(e)}")
                import traceback
                traceback.print_exc()
    
    # Print appearance summary by team
    print("\n=== Appearance Summary by Team ===")
    for team_name, count in sorted(team_appearance_counts.items()):
        print(f"- {team_name}: {count} appearances")
    
    print(f"\nProcessed {len(all_appearances)} total appearances across all teams")
    return all_appearances

def main():
    """
    Main function to migrate data from Excel to Supabase
    """
    print("Starting Excel to Supabase migration...")
    
    # First, print table creation instructions
    create_tables()
    
    # Ask user if tables have been created
    response = input("\nHave you created the required tables in Supabase? (y/n): ")
    if response.lower() != 'y':
        print("Please create the tables first, then run this script again.")
        return
    
    # Migrate teams
    team_data = upsert_teams()
    if not team_data:
        print("Error: Failed to migrate teams.")
        return
    
    # Migrate players
    player_data = upsert_players(team_data)
    if not player_data:
        print("Error: Failed to migrate players.")
        return
    
    # Migrate appearances
    appearances_data = upsert_appearances(team_data, player_data)
    
    # Print summary of players per team
    print("\n=== Player Count Summary ===")
    team_player_counts = {}
    for player in player_data:
        team_id = player["team_id"]
        team_name = next((team["name"] for team in team_data if team["id"] == team_id), "Unknown")
        if team_name not in team_player_counts:
            team_player_counts[team_name] = 0
        team_player_counts[team_name] += 1
    
    for team_name, count in sorted(team_player_counts.items()):
        print(f"- {team_name}: {count} players")
    
    print("\nMigration completed successfully!")
    print(f"- {len(team_data)} teams")
    print(f"- {len(player_data)} players")
    print(f"- {len(appearances_data)} appearances")

if __name__ == "__main__":
    main() 