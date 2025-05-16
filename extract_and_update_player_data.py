import cv2
import os
import time
import re
# Add PIL ANTIALIAS compatibility before importing easyocr
from PIL import Image
# Handle PIL ANTIALIAS deprecation in newer versions
if not hasattr(Image, 'ANTIALIAS'):
    # For newer Pillow versions (>=9.0.0)
    Image.ANTIALIAS = Image.Resampling.LANCZOS

import easyocr
import unicodedata
import sys
from openpyxl import load_workbook
import traceback

# Global variables
excel_path = "APPEARANCE.xlsx"  # Default path
default_match_day = "MD1"
reader = None  # Global reader to reuse across videos

# Excel locations (based on our analysis)
PLAYER_NAME_COL = 1  # Column A
MD1_COL = 5  # Column I (not 93)
MD2_COL = 6  # Column J
MD3_COL = 7  # Column K
MD4_COL = 8  # Column L
MD5_COL = 9  # Column M
PLAYER_NAME_ROW_START = 2  # Row 7 is the header row
ADDITIONAL_PLAYERS_ROW = 60  # Row 60 for unmatched players

# Verify Excel file exists and can be opened
def verify_excel_file(excel_path):
    if not os.path.exists(excel_path):
        print(f"Error: Excel file '{excel_path}' not found.")
        alt_path = input(f"Enter the correct path to your Excel file (or press Enter to use '{excel_path}'): ")
        if alt_path.strip():
            return alt_path
        return None
    
    try:
        # Test if we can open the workbook
        workbook = load_workbook(excel_path, read_only=True)
        workbook.close()
        return excel_path
    except Exception as e:
        print(f"Error opening Excel file: {str(e)}")
        print("This could be due to a corrupted file or an open file in Excel.")
        alt_path = input(f"Enter the path to a different Excel file (or press Enter to try again with '{excel_path}'): ")
        if alt_path.strip():
            return alt_path
        return None

# Initialize OCR reader
def initialize_ocr():
    global reader
    if reader is None:
        print("Initializing OCR reader...")
        reader = easyocr.Reader(['en'])
    return reader

# Extract team names from video filename
def extract_teams_from_filename(video_path):
    # Attempt to extract team names from a pattern like "team1 vs team2 md1.mp4"
    filename = os.path.basename(video_path).lower()
    match = re.match(r'([a-z]+)\s+vs\s+([a-z]+)', filename)
    
    if match:
        home_team = match.group(1).upper()
        away_team = match.group(2).upper()
        print(f"Detected teams from filename: {home_team} vs {away_team}")
        return home_team, away_team
    
    # If pattern doesn't match, prompt user for team names
    print(f"Could not detect team names from filename: {filename}")
    home_team = input("Enter HOME team name (e.g., AJAX): ").upper()
    away_team = input("Enter AWAY team name (e.g., BARCA): ").upper()
    return home_team, away_team

# Extract frames from the video
def extract_frames(video_path, output_dir="frames"):
    print(f"Extracting frames from {video_path}...")
    
    # Create output directory if it doesn't exist
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # Open the video
    cap = cv2.VideoCapture(video_path)
    
    if not cap.isOpened():
        print("Error: Could not open video.")
        return []
    
    # Get video properties
    fps = cap.get(cv2.CAP_PROP_FPS)
    total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    duration = total_frames / fps
    
    print(f"Video duration: {duration:.2f} seconds, FPS: {fps}, Total frames: {total_frames}")
    
    # Extract more frames to ensure we don't miss any player rating screens
    # For short videos, extract frames more frequently
    if duration < 30:
        frame_interval = max(1, int(fps / 3))  # Capture 3 frames per second
    else:
        frame_interval = max(1, int(fps / 2))  # Capture 2 frames per second
    
    print(f"Using frame interval of {frame_interval} (approximately {fps/frame_interval:.1f} frames per second)")
    
    frame_count = 0
    saved_frames = []
    
    while True:
        ret, frame = cap.read()
        if not ret:
            break
        
        if frame_count % frame_interval == 0:
            frame_path = os.path.join(output_dir, f"frame_{frame_count}.jpg")
            cv2.imwrite(frame_path, frame)
            saved_frames.append(frame_path)
            
            # Print progress
            if len(saved_frames) % 10 == 0:
                print(f"Extracted {len(saved_frames)} frames...")
        
        frame_count += 1
    
    cap.release()
    print(f"Total frames extracted: {len(saved_frames)}")
    return saved_frames

# Extract text from image using OCR
def extract_text_from_image(image_path, reader):
    # Read the image
    image = cv2.imread(image_path)
    
    # Get OCR results
    results = reader.readtext(image)
    
    # Return the detected text
    return results

# Process text to extract player information
def process_text_data(ocr_results):
    # Initialize player data containers
    home_players = []
    away_players = []
    current_section = None
    
    # Combine results into a single string
    text_blocks = []
    for detection in ocr_results:
        text = detection[1]
        confidence = detection[2]
        # Lower confidence threshold to catch more text but filter later
        if confidence > 0.2:
            text_blocks.append(text.strip())
    
    # Join all detected text
    all_text = " ".join(text_blocks)
    
    # Determine if we're looking at home or away team ratings
    if "Player Ratings: Home" in all_text or "Player Ratings: Home" in " ".join(text_blocks):
        current_section = "home"
    elif "Player Ratings: Away" in all_text or "Player Ratings: Away" in " ".join(text_blocks):
        current_section = "away"
    
    # Known false positives to exclude
    false_positives = ["ack maia", "back", "pack", "sack", "ack", "acc", "maia", "main", "main menu", 
                       "page", "click", "next", "prev", "menu", "rank", "rankings", "overall", "squad",
                       "default", "formation", "tactics", "attacking", "balanced", "defensive", "irwf"]
    
    # Extract player names
    if current_section:
        # First, try to find full names (multi-word entries)
        for i, text in enumerate(text_blocks):
            # Skip ratings values (usually just numbers like 6.5)
            if re.match(r'^[\d.]+$', text):
                continue
                
            # Skip short position indicators (like GK, CB, etc.)
            if re.match(r'^[A-Z]{2,3}$', text):
                continue
                
            # Skip "Player Ratings:" and navigation elements
            if "Player Ratings" in text or "Back" in text or "Away" == text or "Home" == text:
                continue
                
            # Check against known false positives (UI elements misread as names)
            if any(fp in text.lower() for fp in false_positives):
                continue
                
            # First pass - full names (first and last name)
            if " " in text and len(text.split()) >= 2:
                # Check for common suffixes like "90" or "95" often seen in player ratings
                clean_name = re.sub(r'\s+\d+$', '', text)
                
                # Clean up OCR errors in names
                clean_name = clean_name.replace("~", "I")
                clean_name = clean_name.replace("l1", "h")
                clean_name = clean_name.replace("1", "i")
                clean_name = clean_name.replace("0", "o")
                
                # Fix some specific player names with common OCR errors
                if "idiger" in clean_name.lower() or "iidiger" in clean_name.lower() or "rudiger" in clean_name.lower():
                    clean_name = "Antonio Rüdiger"
                    
                if "rowe" in clean_name.lower() and "smith" in clean_name.lower():
                    clean_name = "Emile Smith Rowe"  # Ensure full name
                
                if "asensio" in clean_name.lower() or "asencio" in clean_name.lower():
                    clean_name = "Marco Asensio"
                    
                if "tadic" in clean_name.lower():
                    clean_name = "Dusan Tadic"
                    
                if "lloris" in clean_name.lower():
                    clean_name = "Hugo Lloris"
                    
                if "aspas" in clean_name.lower() or "llago" in clean_name.lower() or "iago" in clean_name.lower():
                    clean_name = "Iago Aspas"
                    
                if "roon" in clean_name.lower() or "de ro" in clean_name.lower():
                    clean_name = "Marten de Roon"
                
                # Remove common non-name text
                clean_name = re.sub(r'SUB|SUBS|SUBSTITUTION|RATING', '', clean_name, flags=re.IGNORECASE)
                clean_name = re.sub(r'\(\d+\)|\d+\'|\d+\'\'', '', clean_name)  # Remove time markers
                clean_name = clean_name.strip()
                
                # Skip if name is too short after cleaning
                if len(clean_name) < 3:
                    continue
                    
                # Skip if name looks like a UI element
                if any(fp in clean_name.lower() for fp in false_positives):
                    continue
                
                # Add to appropriate list
                if current_section == "home":
                    if not any(are_similar_names(clean_name, existing) for existing in home_players):
                        home_players.append(clean_name)
                else:
                    if not any(are_similar_names(clean_name, existing) for existing in away_players):
                        away_players.append(clean_name)
        
        # Second pass - handle potential single names (but be more selective)
        for i, text in enumerate(text_blocks):
            # Only consider significant text (3+ characters) that isn't recognized elsewhere
            if len(text) >= 3 and text[0].isupper() and not re.match(r'^[\d.]+$', text):
                clean_name = text.strip()
                
                # Skip terms that are likely not player names
                non_name_terms = ["PLAYER", "RATING", "HOME", "AWAY", "BACK", "NEXT", "PREV", "SUB", "CLICK", 
                                 "MENU", "PAGE", "MAIN", "SQUAD", "RANK", "DEFAULT", "ACK", "ACC"]
                if any(term in clean_name.upper() for term in non_name_terms):
                    continue
                
                # Skip position indicators like CB, GK, etc.
                if re.match(r'^[A-Z]{2,3}$', clean_name):
                    continue
                    
                # Skip known false positives
                if any(fp in clean_name.lower() for fp in false_positives):
                    continue
                
                # Add single name if it's not part of an existing entry
                if current_section == "home":
                    # Only add if this name is not part of any existing multi-word names
                    if not any(clean_name in existing or existing in clean_name for existing in home_players):
                        # Check if this might be a last name of a player we know
                        known_last_names = [name.split()[-1] for name in home_players if " " in name]
                        if clean_name not in known_last_names:
                            # Specific name corrections for single names
                            if clean_name.lower() == "awav":
                                continue  # Skip this likely OCR error
                            if clean_name.lower() == "rowe" and any("smith" in player.lower() for player in home_players):
                                continue  # Skip to avoid duplicate with Smith Rowe
                            if clean_name.lower() in ["ack", "maia"]:
                                continue  # Skip common UI misreads
                            
                            home_players.append(clean_name)
                else:
                    if not any(clean_name in existing or existing in clean_name for existing in away_players):
                        known_last_names = [name.split()[-1] for name in away_players if " " in name]
                        if clean_name not in known_last_names:
                            if clean_name.lower() in ["ack", "maia"]:
                                continue  # Skip common UI misreads
                            away_players.append(clean_name)
    
    return home_players, away_players

# Function to check if two names are referring to the same player - STRICT VERSION
def are_similar_names(name1, name2):
    # Convert to comparable format
    norm1 = normalize_player_name(name1)
    norm2 = normalize_player_name(name2)
    
    # Direct match
    if norm1 == norm2:
        return True
    
    # Simple string similarity check for OCR errors
    def string_similarity(s1, s2):
        if not s1 or not s2:
            return 0
        if len(s1) < 3 or len(s2) < 3:
            return 1.0 if s1 == s2 else 0.0
            
        # Count matching characters in order
        matches = 0
        min_len = min(len(s1), len(s2))
        for i in range(min_len):
            if s1[i] == s2[i]:
                matches += 1
        
        # Calculate similarity ratio
        return matches / min_len
    
    # Only allow high similarity whole-name matches (accounting for minor OCR errors)
    if string_similarity(norm1, norm2) >= 0.9:  # High threshold of 90%
        return True
    
    # Check if they're the same via predefined mapping
    # Use our predefined strict matching logic for consistency
    return names_match(norm1, norm2)

# Function to merge similar player names, keeping the most complete versions
def merge_similar_names(player_names):
    if not player_names:
        return []
    
    # Sort by length (longer names first - tend to be more complete)
    sorted_names = sorted(player_names, key=len, reverse=True)
    
    merged = []
    for name in sorted_names:
        # Skip if this name is already covered by an existing entry
        if any(are_similar_names(name, existing) for existing in merged):
            continue
        
        # Create a mapping of common names to their standardized versions
        standard_names = {
            # Format: pattern_to_match_in_lower: standardized_name
            "smith rowe": "Emile Smith Rowe",
            "smith": "Emile Smith Rowe" if any("rowe" in player.lower() for player in player_names) else None,
            "rowe": "Emile Smith Rowe" if any("smith" in player.lower() for player in player_names) else None,
            "emile": "Emile Smith Rowe" if any("smith" in player.lower() or "rowe" in player.lower() for player in player_names) else None,
            "asens": "Marco Asensio",
            "asenc": "Marco Asensio",
            "rudiger": "Antonio Rüdiger",
            "ridiger": "Antonio Rüdiger",
            "ruediger": "Antonio Rüdiger",
            "rlidiger": "Antonio Rüdiger",
            "rudlger": "Antonio Rüdiger",
            "ruoiger": "Antonio Rüdiger",
            "mbappe": "Kylian Mbappé",
            "mbappi": "Kylian Mbappé",
            "embape": "Kylian Mbappé",
            "killian": "Kylian Mbappé" if any("mbappe" in player.lower() for player in player_names) else None,
            "kylian": "Kylian Mbappé" if any("mbappe" in player.lower() for player in player_names) else None,
            "lewandowski": "Robert Lewandowski",
            "lewandovski": "Robert Lewandowski",
            "modric": "Luka Modrić",
            "benzema": "Karim Benzema",
            "benzima": "Karim Benzema",
            "hazard": "Eden Hazard",
            "hazaard": "Eden Hazard",
            "neymar": "Neymar Jr",
            "neyner": "Neymar Jr",
            "salah": "Mohamed Salah",
            "salaa": "Mohamed Salah",
            "mo salah": "Mohamed Salah",
            "mo": "Mohamed Salah" if any("salah" in player.lower() for player in player_names) else None,
            "mohamed": "Mohamed Salah" if any("salah" in player.lower() for player in player_names) else None,
            "ronaldo": "Cristiano Ronaldo",
            "ronaldd": "Cristiano Ronaldo",
            "cristiano": "Cristiano Ronaldo",
            "cr7": "Cristiano Ronaldo",
            "de bruyne": "Kevin De Bruyne",
            "debruvne": "Kevin De Bruyne",
            "debruine": "Kevin De Bruyne",
            "kevin": "Kevin De Bruyne" if any("bruyne" in player.lower() for player in player_names) else None,
            "haaland": "Erling Haaland",
            "halland": "Erling Haaland",
            "haland": "Erling Haaland",
            "erling": "Erling Haaland" if any("haaland" in player.lower() or "haland" in player.lower() for player in player_names) else None,
            "de roon": "Marten de Roon",
            "roon": "Marten de Roon" if any("de" in player.lower() for player in player_names) else None,
            "marten": "Marten de Roon" if any("roon" in player.lower() for player in player_names) else None,
            "tadic": "Dusan Tadić",
            "dusan": "Dusan Tadić" if any("tadic" in player.lower() for player in player_names) else None,
            "lloris": "Hugo Lloris",
            # Only standardize "hugo" to "Hugo Lloris" if "lloris" is also in the player names
            "hugo": "Hugo Lloris" if any("lloris" in player.lower() for player in player_names) else "Hugo",
            "aspas": "Iago Aspas",
            "iago": "Iago Aspas" if any("aspas" in player.lower() for player in player_names) else None
        }
        
        # Check if this name matches any of our standard patterns
        name_lower = name.lower()
        standardized = None
        
        for pattern, standard_name in standard_names.items():
            if pattern in name_lower and standard_name is not None:
                standardized = standard_name
                break
        
        # If we found a standardized version, use it
        if standardized:
            name = standardized
        
        # Add to merged list
        merged.append(name)
    
    return merged

# Find player ratings screens in the extracted frames
def identify_player_ratings_frames(frames, reader):
    print("Analyzing frames to find player ratings screens...")
    home_players = []
    away_players = []
    
    # Track which frames contain home vs away ratings
    home_frames = []
    away_frames = []
    
    for frame_path in frames:
        # Extract text from the frame
        ocr_results = extract_text_from_image(frame_path, reader)
        
        # Check if this frame contains player ratings
        contains_home = False
        contains_away = False
        for detection in ocr_results:
            if "Player Ratings: Home" in detection[1]:
                contains_home = True
            if "Player Ratings: Away" in detection[1]:
                contains_away = True
                
        if contains_home:
            print(f"Found HOME player ratings in {frame_path}")
            home_frames.append(frame_path)
            # Process the text to extract player names
            h_players, _ = process_text_data(ocr_results)
            home_players.extend(h_players)
        
        if contains_away:
            print(f"Found AWAY player ratings in {frame_path}")
            away_frames.append(frame_path)
            # Process the text to extract player names
            _, a_players = process_text_data(ocr_results)
            away_players.extend(a_players)
    
    print(f"Found {len(home_frames)} frames with home ratings and {len(away_frames)} frames with away ratings")
    
    # If we haven't found enough players, reprocess frames with more aggressive settings
    if len(home_players) < 11 or len(away_players) < 11:
        print("Didn't find all 11 players, reprocessing with more aggressive settings...")
        
        # Process all frames that might contain player info, even if headers not detected
        for frame_path in frames:
            ocr_results = extract_text_from_image(frame_path, reader)
            
            # Try to determine if this might be a player rating frame
            contains_player_data = False
            for detection in ocr_results:
                # Look for rating-like values
                if re.match(r'^\d\.\d$', detection[1]):  # Match pattern like "6.5", "7.0", etc.
                    contains_player_data = True
                    break
            
            if contains_player_data:
                print(f"Reprocessing potential player frame: {frame_path}")
                h_temp, a_temp = process_text_data(ocr_results)
                
                # Only add players we don't already have
                for player in h_temp:
                    if not any(are_similar_names(player, existing) for existing in home_players):
                        home_players.append(player)
                for player in a_temp:
                    if not any(are_similar_names(player, existing) for existing in away_players):
                        away_players.append(player)
    
    # Apply smarter merging to get the most complete player names
    cleaned_home_players = merge_similar_names(home_players)
    cleaned_away_players = merge_similar_names(away_players)
    
    # Check if we found close to 11 players
    print(f"Extracted {len(cleaned_home_players)} home players and {len(cleaned_away_players)} away players")
    if len(cleaned_home_players) < 11:
        print("WARNING: Found fewer than 11 home players - some players may be missing")
    if len(cleaned_away_players) < 11:
        print("WARNING: Found fewer than 11 away players - some players may be missing")
    
    return cleaned_home_players, cleaned_away_players

# Normalize player name to handle special characters and variations
def normalize_player_name(name):
    if not name:
        return ""
        
    # Convert to string and uppercase for consistency
    name = str(name).upper()
    
    # Remove excess whitespace
    name = " ".join(name.split())
    
    # Replace common special characters
    name = name.replace("Ü", "U").replace("Ö", "O").replace("Ä", "A")
    name = name.replace("É", "E").replace("È", "E").replace("Ê", "E").replace("Ë", "E")
    name = name.replace("Á", "A").replace("À", "A").replace("Â", "A").replace("Ã", "A")
    name = name.replace("Í", "I").replace("Ì", "I").replace("Î", "I").replace("Ï", "I")
    name = name.replace("Ó", "O").replace("Ò", "O").replace("Ô", "O").replace("Õ", "O")
    name = name.replace("Ú", "U").replace("Ù", "U").replace("Û", "U").replace("Ü", "U")
    name = name.replace("Ñ", "N").replace("Ç", "C").replace("ß", "SS")
    
    # Remove dots from initials to standardize format (e.g., "A. PLAYER" -> "A PLAYER")
    name = re.sub(r'(\b[A-Z])\.', r'\1', name)
    
    # Check for prefixes in player names
    has_prefix = False
    for prefix in ["DE ", "VAN ", "VON ", "EL ", "AL ", "LA ", "DI ", "DA ", "DOS ", "DER ", "TER "]:
        if prefix in name:
            has_prefix = True
    
    # Remove non-alphabetic characters except spaces
    name = re.sub(r'[^A-Z ]', '', name)
    
    # Fix common OCR errors and standardize names
    name = name.replace("II", "I")  # Double I often confused with Ü
    name = name.replace("VV", "W")  # Double V often confused with W
    name = name.replace("RN", "M")  # RN often confused with M
    name = name.replace("0", "O")   # 0 often confused with O
    name = name.replace("IJ", "U")  # IJ is often a misread of Ü
    name = name.replace("1", "I")   # 1 often confused with I
    name = name.replace("8", "B")   # 8 often confused with B
    name = name.replace("5", "S")   # 5 often confused with S
    name = name.replace("6", "G")   # 6 often confused with G
    name = name.replace("9", "G")   # 9 often confused with G
    name = name.replace("KM", "KIM")  # Common OCR error for Asian names
    
    # Common player name corrections
    name = name.replace("RUDIGER", "RUDIGER")  # Standardize
    name = name.replace("RIDIGER", "RUDIGER")  # Standardize 
    name = name.replace("RIIDIGER", "RUDIGER")  # Standardize
    name = name.replace("RUEDICER", "RUDIGER")  # OCR error
    name = name.replace("RLIDIGER", "RUDIGER")  # OCR error
    name = name.replace("RUDLGER", "RUDIGER")  # OCR error
    name = name.replace("RUOIGER", "RUDIGER")  # OCR error
    name = name.replace("TONV", "TONY")  # Fix common error
    name = name.replace("TONYADAMS", "TONY ADAMS")  # Fix missing space
    name = name.replace("SMITHROWE", "SMITH ROWE")  # Fix missing space
    name = name.replace("EMILSMITH", "EMILE SMITH")  # Fix common error
    name = name.replace("SMITE", "SMITH")  # Fix OCR error
    name = name.replace("FHILL", "PHIL")  # Fix OCR error
    name = name.replace("AWAV", "AWAY")  # Common OCR error, not a player name
    name = name.replace("ACKMAIA", "")  # Remove common UI misread
    name = name.replace("MBAPPI", "MBAPPE")  # Fix common OCR error
    name = name.replace("EMBAPE", "MBAPPE")  # Fix common OCR error
    name = name.replace("LEWANDOVSKI", "LEWANDOWSKI")  # Fix common misspelling
    name = name.replace("MODRIC", "MODRIC")  # Fix accent issues
    name = name.replace("BENZIMA", "BENZEMA")  # Fix common misspelling
    name = name.replace("HAZAARD", "HAZARD")  # Fix common OCR error
    name = name.replace("KROOS", "KROOS")  # Standardize
    name = name.replace("NEYNER", "NEYMAR")  # Fix common OCR error
    name = name.replace("KAVEMANI", "CAVANI")  # Fix common OCR error
    name = name.replace("SALAA", "SALAH")  # Fix common OCR error
    name = name.replace("RONALDD", "RONALDO")  # Fix common OCR error
    name = name.replace("DEBRUVNE", "DE BRUYNE")  # Fix common OCR error
    name = name.replace("DEBRUINE", "DE BRUYNE")  # Fix common OCR error
    name = name.replace("HALLAND", "HAALAND")  # Fix common misspelling
    name = name.replace("HALAND", "HAALAND")  # Fix common misspelling
    name = name.replace("ACK", "")  # Remove common UI misread
    name = name.replace("MAIA", "")  # Remove common UI misread
    name = name.replace("IRWF", "")  # Remove this OCR error
    
    # Fix double spacing that might have been introduced
    name = re.sub(r'\s+', ' ', name)
    
    # Just return the normalized name as string (not a tuple)
    return name.strip()

# Get parts of a player name (first, middle, last) with special handling
def get_name_parts(name):
    if not name:
        return [], ""
        
    name_str = str(name).strip().upper()
    parts = name_str.split()
    
    # Handle case with no parts
    if not parts:
        return [], ""
    
    # Handle single-part names
    if len(parts) == 1:
        return [], parts[0]
    
    # For names with prefixes like "de", "van", etc., handle special cases
    prefixes = ["DE", "VAN", "VON", "EL", "AL", "LA", "DI", "DA", "DOS", "DER", "TER"]
    
    # Check if we have a prefix as the first part
    has_prefix = parts[0] in prefixes
    
    # Extract last name with potential prefixes
    if has_prefix and len(parts) >= 3:
        # For "DE JONG" style names, consider "DE JONG" as the last name
        last_name = f"{parts[-2]} {parts[-1]}"
        first_parts = parts[:-2]
    elif has_prefix and len(parts) == 2:
        # For just "DE JONG" with no first name
        last_name = name_str
        first_parts = []
    else:
        # Normal case - last part is the last name
        last_name = parts[-1]
        first_parts = parts[:-1]
    
    # Check for compound last names like "SMITH ROWE"
    compound_last_names = [
        ("SMITH", "ROWE"),
        ("DE", "JONG"),
        ("VAN", "DIJK"),
        ("TER", "STEGEN"),
        ("DE", "BRUYNE"),
        ("DE", "LIGT"),
        ("DE", "GEA"),
        ("DE", "ROON"),
        ("DE", "LA"),
        ("DE", "LOS"),
        ("MC", "TOMINAY"),  # Variations of Mc/Mac names
        ("MAC", "ALLISTER")
    ]
    
    # Check if we have a compound last name
    if len(first_parts) >= 1:
        for prefix, suffix in compound_last_names:
            # Check if we have a match for parts like "SMITH ROWE"
            if (first_parts[-1] == prefix and last_name == suffix):
                last_name = f"{prefix} {suffix}"
                first_parts = first_parts[:-1]
                break
    
    return first_parts, last_name

# Enhanced name matching function - STRICT FULL NAME MATCHING ONLY
def names_match(excel_name, extracted_name):
    # Skip empty names
    if not excel_name or not extracted_name:
        return False
    
    # Normalize both names to handle special characters
    norm_excel = normalize_player_name(excel_name)
    norm_extracted = normalize_player_name(extracted_name)
    
    # Direct match first
    if norm_excel == norm_extracted:
        return True
    
    # Simple string similarity check for OCR errors
    def string_similarity(s1, s2):
        if not s1 or not s2:
            return 0
        if len(s1) < 3 or len(s2) < 3:  # For very short strings, require exact match
            return 1.0 if s1 == s2 else 0.0
            
        # Count matching characters in order
        matches = 0
        min_len = min(len(s1), len(s2))
        for i in range(min_len):
            if s1[i] == s2[i]:
                matches += 1
        
        # Calculate similarity ratio
        return matches / min_len
    
    # Only allow very high similarity whole-name matches (accounting for minor OCR errors)
    if string_similarity(norm_excel, norm_extracted) >= 0.9:  # Increased threshold to 90% 
        return True
    
    # Extract name parts for both
    excel_first_parts, excel_last = get_name_parts(norm_excel)
    extracted_first_parts, extracted_last = get_name_parts(norm_extracted)
    
    # REQUIRE matching last names
    if excel_last != extracted_last:
        # Check for very high similarity in last names (only for OCR errors)
        if string_similarity(excel_last, extracted_last) >= 0.8:
            # Continue checking first names
            pass
        else:
            # Hardcoded known matches for players with different name formats or common OCR errors
            known_matches = [
                # Format: ([excel_pattern_parts], [extracted_pattern_parts])
                (["RUDIGER"], ["RIIDIGER", "RIDIGER", "RUEDIGER", "RLIDIGER", "RUDLGER", "RUOIGER"]),
                (["ASPAS"], ["IAGO ASPAS"]),
                (["IAGO ASPAS"], ["ASPAS"]),
                (["RÜDIGER"], ["ANTONIO RUDIGER"]),
                (["ANTONIO RÜDIGER"], ["RUDIGER"]),
                (["KUBO"], ["TAKEFUSA KUBO"]),
                (["TAKEFUSA KUBO"], ["KUBO"]),
                (["DE ROON"], ["MARTEN DE ROON"]),
                (["MARTEN DE ROON"], ["DE ROON"]),
                (["EMILE SMITH ROWE"], ["SMITH ROWE"]),
                (["SMITH ROWE"], ["EMILE SMITH ROWE"]),
                (["KYLIAN MBAPPÉ"], ["MBAPPE", "MBAPPI", "EMBAPE"]),
                (["MBAPPE"], ["KYLIAN MBAPPE", "KYLIAN MBAPPÉ"]),
                (["ROBERT LEWANDOWSKI"], ["LEWANDOWSKI"]),
                (["LEWANDOWSKI"], ["ROBERT LEWANDOWSKI"]),
                (["LUKA MODRIC"], ["MODRIC"]),
                (["MODRIC"], ["LUKA MODRIC"]),
                (["KARIM BENZEMA"], ["BENZEMA"]),
                (["BENZEMA"], ["KARIM BENZEMA"]),
                (["EDEN HAZARD"], ["HAZARD"]),
                (["HAZARD"], ["EDEN HAZARD"]),
                (["NEYMAR JR"], ["NEYMAR"]),
                (["NEYMAR"], ["NEYMAR JR"]),
                (["MOHAMED SALAH"], ["SALAH", "MO SALAH"]),
                (["SALAH"], ["MOHAMED SALAH", "MO SALAH"]),
                (["CRISTIANO RONALDO"], ["RONALDO"]),
                (["RONALDO"], ["CRISTIANO RONALDO"]),
                (["KEVIN DE BRUYNE"], ["DE BRUYNE"]),
                (["DE BRUYNE"], ["KEVIN DE BRUYNE"]),
                (["ERLING HAALAND"], ["HAALAND"]),
                (["HAALAND"], ["ERLING HAALAND"]),
                # Add specific matches for Hugo vs Hugo Lloris to ensure they're treated as separate players
                (["HUGO LLORIS"], ["LLORIS"]),
                (["LLORIS"], ["HUGO LLORIS"]),
            ]
            
            # Make sure "HUGO" alone doesn't match with "HUGO LLORIS"
            if (norm_excel == "HUGO" and norm_extracted == "HUGO LLORIS") or \
               (norm_excel == "HUGO LLORIS" and norm_extracted == "HUGO"):
                return False
                
            # Check against known full-name matches only
            excel_full = norm_excel
            extracted_full = norm_extracted
            
            for excel_patterns, extracted_patterns in known_matches:
                if excel_full == excel_patterns[0] and extracted_full in extracted_patterns:
                    return True
                if extracted_full == excel_patterns[0] and excel_full in extracted_patterns:
                    return True
            
            return False
    
    # If we get here, last names match - now check first names
    
    # REQUIRE all first parts match or are properly accounted for
    if len(excel_first_parts) != len(extracted_first_parts):
        # Different number of first name parts - only allow if one contains the full set of the other
        if len(excel_first_parts) > len(extracted_first_parts):
            # Excel has more first name parts - check if all extracted parts match excel parts
            for i, extracted_part in enumerate(extracted_first_parts):
                if extracted_part != excel_first_parts[i]:
                    return False
        elif len(extracted_first_parts) > len(excel_first_parts):
            # Extracted has more first name parts - check if all excel parts match extracted parts
            for i, excel_part in enumerate(excel_first_parts):
                if excel_part != extracted_first_parts[i]:
                    return False
        else:
            return False
    else:
        # Same number of first name parts - check all parts match exactly
        for i in range(len(excel_first_parts)):
            if excel_first_parts[i] != extracted_first_parts[i]:
                # Check for initials (e.g., "J." vs "JOHN")
                if len(excel_first_parts[i]) == 1 and len(extracted_first_parts[i]) > 1:
                    if excel_first_parts[i][0] != extracted_first_parts[i][0]:
                        return False
                elif len(extracted_first_parts[i]) == 1 and len(excel_first_parts[i]) > 1:
                    if extracted_first_parts[i][0] != excel_first_parts[i][0]:
                        return False
                else:
                    # Check for high similarity (OCR errors only)
                    if string_similarity(excel_first_parts[i], extracted_first_parts[i]) < 0.8:
                        return False
    
    # If we get here, both last name and all first name parts match or are properly accounted for
    return True

# Update MD value specifically based on match day
def update_md_value(sheet, row, match_day="MD1", col_value=1):
    """Update the MD cell for the specific match day with a value and ensure it's saved properly"""
    # Determine which column to use based on the match day
    col = MD1_COL
    if match_day == "MD2":
        col = MD2_COL
    elif match_day == "MD3":
        col = MD3_COL
    elif match_day == "MD4":
        col = MD4_COL
    elif match_day == "MD5":
        col = MD5_COL
    
    # Get the cell
    md_cell = sheet.cell(row=row, column=col)
    # Force as integer value
    md_cell.value = int(col_value)
    # Format as number to ensure Excel treats it as a proper value
    md_cell.number_format = '0'

# Update a specific team sheet
def update_team_sheet(sheet, player_names, match_day="MD1"):
    # Track which players have been matched
    matched_players = []
    already_matched_excel_rows = []  # Keep track of which Excel rows we've already matched
    
    # Create a copy of player_names to track unmatched players
    unmatched_players = player_names.copy()
    
    # Start checking from row 8 (after header row)
    player_name_start_row = PLAYER_NAME_ROW_START + 1
    
    # Find existing players and mark them as present
    for row in range(player_name_start_row, ADDITIONAL_PLAYERS_ROW):
        cell_value = sheet.cell(row=row, column=PLAYER_NAME_COL).value
        if cell_value and row not in already_matched_excel_rows:
            # Try to match against extracted players
            for extracted_name in player_names:
                if names_match(cell_value, extracted_name):
                    # Set the value 1 in the appropriate MD column
                    update_md_value(sheet, row, match_day)
                    print(f"Matched player: {cell_value} with {extracted_name} for {match_day}")
                    
                    # Only add to matched_players if this is the first time we're matching this exact name
                    if extracted_name not in matched_players:
                        matched_players.append(extracted_name)
                        # Only remove from unmatched once per extracted name
                        if extracted_name in unmatched_players:
                            unmatched_players.remove(extracted_name)
                    
                    # Mark this Excel row as matched
                    already_matched_excel_rows.append(row)
                    break
    
    # Clean up unmatched player names to prevent duplicates due to OCR errors
    clean_unmatched = []
    for player in unmatched_players:
        # Check if this is just an OCR error of an already matched player
        if not any(names_match(matched, player) for matched in matched_players):
            clean_unmatched.append(player)
            
    unmatched_players = clean_unmatched
    
    # Add unmatched players to A60 area
    if unmatched_players:
        print(f"Adding {len(unmatched_players)} unmatched players starting at row {ADDITIONAL_PLAYERS_ROW}")
        
        # Add header in the additional players area if it doesn't exist
        if not sheet.cell(row=ADDITIONAL_PLAYERS_ROW, column=PLAYER_NAME_COL).value:
            sheet.cell(row=ADDITIONAL_PLAYERS_ROW, column=PLAYER_NAME_COL).value = "ADDITIONAL PLAYERS"
            sheet.cell(row=ADDITIONAL_PLAYERS_ROW, column=PLAYER_NAME_COL + 1).value = match_day
        
        # Add unmatched players below the header
        current_row = ADDITIONAL_PLAYERS_ROW + 1
        for player in unmatched_players:
            sheet.cell(row=current_row, column=PLAYER_NAME_COL).value = player
            # Use our helper function to ensure consistent formatting for the correct match day
            update_md_value(sheet, current_row, match_day)
            print(f"Added unmatched player: {player} for {match_day}")
            current_row += 1
    
    # Return the counts of matched and unmatched players
    return len(matched_players), len(unmatched_players)

# Update Excel file with the extracted player data
def update_excel(excel_path, home_team, away_team, home_players, away_players, match_day):
    print(f"Updating Excel file: {excel_path} for {match_day}")
    
    try:
        # Load workbook with proper error handling
        try:
            workbook = load_workbook(excel_path)
        except Exception as e:
            print(f"Error opening Excel file: {str(e)}")
            print("This may be due to a corrupted file, or the file being open in Excel.")
            print("Please close Excel if it's open and try again.")
            return False
        
        # Track counts for summary
        home_extracted = len(home_players)
        home_matched = 0
        home_unmatched = 0
        away_extracted = len(away_players)
        away_matched = 0 
        away_unmatched = 0
        
        # Process home team
        if home_team in workbook.sheetnames:
            print(f"\nUpdating {home_team} sheet for {match_day}...")
            home_matched, home_unmatched = update_team_sheet(workbook[home_team], home_players, match_day)
            print(f"SUMMARY for {home_team}:")
            print(f"  - Players Extracted: {home_extracted}")
            print(f"  - Players Matched: {home_matched} (may include players matched to multiple Excel entries)")
            print(f"  - Players Unmatched: {home_unmatched} (added as new entries)")
        else:
            print(f"Warning: Sheet for {home_team} not found in Excel file")
            print(f"Available sheets: {', '.join(workbook.sheetnames)}")
            
            # Ask user if they want to use an alternate sheet
            print(f"Would you like to use an alternate sheet for {home_team}?")
            for i, sheet_name in enumerate(workbook.sheetnames, 1):
                print(f"{i}. {sheet_name}")
            
            try:
                choice = input("Enter sheet number or 0 to skip: ")
                if choice.isdigit() and 1 <= int(choice) <= len(workbook.sheetnames):
                    selected_sheet = workbook.sheetnames[int(choice) - 1]
                    print(f"Using sheet {selected_sheet} for {home_team}")
                    home_matched, home_unmatched = update_team_sheet(workbook[selected_sheet], home_players, match_day)
            except:
                print(f"Skipping {home_team} updates")
        
        # Process away team
        if away_team in workbook.sheetnames:
            print(f"\nUpdating {away_team} sheet for {match_day}...")
            away_matched, away_unmatched = update_team_sheet(workbook[away_team], away_players, match_day)
            print(f"SUMMARY for {away_team}:")
            print(f"  - Players Extracted: {away_extracted}")
            print(f"  - Players Matched: {away_matched} (may include players matched to multiple Excel entries)")
            print(f"  - Players Unmatched: {away_unmatched} (added as new entries)")
        else:
            print(f"Warning: Sheet for {away_team} not found in Excel file")
            print(f"Available sheets: {', '.join(workbook.sheetnames)}")
            
            # Ask user if they want to use an alternate sheet
            print(f"Would you like to use an alternate sheet for {away_team}?")
            for i, sheet_name in enumerate(workbook.sheetnames, 1):
                print(f"{i}. {sheet_name}")
            
            try:
                choice = input("Enter sheet number or 0 to skip: ")
                if choice.isdigit() and 1 <= int(choice) <= len(workbook.sheetnames):
                    selected_sheet = workbook.sheetnames[int(choice) - 1]
                    print(f"Using sheet {selected_sheet} for {away_team}")
                    away_matched, away_unmatched = update_team_sheet(workbook[selected_sheet], away_players, match_day)
            except:
                print(f"Skipping {away_team} updates")
        
        # Print overall summary
        total_extracted = home_extracted + away_extracted
        total_matched = home_matched + away_matched
        total_unmatched = home_unmatched + away_unmatched
        print(f"\nOVERALL SUMMARY:")
        print(f"  - Total Extracted: {total_extracted} ({home_extracted} home + {away_extracted} away)")
        print(f"  - Total Matched: {total_matched} ({home_matched} home + {away_matched} away)")
        print(f"  - Total Unmatched: {total_unmatched} ({home_unmatched} home + {away_unmatched} away)")
        
        # Save workbook with a try-except to handle file permission issues
        try:
            workbook.save(excel_path)
            print(f"Excel file updated successfully for {match_day}")
            return True
        except PermissionError:
            # If the original file is open, save to a backup file
            backup_path = f"{excel_path.split('.')[0]}_updated.xlsx"
            workbook.save(backup_path)
            print(f"Warning: Could not save to {excel_path} (file may be open in Excel)")
            print(f"Data saved to backup file: {backup_path}")
            print(f"Please close the original Excel file and rename the backup file to {excel_path}")
            return True
    except Exception as e:
        print(f"Error updating Excel file: {str(e)}")
        print(traceback.format_exc())  # Print full traceback for debugging
        
        # Try saving to a backup file if there's an issue
        try:
            backup_path = f"{excel_path.split('.')[0]}_error_backup.xlsx"
            workbook.save(backup_path)
            print(f"Data saved to backup file: {backup_path}")
        except Exception as backup_error:
            print(f"Failed to save backup file: {str(backup_error)}")
            print("Please ensure Excel is not open and the file is not read-only.")
        return False

# Process a single video file
def process_video(video_path):
    start_time = time.time()
    print(f"\n{'='*80}")
    print(f"Starting player extraction from video: {video_path}")
    print(f"{'='*80}\n")
    
    # Extract team names from filename
    home_team, away_team = extract_teams_from_filename(video_path)
    
    # Extract match day from video filename
    match_day_value = default_match_day
    if "md" in video_path.lower():
        # Try to extract the match day number from the filename
        md_match = re.search(r'md(\d+)', video_path.lower())
        if md_match:
            md_number = md_match.group(1)
            match_day_value = f"MD{md_number}"
            print(f"Detected match day from filename: {match_day_value}")
    
    # Initialize OCR
    reader = initialize_ocr()
    
    # Extract frames
    frames = extract_frames(video_path)
    
    if not frames:
        print("Error: No frames were extracted from the video.")
        return False
    
    # Identify player ratings frames and extract player names
    home_players, away_players = identify_player_ratings_frames(frames, reader)
    
    # Debug: Print raw OCR results for a few frames to identify potential issues
    print("\nDEBUG: Checking for OCR errors in 3 random frames...")
    import random
    sample_frames = random.sample(frames, min(3, len(frames)))
    for frame_path in sample_frames:
        print(f"\nOCR results for {frame_path}:")
        ocr_results = extract_text_from_image(frame_path, reader)
        for detection in ocr_results:
            text = detection[1]
            confidence = detection[2]
            if confidence > 0.2:
                print(f"Text: {text}, Confidence: {confidence:.2f}")
                if "irwf" in text.lower():
                    print(f"WARNING: Found 'irwf' in '{text}' - this is likely an OCR error!")
    
    # Print extracted player lists with counts
    print(f"\nExtracted Home Players ({home_team}) - Total: {len(home_players)}")
    for i, player in enumerate(home_players, 1):
        print(f"{i}. {player}")
    
    print(f"\nExtracted Away Players ({away_team}) - Total: {len(away_players)}")
    for i, player in enumerate(away_players, 1):
        print(f"{i}. {player}")
    
    # Update Excel file
    update_excel(excel_path, home_team, away_team, home_players, away_players, match_day_value)
    
    # Clean up frames
    if os.path.exists("frames"):
        import shutil
        shutil.rmtree("frames")
    
    elapsed_time = time.time() - start_time
    print(f"\nProcess completed in {elapsed_time:.2f} seconds.")
    return True

def main():
    # Process these specific videos
    video_files = ["NEWCASTLE UNITED VS BORUSSIA DORTMUND MD1.mp4"]
    
    # First, verify the Excel file
    global excel_path
    valid_excel_path = verify_excel_file(excel_path)
    if not valid_excel_path:
        print("Error: Unable to proceed without a valid Excel file.")
        return
    excel_path = valid_excel_path
    
    # Check if the files exist
    existing_files = []
    for video in video_files:
        if os.path.exists(video):
            existing_files.append(video)
        else:
            print(f"Warning: Video file not found: {video}")
    
    if not existing_files:
        print("Error: No video files found to process.")
        return
    
    # Process each video file
    success_count = 0
    total_results = {
        "videos_processed": 0,
        "total_players_extracted": 0,
        "total_players_matched": 0,
        "total_players_unmatched": 0,
        "details": []
    }
    
    for video_file in existing_files:
        start_time = time.time()
        print(f"\n{'='*80}")
        print(f"Starting player extraction from video: {video_file}")
        print(f"{'='*80}\n")
        
        try:
            # Extract team names from filename
            home_team, away_team = extract_teams_from_filename(video_file)
            
            # Extract match day from video filename
            match_day_value = default_match_day
            if "md" in video_file.lower():
                # Try to extract the match day number from the filename
                md_match = re.search(r'md(\d+)', video_file.lower())
                if md_match:
                    md_number = md_match.group(1)
                    match_day_value = f"MD{md_number}"
                    print(f"Detected match day from filename: {match_day_value}")
            
            # Initialize OCR
            reader = initialize_ocr()
            
            # Extract frames
            frames = extract_frames(video_file)
            
            if not frames:
                print("Error: No frames were extracted from the video.")
                continue
            
            # Identify player ratings frames and extract player names
            home_players, away_players = identify_player_ratings_frames(frames, reader)
            
            # Debug: Check for common OCR errors in a sample of frames
            print("\nDEBUG: Checking for OCR errors in 3 random frames...")
            import random
            sample_frames = random.sample(frames, min(3, len(frames)))
            for frame_path in sample_frames:
                print(f"\nOCR results for {frame_path}:")
                ocr_results = extract_text_from_image(frame_path, reader)
                for detection in ocr_results:
                    text = detection[1]
                    confidence = detection[2]
                    if confidence > 0.2:
                        print(f"Text: {text}, Confidence: {confidence:.2f}")
                        if "irwf" in text.lower():
                            print(f"WARNING: Found 'irwf' in '{text}' - this is likely an OCR error!")
            
            # Print extracted player lists with counts
            print(f"\nExtracted Home Players ({home_team}) - Total: {len(home_players)}")
            for i, player in enumerate(home_players, 1):
                print(f"{i}. {player}")
            
            print(f"\nExtracted Away Players ({away_team}) - Total: {len(away_players)}")
            for i, player in enumerate(away_players, 1):
                print(f"{i}. {player}")
            
            # Update Excel file and get counts
            try:
                # Load workbook
                workbook = load_workbook(excel_path)
                
                # Track counts for this video
                video_results = {
                    "video_file": video_file,
                    "home_team": home_team,
                    "away_team": away_team,
                    "match_day": match_day_value,
                    "home_players_extracted": len(home_players),
                    "away_players_extracted": len(away_players),
                    "home_players_matched": 0,
                    "home_players_unmatched": 0,
                    "away_players_matched": 0,
                    "away_players_unmatched": 0
                }
                
                # Process home team
                if home_team in workbook.sheetnames:
                    print(f"\nUpdating {home_team} sheet for {match_day_value}...")
                    matched_count, unmatched_count = update_team_sheet(workbook[home_team], home_players, match_day_value)
                    video_results["home_players_matched"] = matched_count
                    video_results["home_players_unmatched"] = unmatched_count
                    print(f"SUMMARY for {home_team}: {matched_count} players matched, {unmatched_count} players added as unmatched")
                else:
                    print(f"Warning: Sheet for {home_team} not found in Excel file")
                
                # Process away team
                if away_team in workbook.sheetnames:
                    print(f"\nUpdating {away_team} sheet for {match_day_value}...")
                    matched_count, unmatched_count = update_team_sheet(workbook[away_team], away_players, match_day_value)
                    video_results["away_players_matched"] = matched_count
                    video_results["away_players_unmatched"] = unmatched_count
                    print(f"SUMMARY for {away_team}: {matched_count} players matched, {unmatched_count} players added as unmatched")
                else:
                    print(f"Warning: Sheet for {away_team} not found in Excel file")
                
                # Update total counts
                total_players_matched = video_results["home_players_matched"] + video_results["away_players_matched"]
                total_players_unmatched = video_results["home_players_unmatched"] + video_results["away_players_unmatched"]
                
                print(f"\nVIDEO SUMMARY FOR {video_file}:")
                print(f"{'='*50}")
                print(f"Home Team ({home_team}):")
                print(f"  - Players Extracted: {video_results['home_players_extracted']}")
                print(f"  - Players Matched: {video_results['home_players_matched']} (may match with multiple Excel entries)")
                print(f"  - Players Unmatched (Added): {video_results['home_players_unmatched']}")
                print(f"Away Team ({away_team}):")
                print(f"  - Players Extracted: {video_results['away_players_extracted']}")
                print(f"  - Players Matched: {video_results['away_players_matched']} (may match with multiple Excel entries)")
                print(f"  - Players Unmatched (Added): {video_results['away_players_unmatched']}")
                print(f"TOTAL: {video_results['home_players_extracted'] + video_results['away_players_extracted']} extracted, {total_players_matched} matched, {total_players_unmatched} unmatched")
                print(f"NOTE: A single player may match with multiple entries in the Excel file (e.g., Michael Owen matching both 'M. OWEN' and 'J. BOWEN')")
                print(f"{'='*50}")
                
                # Save workbook with a try-except to handle file permission issues
                try:
                    workbook.save(excel_path)
                    print(f"Excel file updated successfully for {match_day_value}")
                    success_count += 1
                except PermissionError:
                    # If the original file is open, save to a backup file
                    backup_path = f"{excel_path.split('.')[0]}_updated.xlsx"
                    workbook.save(backup_path)
                    print(f"Warning: Could not save to {excel_path} (file may be open in Excel)")
                    print(f"Data saved to backup file: {backup_path}")
                    print(f"Please close the original Excel file and rename the backup file to {excel_path}")
                    success_count += 1
                
                # Add to total results
                total_results["videos_processed"] += 1
                total_results["total_players_extracted"] += (video_results["home_players_extracted"] + video_results["away_players_extracted"])
                total_results["total_players_matched"] += total_players_matched
                total_results["total_players_unmatched"] += total_players_unmatched
                total_results["details"].append(video_results)
                
            except Exception as e:
                print(f"Error updating Excel file: {str(e)}")
                print(traceback.format_exc())  # Print full traceback for debugging
                
                # Try saving to a backup file if there's an issue
                try:
                    backup_path = f"{excel_path.split('.')[0]}_error_backup.xlsx"
                    workbook.save(backup_path)
                    print(f"Data saved to backup file: {backup_path}")
                except Exception as backup_error:
                    print(f"Failed to save backup file: {str(backup_error)}")
                    print("Please ensure Excel is not open and the file is not read-only.")
            
            # Clean up frames
            if os.path.exists("frames"):
                import shutil
                shutil.rmtree("frames")
            
            elapsed_time = time.time() - start_time
            print(f"\nVideo {video_file} processed in {elapsed_time:.2f} seconds.")
            
        except Exception as e:
            print(f"Error processing video {video_file}: {str(e)}")
            print(traceback.format_exc())  # Print full traceback for debugging
    
    # Print overall summary for all videos
    print(f"\n{'#'*80}")
    print(f"OVERALL SUMMARY FOR ALL VIDEOS")
    print(f"{'#'*80}")
    print(f"Videos Processed: {total_results['videos_processed']} of {len(existing_files)}")
    print(f"Total Players Extracted: {total_results['total_players_extracted']}")
    print(f"Total Players Matched: {total_results['total_players_matched']}")
    print(f"Total Players Unmatched (Added): {total_results['total_players_unmatched']}")
    
    # Detail summary for each video
    print(f"\nDETAILED RESULTS BY VIDEO:")
    for i, result in enumerate(total_results["details"], 1):
        print(f"\n{i}. {result['video_file']} ({result['match_day']}):")
        print(f"   - Home Team ({result['home_team']}): {result['home_players_extracted']} extracted, {result['home_players_matched']} matched, {result['home_players_unmatched']} added")
        print(f"   - Away Team ({result['away_team']}): {result['away_players_extracted']} extracted, {result['away_players_matched']} matched, {result['away_players_unmatched']} added")
    
    print(f"\nComplete! {success_count} out of {len(existing_files)} videos processed successfully.")

if __name__ == "__main__":
    main()