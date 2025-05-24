import os
import time
import tempfile
import json
import uuid
import threading
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, send_file
from flask_session import Session
from werkzeug.utils import secure_filename
from dotenv import load_dotenv
from supabase import create_client, Client
import file_manager
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

# Load environment variables
load_dotenv()

# Initialize Supabase client
supabase_url = os.environ.get("SUPABASE_URL")
supabase_key = os.environ.get("SUPABASE_KEY")
supabase: Client = create_client(supabase_url, supabase_key)

# Flask app configuration
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev_key_for_testing")

# Configure server-side sessions
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_FILE_DIR'] = os.path.join(os.getcwd(), 'flask_sessions')
app.config['SESSION_PERMANENT'] = False
app.config['SESSION_USE_SIGNER'] = True
Session(app)

# Configure app folders
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['FRAMES_FOLDER'] = os.path.join('static', 'frames')

# Make sure folders exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['FRAMES_FOLDER'], exist_ok=True)
os.makedirs(app.config['SESSION_FILE_DIR'], exist_ok=True)

# Global variables to track tasks
processing_tasks = {}
task_lock = threading.Lock()

# Helper functions
def get_players_by_team_id(team_id):
    """Get all players belonging to a specific team"""
    result = supabase.table("players").select("*").eq("team_id", team_id).execute()
    return result.data if result.data else []

def update_player_appearances(player_id, match_id):
    """Add or update player appearance record"""
    # Check if this appearance already exists
    result = supabase.table("appearances").select("*").eq("player_id", player_id).eq("match_id", match_id).execute()
    
    if result.data and len(result.data) > 0:
        # Already exists, don't create duplicate
        return False
        
    # Create new appearance record
    result = supabase.table("appearances").insert({
        "player_id": player_id,
        "match_id": match_id
    }).execute()
    
    if result.data and len(result.data) > 0:
        # Update total_appearances count in players table
        player = supabase.table("players").select("total_appearances").eq("id", player_id).execute()
        if player.data and len(player.data) > 0:
            current_count = player.data[0].get("total_appearances", 0) or 0
            supabase.table("players").update({"total_appearances": current_count + 1}).eq("id", player_id).execute()
        return True
    
    return False

def store_unmatched_player(name, team_id, match_id):
    """Store unmatched player names in the database for future reference"""
    try:
        # Check if this unmatched player + team combination already exists
        result = supabase.table("unmatched_players").select("*").eq("name", name).eq("team_id", team_id).execute()
        
        if result.data and len(result.data) > 0:
            # If it exists, update the last_seen field and increment occurrence count
            player_id = result.data[0]["id"]
            occurrence_count = result.data[0].get("occurrence_count", 1) + 1
            
            supabase.table("unmatched_players").update({
                "last_seen": time.strftime("%Y-%m-%d"),
                "occurrence_count": occurrence_count,
                "last_match_id": match_id
            }).eq("id", player_id).execute()
            
            print(f"Updated unmatched player: {name} for team {team_id} (seen {occurrence_count} times)")
            return player_id
        else:
            # Create new unmatched player record
            result = supabase.table("unmatched_players").insert({
                "name": name,
                "team_id": team_id,
                "first_seen": time.strftime("%Y-%m-%d"),
                "last_seen": time.strftime("%Y-%m-%d"),
                "occurrence_count": 1,
                "last_match_id": match_id
            }).execute()
            
            if result.data and len(result.data) > 0:
                print(f"Added new unmatched player: {name} for team {team_id}")
                return result.data[0]["id"]
            
    except Exception as e:
        print(f"Error storing unmatched player {name}: {str(e)}")
    
    return None

# Routes
@app.route('/')
def index():
    """Render the home page with stats"""
    # Get all teams for dropdown
    teams = supabase.table("teams").select("*").order("name").execute()
    teams_data = teams.data if teams.data else []
    
    # Create stats with default values
    default_stats = {
        'players': '150+',
        'teams': '12',
        'matches': '48',
        'appearances': '1200+'
    }
    
    # Try to get actual stats if possible
    try:
        # Count players - just get all and count length
        players_data = supabase.table("players").select("id").execute()
        if players_data.data:
            default_stats['players'] = str(len(players_data.data))
        
        # Count teams from our already fetched data
        if teams_data:
            default_stats['teams'] = str(len(teams_data))
        
        # Count matches
        matches_data = supabase.table("matches").select("id").execute()
        if matches_data.data:
            default_stats['matches'] = str(len(matches_data.data))
        
        # Count appearances
        appearances_data = supabase.table("appearances").select("id").execute()
        if appearances_data.data:
            default_stats['appearances'] = str(len(appearances_data.data))
    
    except Exception as e:
        print(f"Error getting stats: {str(e)}")
        # We'll use the default stats defined above
    
    return render_template('index.html', teams=teams_data, stats=default_stats)

@app.route('/form_demo')
def form_demo():
    """Render the form demo page with Vision OS styled elements"""
    return render_template('form_demo.html')

@app.route('/players')
def players():
    """View all players, optionally filtered by team"""
    team_id = request.args.get('team_id')
    if not team_id:
        players_data = supabase.table("players").select("*, team:team_id(name)").order("name").execute()
    else:
        players_data = supabase.table("players").select("*, team:team_id(name)").eq("team_id", team_id).order("name").execute()
    
    teams = supabase.table("teams").select("*").order("name").execute()
    
    return render_template(
        'players.html', 
        players=players_data.data if players_data.data else [],
        teams=teams.data if teams.data else []
    )

@app.route('/matches')
def matches():
    """View all matches"""
    # Get all matches with team information
    matches_data = supabase.table("matches").select("*, home_team:home_team_id(name), away_team:away_team_id(name)").order("date", desc=True).execute()
    
    # Get all teams for the create match day form
    all_teams = supabase.table("teams").select("id, name").order("name").execute()
    
    # If there are matches, get additional information for each
    if matches_data.data:
        for match in matches_data.data:
            # Get appearance counts
            try:
                appearances = supabase.table("appearances").select("*, player:player_id(team_id)").eq("match_id", match["id"]).execute()
                
                # Count by team
                home_count = 0
                away_count = 0
                
                if appearances.data:
                    for appearance in appearances.data:
                        if 'player' in appearance and appearance['player']:
                            team_id = appearance['player'].get('team_id')
                            if team_id == match['home_team_id']:
                                home_count += 1
                            elif team_id == match['away_team_id']:
                                away_count += 1
                
                match["home_appearances"] = home_count
                match["away_appearances"] = away_count
                
                # Get unmatched player counts
                unmatched = supabase.table("unmatched_players").select("team_id").eq("last_match_id", match["id"]).execute()
                
                # Count by team
                home_unmatched = 0
                away_unmatched = 0
                
                if unmatched.data:
                    for player in unmatched.data:
                        team_id = player.get('team_id')
                        if team_id == match['home_team_id']:
                            home_unmatched += 1
                        elif team_id == match['away_team_id']:
                            away_unmatched += 1
                
                match["home_unmatched"] = home_unmatched
                match["away_unmatched"] = away_unmatched
                
            except Exception as e:
                print(f"Error getting stats for match {match['id']}: {str(e)}")
                match["home_appearances"] = 0
                match["away_appearances"] = 0
                match["home_unmatched"] = 0
                match["away_unmatched"] = 0
    
    return render_template('matches.html', 
                           matches=matches_data.data if matches_data.data else [],
                           all_teams=all_teams.data if all_teams.data else [])

@app.route('/match/<match_id>')
def match_details(match_id):
    """View details about a specific match"""
    # Get match info
    match = supabase.table("matches").select("*, home_team:home_team_id(name), away_team:away_team_id(name)").eq("id", match_id).execute()
    
    if not match.data or len(match.data) == 0:
        flash('Match not found', 'danger')
        return redirect(url_for('matches'))
    
    # Get appearances for this match
    appearances = supabase.table("appearances").select("*, player:player_id(name, team_id)").eq("match_id", match_id).execute()
    
    # Get unmatched players for this match
    unmatched_players = supabase.table("unmatched_players").select("*").eq("last_match_id", match_id).execute()
    
    # Separate by team
    home_appearances = []
    away_appearances = []
    home_unmatched = []
    away_unmatched = []
    
    # Process regular appearances
    if appearances.data:
        for appearance in appearances.data:
            if 'player' in appearance and appearance['player']:
                team_id = appearance['player'].get('team_id')
                if team_id == match.data[0]['home_team_id']:
                    home_appearances.append(appearance)
                elif team_id == match.data[0]['away_team_id']:
                    away_appearances.append(appearance)
    
    # Process unmatched players
    if unmatched_players.data:
        for player in unmatched_players.data:
            team_id = player.get('team_id')
            if team_id == match.data[0]['home_team_id']:
                home_unmatched.append(player)
            elif team_id == match.data[0]['away_team_id']:
                away_unmatched.append(player)
    
    return render_template(
        'match_details.html', 
        match=match.data[0], 
        home_appearances=home_appearances,
        away_appearances=away_appearances,
        home_unmatched=home_unmatched,
        away_unmatched=away_unmatched
    )

@app.route('/match/<match_id>/edit')
def edit_match_appearances(match_id):
    """Edit player appearances for a specific match"""
    # Get match data
    match = supabase.table("matches").select("*, home_team:home_team_id(name), away_team:away_team_id(name)").eq("id", match_id).execute()
    
    if not match.data or len(match.data) == 0:
        flash('Match not found', 'danger')
        return redirect(url_for('matches'))
    
    # Get all players for both teams
    home_team_id = match.data[0]['home_team_id']
    away_team_id = match.data[0]['away_team_id']
    
    home_players = supabase.table("players").select("*").eq("team_id", home_team_id).order("name").execute()
    away_players = supabase.table("players").select("*").eq("team_id", away_team_id).order("name").execute()
    
    # Get current appearances for this match
    appearances = supabase.table("appearances").select("*").eq("match_id", match_id).execute()
    
    # Create a set of player IDs who appeared in this match for quick lookup
    appeared_player_ids = set()
    if appearances.data:
        for appearance in appearances.data:
            appeared_player_ids.add(appearance['player_id'])
    
    # Get unmatched players for this match
    unmatched_players_result = supabase.table("unmatched_players").select("*").eq("last_match_id", match_id).execute()
    
    # Format unmatched players data for the template
    unmatched_players = []
    if unmatched_players_result.data and len(unmatched_players_result.data) > 0:
        for player in unmatched_players_result.data:
            team = "home" if player["team_id"] == home_team_id else "away"
            unmatched_players.append({
                "id": player["id"],
                "name": player["name"],
                "team": team,
                "occurrence_count": player.get("occurrence_count", 1)
            })
    
    # Format team data
    home_team_name = match.data[0]['home_team']['name'] if 'home_team' in match.data[0] else "Unknown Team"
    away_team_name = match.data[0]['away_team']['name'] if 'away_team' in match.data[0] else "Unknown Team"
    match_day = match.data[0].get('match_day', 'Unknown')
    match_date = match.data[0].get('date', 'Unknown')
    
    return render_template(
        'edit_match_appearances.html',
        match_id=match_id,
        match_day_id=match_id,  # Used by unmatched players forms
        home_team_name=home_team_name,
        away_team_name=away_team_name,
        match_day=match_day,
        match_date=match_date,
        home_players=home_players.data if home_players.data else [],
        away_players=away_players.data if away_players.data else [],
        appeared_player_ids=appeared_player_ids,
        unmatched_players=unmatched_players
    )

@app.route('/match/<match_id>/update_appearances', methods=['POST'])
def update_match_appearances(match_id):
    """Update player appearances for a match"""
    # Get match data to verify it exists
    match = supabase.table("matches").select("id").eq("id", match_id).execute()
    
    if not match.data or len(match.data) == 0:
        return jsonify({"success": False, "error": "Match not found"})
    
    # Get player IDs from form
    player_data = request.json
    
    if not player_data or not isinstance(player_data, dict):
        return jsonify({"success": False, "error": "Invalid player data"})
    
    selected_players = player_data.get('player_ids', [])
    
    try:
        # First, get all current appearances for this match
        current_appearances = supabase.table("appearances").select("id, player_id").eq("match_id", match_id).execute()
        
        # Create sets for easier comparison
        current_player_ids = set()
        if current_appearances.data:
            for appearance in current_appearances.data:
                current_player_ids.add(appearance['player_id'])
        
        selected_player_ids = set(selected_players)
        
        # Players to add (in selected but not in current)
        players_to_add = selected_player_ids - current_player_ids
        
        # Players to remove (in current but not in selected)
        players_to_remove = current_player_ids - selected_player_ids
        
        # Add new appearances
        for player_id in players_to_add:
            supabase.table("appearances").insert({
                "player_id": player_id,
                "match_id": match_id
            }).execute()
            
            # Update total_appearances count in players table
            player = supabase.table("players").select("total_appearances").eq("id", player_id).execute()
            if player.data and len(player.data) > 0:
                current_count = player.data[0].get("total_appearances", 0) or 0
                supabase.table("players").update({"total_appearances": current_count + 1}).eq("id", player_id).execute()
        
        # Remove appearances that are no longer selected
        for player_id in players_to_remove:
            # Find the appearance ID to delete
            appearance_to_remove = next((a for a in current_appearances.data if a['player_id'] == player_id), None)
            
            if appearance_to_remove:
                # Delete the appearance
                supabase.table("appearances").delete().eq("id", appearance_to_remove['id']).execute()
                
                # Update total_appearances count in players table
                player = supabase.table("players").select("total_appearances").eq("id", player_id).execute()
                if player.data and len(player.data) > 0:
                    current_count = player.data[0].get("total_appearances", 0) or 0
                    if current_count > 0:  # Ensure we don't go negative
                        supabase.table("players").update({"total_appearances": current_count - 1}).eq("id", player_id).execute()
    
        return jsonify({
            "success": True,
            "added": len(players_to_add),
            "removed": len(players_to_remove)
        })
        
    except Exception as e:
        print(f"Error updating appearances: {str(e)}")
        return jsonify({"success": False, "error": str(e)}) 

@app.route('/create_match_day', methods=['POST'])
def create_match_day():
    """Create a new match day"""
    try:
        # Get data from request
        data = request.json if request.is_json else request.form.to_dict()
        match_day = data.get('match_day')
        match_date = data.get('match_date')
        home_team_id = data.get('home_team_id')
        away_team_id = data.get('away_team_id')
        
        # Set default date to today if not provided
        if not match_date:
            from datetime import date
            match_date = date.today().isoformat()
        
        # Validate input
        if not match_day or not home_team_id or not away_team_id:
            flash("Match day and team selection are required", "danger")
            return redirect(url_for('matches'))
            
        # Insert the new match day
        result = supabase.table("matches").insert({
            "match_day": match_day,
            "date": match_date,
            "home_team_id": home_team_id,
            "away_team_id": away_team_id
        }).execute()
        
        if not result.data or len(result.data) == 0:
            flash("Failed to create match day", "danger")
            return redirect(url_for('matches'))
            
        # Redirect to edit appearances page for the new match
        match_id = result.data[0]["id"]
        flash("Match day created successfully", "success")
        return redirect(url_for('edit_match_appearances', match_id=match_id))
        
    except Exception as e:
        print(f"Error creating match day: {str(e)}")
        flash(f"Error: {str(e)}", "danger")
        return redirect(url_for('matches'))

@app.route('/add_unmatched_player/<match_day_id>', methods=['POST'])
def add_unmatched_player(match_day_id):
    """Add an unmatched player for a specific match day"""
    try:
        # Get data from request
        data = request.json if request.is_json else request.form.to_dict()
        name = data.get('name')
        team = data.get('team')
        first_seen = data.get('first_seen')
        last_seen = data.get('last_seen')
        last_match_id = data.get('last_match_id', match_day_id)  # Default to match_day_id if not provided
        source_player_id = data.get('source_player_id')
        from_different_team = data.get('from_different_team', False)
        
        # Validate input
        if not name or not team:
            return jsonify({"success": False, "error": "Name and team are required"})
        
        # Get match details
        match_result = supabase.table("matches").select("*").eq("id", match_day_id).execute()
        
        if not match_result.data or len(match_result.data) == 0:
            return jsonify({"success": False, "error": "Match not found"})
            
        match = match_result.data[0]
        team_id = match["home_team_id"] if team == "home" else match["away_team_id"]
        
        # Check if player already exists with this name and team
        existing_player = supabase.table("unmatched_players").select("*").eq("name", name).eq("team_id", team_id).execute()
        
        if existing_player.data and len(existing_player.data) > 0:
            # Increment existing player
            player = existing_player.data[0]
            player_id = player["id"]
            new_occurrence_count = (player.get("occurrence_count") or 1) + 1
            
            result = supabase.table("unmatched_players").update({
                "occurrence_count": new_occurrence_count,
                "last_seen": time.strftime("%Y-%m-%d"),
                "last_match_id": last_match_id
            }).eq("id", player_id).execute()
            
            return jsonify({
                "success": True, 
                "player_id": player_id,
                "occurrence_count": new_occurrence_count,
                "is_existing": True,
                "message": "Player already exists, count incremented"
            })
        
        # Insert the unmatched player
        insert_data = {
            "name": name,
            "team_id": team_id,
            "last_match_id": last_match_id,
            "first_seen": first_seen or time.strftime("%Y-%m-%d"),
            "last_seen": last_seen or time.strftime("%Y-%m-%d"),
            "status": "unmatched",
            "occurrence_count": 1
        }
        
        # If from a registered player, add source info
        if source_player_id:
            insert_data["notes"] = f"Originally from player ID: {source_player_id} (different team: {from_different_team})"
        
        result = supabase.table("unmatched_players").insert(insert_data).execute()
        
        if not result.data or len(result.data) == 0:
            return jsonify({"success": False, "error": "Failed to add player"})
            
        return jsonify({
            "success": True, 
            "player_id": result.data[0]["id"],
            "occurrence_count": 1,
            "is_existing": False,
            "message": "Player added successfully"
        })
        
    except Exception as e:
        print(f"Error adding unmatched player: {str(e)}")
        return jsonify({"success": False, "error": str(e)})

@app.route('/edit_unmatched_player/<match_day_id>', methods=['POST'])
def edit_unmatched_player(match_day_id):
    """Update an unmatched player for a specific match day"""
    try:
        # Get data from request
        data = request.json if request.is_json else request.form.to_dict()
        player_id = data.get('player_id')
        name = data.get('name')
        team = data.get('team')
        
        # Validate input
        if not player_id or not name or not team:
            return jsonify({"success": False, "error": "Player ID, name, and team are required"})
        
        # Get match details
        match_result = supabase.table("matches").select("*").eq("id", match_day_id).execute()
        
        if not match_result.data or len(match_result.data) == 0:
            return jsonify({"success": False, "error": "Match not found"})
            
        match = match_result.data[0]
        team_id = match["home_team_id"] if team == "home" else match["away_team_id"]
        
        # Update the unmatched player
        result = supabase.table("unmatched_players").update({
            "name": name,
            "team_id": team_id
        }).eq("id", player_id).eq("match_id", match_day_id).execute()
        
        if not result.data or len(result.data) == 0:
            return jsonify({"success": False, "error": "Failed to update player"})
            
        return jsonify({
            "success": True,
            "message": "Player updated successfully"
        })
        
    except Exception as e:
        print(f"Error updating unmatched player: {str(e)}")
        return jsonify({"success": False, "error": str(e)})

@app.route('/delete_unmatched_player/<match_day_id>', methods=['POST'])
def delete_unmatched_player(match_day_id):
    """Delete an unmatched player for a specific match day"""
    try:
        # Get data from request
        data = request.json if request.is_json else request.form.to_dict()
        player_id = data.get('player_id')
        
        # Validate input
        if not player_id:
            return jsonify({"success": False, "error": "Player ID is required"})
        
        # Delete the unmatched player
        result = supabase.table("unmatched_players").delete().eq("id", player_id).eq("last_match_id", match_day_id).execute()
        
        if not result.data or len(result.data) == 0:
            return jsonify({"success": False, "error": "Failed to delete player"})
            
        return jsonify({
            "success": True,
            "message": "Player deleted successfully"
        })
        
    except Exception as e:
        print(f"Error deleting unmatched player: {str(e)}")
        return jsonify({"success": False, "error": str(e)})

@app.route('/match_player/<match_day_id>', methods=['POST'])
def match_player(match_day_id):
    """Match an unmatched player to an existing player"""
    try:
        # Get data from request
        data = request.json if request.is_json else request.form.to_dict()
        unmatched_player_id = data.get('unmatched_player_id')
        existing_player_id = data.get('existing_player_id')
        
        # Validate input
        if not unmatched_player_id or not existing_player_id:
            return jsonify({"success": False, "error": "Unmatched player ID and existing player ID are required"})
        
        # Get the unmatched player
        unmatched_result = supabase.table("unmatched_players").select("*").eq("id", unmatched_player_id).execute()
        
        if not unmatched_result.data or len(unmatched_result.data) == 0:
            return jsonify({"success": False, "error": "Unmatched player not found"})
            
        unmatched_player = unmatched_result.data[0]
        
        # Check if selecting an unmatched player or regular player
        if existing_player_id.startswith("unmatched_"):
            # Matching to another unmatched player
            # Extract the real ID
            real_unmatched_id = existing_player_id.replace("unmatched_", "")
            
            # Get the target unmatched player
            target_unmatched = supabase.table("unmatched_players").select("*").eq("id", real_unmatched_id).execute()
            
            if not target_unmatched.data or len(target_unmatched.data) == 0:
                return jsonify({"success": False, "error": "Target unmatched player not found"})
                
            target_player = target_unmatched.data[0]
            current_count = target_player.get("occurrence_count", 1) or 1
            
            # Update the target unmatched player's occurrence count
            supabase.table("unmatched_players").update({
                "occurrence_count": current_count + 1,
                "last_seen": time.strftime("%Y-%m-%d"),
                "last_match_id": match_day_id
            }).eq("id", real_unmatched_id).execute()
            
            # Mark the original unmatched player as matched to the target
            supabase.table("unmatched_players").update({
                "status": "merged",
                "matched_player_id": real_unmatched_id
            }).eq("id", unmatched_player_id).execute()
            
        else:
            # Matching to a regular player
            # Update the status of the unmatched player
            status_result = supabase.table("unmatched_players").update({
                "status": "matched",
                "matched_player_id": existing_player_id
            }).eq("id", unmatched_player_id).execute()
            
            if not status_result.data or len(status_result.data) == 0:
                return jsonify({"success": False, "error": "Failed to update unmatched player status"})
            
            # Create an appearance for the existing player if not already present
            appearance_result = supabase.table("appearances").select("*").eq("player_id", existing_player_id).eq("match_id", match_day_id).execute()
            
            if not appearance_result.data or len(appearance_result.data) == 0:
                # Create new appearance
                supabase.table("appearances").insert({
                    "player_id": existing_player_id,
                    "match_id": match_day_id
                }).execute()
                
                # Update total appearances for the player
                player = supabase.table("players").select("total_appearances").eq("id", existing_player_id).execute()
                if player.data and len(player.data) > 0:
                    current_count = player.data[0].get("total_appearances", 0) or 0
                    supabase.table("players").update({"total_appearances": current_count + 1}).eq("id", existing_player_id).execute()
        
        return jsonify({
            "success": True,
            "message": "Player matched successfully"
        })
        
    except Exception as e:
        print(f"Error matching player: {str(e)}")
        return jsonify({"success": False, "error": str(e)})

@app.route('/export/teams/excel')
def export_teams_excel():
    """Initiate Excel export process"""
    # Generate a unique ID for this export task
    export_id = str(uuid.uuid4())
    
    # Initialize export status
    with task_lock:
        processing_tasks[export_id] = {
            'status': 'starting',
            'current_step': 1,
            'step_details': {},
            'log_messages': [],
            'start_time': time.time(),
            'task_type': 'excel_export',
            'progress': 0,
            'teams_processed': 0,
            'total_teams': 0
        }
    
    # Start background thread for processing
    thread = threading.Thread(target=generate_excel_export, args=(export_id,))
    thread.daemon = True
    thread.start()
    
    # Redirect to processing page
    return redirect(url_for('export_progress', export_id=export_id))

@app.route('/export/progress/<export_id>')
def export_progress(export_id):
    """Show progress page for Excel export"""
    if export_id not in processing_tasks:
        flash('Invalid export ID', 'danger')
        return redirect(url_for('index'))
    
    task = processing_tasks[export_id]
    
    # Check if task is already complete
    if task['status'] == 'complete':
        return redirect(url_for('download_excel', export_id=export_id))
    
    # If there's an error, flash message and redirect
    if task['status'] == 'error':
        flash(f"Error generating Excel file: {task.get('error', 'Unknown error')}", 'danger')
        return redirect(url_for('index'))
    
    # Show processing page
    return render_template(
        'export_progress.html',
        export_id=export_id
    )

@app.route('/export/status/<export_id>')
def export_status(export_id):
    """API endpoint to get export status"""
    with task_lock:
        if export_id not in processing_tasks:
            return jsonify({'error': 'Invalid export ID'}), 404
        
        task = processing_tasks[export_id]
        
        response = {
            'status': task['status'],
            'progress': task.get('progress', 0),
            'teams_processed': task.get('teams_processed', 0),
            'total_teams': task.get('total_teams', 0),
            'timestamp': time.time()
        }
        
        # Add step-specific details if available
        if 'step_details' in task:
            response['step_details'] = task['step_details']
        
        # Add any log messages
        if 'log_messages' in task:
            response['log_messages'] = task['log_messages']
            # Reset log messages after sending (to avoid duplicates)
            task['log_messages'] = []
        
        if task['status'] == 'complete':
            response['download_url'] = url_for('download_excel', export_id=export_id)
        
        if task['status'] == 'error':
            response['error'] = task.get('error', 'Unknown error')
    
    return jsonify(response)

@app.route('/export/download/<export_id>')
def download_excel(export_id):
    """Download the generated Excel file"""
    if export_id not in processing_tasks:
        flash('Invalid export ID', 'danger')
        return redirect(url_for('index'))
    
    task = processing_tasks[export_id]
    
    if task['status'] != 'complete' or 'file_path' not in task:
        flash('Excel file not ready for download', 'warning')
        return redirect(url_for('export_progress', export_id=export_id))
    
    # Generate a filename with date
    current_date = time.strftime("%Y%m%d")
    filename = f"player_appearances_{current_date}.xlsx"
    
    # Send the file
    return send_file(
        task['file_path'],
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

def generate_excel_export(export_id):
    """Background task to generate Excel export"""
    with task_lock:
        task = processing_tasks[export_id]
        # Add log entry
        task['log_messages'].append({
            'message': "Starting Excel export process",
            'type': 'info'
        })
    
    try:
        # Create a progress logging function
        def log_progress(message, progress=None, step=None, details=None, message_type='info'):
            with task_lock:
                task = processing_tasks[export_id]
                task['log_messages'].append({
                    'message': message,
                    'type': message_type
                })
                if progress is not None:
                    task['progress'] = progress
                if step and details:
                    task['step_details'][str(step)] = details
                print(f"Export progress: {message}")
        
        # Create a new workbook
        wb = Workbook()
        
        log_progress("Fetching team data from database", 5, 1, "Fetching teams")
        
        # Get all teams
        teams = supabase.table("teams").select("*").order("name").execute()
        
        if not teams.data:
            with task_lock:
                task['status'] = 'error'
                task['error'] = 'No teams found to export'
            return
        
        # Set total teams count
        with task_lock:
            task['total_teams'] = len(teams.data)
        
        log_progress(f"Found {len(teams.data)} teams", 10, 1, f"Found {len(teams.data)} teams")
        
        # Get all matches
        log_progress("Fetching match data", 15, 2, "Fetching matches")
        matches = supabase.table("matches").select("id, match_day, date").order("date").execute()
        match_days = []
        match_map = {}
        
        if matches.data:
            # Create a mapping of match IDs to match days
            for match in matches.data:
                match_id = match["id"]
                match_day = match["match_day"]
                match_map[match_id] = match_day
                if match_day not in match_days:
                    match_days.append(match_day)
            
            log_progress(f"Found {len(matches.data)} matches across {len(match_days)} match days", 
                        20, 2, f"Found {len(matches.data)} matches")
        else:
            log_progress("No matches found", 20, 2, "No matches found", "warning")
        
        # Sort match days
        match_days.sort()
        
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        
        # Define styles
        log_progress("Setting up Excel styles and formats", 25, 3, "Setting up styles")
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        subheader_font = Font(bold=True, size=11)
        subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        unmatched_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # Create a summary sheet first
        log_progress("Creating summary sheet", 30, 4, "Creating summary")
        summary = wb.create_sheet(title="Summary", index=0)
        summary['A1'] = "Teams and Players Summary"
        summary['A1'].font = Font(bold=True, size=14)
        summary.merge_cells('A1:D1')
        
        summary['A3'] = "Team Name"
        summary['A3'].font = header_font
        summary['A3'].fill = header_fill
        summary['A3'].alignment = header_alignment
        summary['A3'].border = thin_border
        
        summary['B3'] = "Players"
        summary['B3'].font = header_font
        summary['B3'].fill = header_fill
        summary['B3'].alignment = header_alignment
        summary['B3'].border = thin_border
        
        summary['C3'] = "Unmatched Names"
        summary['C3'].font = header_font
        summary['C3'].fill = header_fill
        summary['C3'].alignment = header_alignment
        summary['C3'].border = thin_border
        
        summary['D3'] = "Total Appearances"
        summary['D3'].font = header_font
        summary['D3'].fill = header_fill
        summary['D3'].alignment = header_alignment
        summary['D3'].border = thin_border
        
        # Set column widths for summary
        summary.column_dimensions['A'].width = 30
        summary.column_dimensions['B'].width = 15
        summary.column_dimensions['C'].width = 20
        summary.column_dimensions['D'].width = 20
        
        # Process each team
        log_progress("Starting to process individual team data", 35, 5, "Processing teams")
        
        total_players = 0
        total_unmatched = 0
        total_appearances = 0
        summary_row_index = 4
        
        # Calculate base progress percentage and increment per team
        base_progress = 35
        progress_per_team = 55 / len(teams.data)  # 55% of progress bar allocated to team processing
        
        for team_index, team in enumerate(teams.data):
            team_id = team["id"]
            team_name = team["name"]
            
            current_progress = base_progress + (team_index * progress_per_team)
            log_progress(f"Processing team {team_index+1}/{len(teams.data)}: {team_name}", 
                        int(current_progress), 5, f"Team {team_index+1}/{len(teams.data)}")
            
            # Update teams processed counter
            with task_lock:
                task['teams_processed'] = team_index + 1
            
            # Create a sheet for this team
            # Ensure sheet name is valid (max 31 chars, no special chars)
            sheet_name = team_name[:31].replace('/', '_').replace('\\', '_').replace('?', '_').replace('*', '_').replace('[', '_').replace(']', '_').replace(':', '_')
            ws = wb.create_sheet(title=sheet_name)
            
            # Set column widths
            ws.column_dimensions['A'].width = 30  # Player name
            ws.column_dimensions['B'].width = 15  # Total appearances
            
            # Add headers
            ws['A1'] = team_name
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:E1')
            
            ws['A3'] = "Player Name"
            ws['A3'].font = header_font
            ws['A3'].fill = header_fill
            ws['A3'].alignment = header_alignment
            ws['A3'].border = thin_border
            
            ws['B3'] = "Total Appearances"
            ws['B3'].font = header_font
            ws['B3'].fill = header_fill
            ws['B3'].alignment = header_alignment
            ws['B3'].border = thin_border
            
            # Add match day columns
            col_index = 3  # Start from column C
            for match_day in match_days:
                col_letter = get_column_letter(col_index)
                ws[f'{col_letter}3'] = match_day
                ws[f'{col_letter}3'].font = header_font
                ws[f'{col_letter}3'].fill = header_fill
                ws[f'{col_letter}3'].alignment = header_alignment
                ws[f'{col_letter}3'].border = thin_border
                ws.column_dimensions[col_letter].width = 12
                col_index += 1
            
            # Get players for this team
            players = supabase.table("players").select("*").eq("team_id", team_id).order("name").execute()
            player_count = len(players.data) if players.data else 0
            total_players += player_count
            
            row_index = 4  # Start from row 4 for player data
            
            # Add player data
            if players.data:
                for player in players.data:
                    player_id = player["id"]
                    player_name = player["name"]
                    
                    # Add player name
                    ws[f'A{row_index}'] = player_name
                    ws[f'A{row_index}'].border = thin_border
                    
                    # Get player appearances
                    appearances = supabase.table("appearances").select("match_id").eq("player_id", player_id).execute()
                    
                    # Count total appearances
                    total_appearances_for_player = len(appearances.data) if appearances.data else 0
                    total_appearances += total_appearances_for_player
                    
                    ws[f'B{row_index}'] = total_appearances_for_player
                    ws[f'B{row_index}'].alignment = Alignment(horizontal="center")
                    ws[f'B{row_index}'].border = thin_border
                    
                    # Mark appearances by match day
                    if appearances.data:
                        col_index = 3  # Start from column C
                        for match_day in match_days:
                            col_letter = get_column_letter(col_index)
                            
                            # Check if player appeared in any match with this match day
                            appeared = False
                            for appearance in appearances.data:
                                match_id = appearance["match_id"]
                                if match_id in match_map and match_map[match_id] == match_day:
                                    appeared = True
                                    break
                            
                            if appeared:
                                ws[f'{col_letter}{row_index}'] = 1
                                ws[f'{col_letter}{row_index}'].alignment = Alignment(horizontal="center")
                            else:
                                ws[f'{col_letter}{row_index}'] = 0
                                ws[f'{col_letter}{row_index}'].alignment = Alignment(horizontal="center")
                            
                            ws[f'{col_letter}{row_index}'].border = thin_border
                            col_index += 1
                    
                    row_index += 1
            
            # Add a separator
            row_index += 1
            ws[f'A{row_index}'] = "Unmatched Player Names"
            ws[f'A{row_index}'].font = subheader_font
            ws[f'A{row_index}'].fill = subheader_fill
            ws.merge_cells(f'A{row_index}:E{row_index}')
            
            # Get unmatched players for this team
            unmatched_players = supabase.table("unmatched_players").select("*").eq("team_id", team_id).order("occurrence_count", desc=True).execute()
            unmatched_count = len(unmatched_players.data) if unmatched_players.data else 0
            total_unmatched += unmatched_count
            
            row_index += 1
            ws[f'A{row_index}'] = "Player Name"
            ws[f'A{row_index}'].font = subheader_font
            ws[f'A{row_index}'].fill = unmatched_fill
            ws[f'A{row_index}'].border = thin_border
            
            ws[f'B{row_index}'] = "Occurrences"
            ws[f'B{row_index}'].font = subheader_font
            ws[f'B{row_index}'].fill = unmatched_fill
            ws[f'B{row_index}'].alignment = Alignment(horizontal="center")
            ws[f'B{row_index}'].border = thin_border
            
            ws[f'C{row_index}'] = "First Seen"
            ws[f'C{row_index}'].font = subheader_font
            ws[f'C{row_index}'].fill = unmatched_fill
            ws[f'C{row_index}'].alignment = Alignment(horizontal="center")
            ws[f'C{row_index}'].border = thin_border
            
            ws[f'D{row_index}'] = "Last Seen"
            ws[f'D{row_index}'].font = subheader_font
            ws[f'D{row_index}'].fill = unmatched_fill
            ws[f'D{row_index}'].alignment = Alignment(horizontal="center")
            ws[f'D{row_index}'].border = thin_border
            
            ws[f'E{row_index}'] = "Last Match"
            ws[f'E{row_index}'].font = subheader_font
            ws[f'E{row_index}'].fill = unmatched_fill
            ws[f'E{row_index}'].alignment = Alignment(horizontal="center")
            ws[f'E{row_index}'].border = thin_border
            
            row_index += 1
            
            # Add unmatched player data
            if unmatched_players.data:
                for player in unmatched_players.data:
                    player_name = player["name"]
                    occurrences = player.get("occurrence_count", 1)
                    first_seen = player.get("first_seen", "")
                    last_seen = player.get("last_seen", "")
                    last_match_id = player.get("last_match_id", "")
                    last_match_day = match_map.get(last_match_id, "") if last_match_id else ""
                    
                    ws[f'A{row_index}'] = player_name
                    ws[f'A{row_index}'].border = thin_border
                    
                    ws[f'B{row_index}'] = occurrences
                    ws[f'B{row_index}'].alignment = Alignment(horizontal="center")
                    ws[f'B{row_index}'].border = thin_border
                    
                    ws[f'C{row_index}'] = first_seen
                    ws[f'C{row_index}'].alignment = Alignment(horizontal="center")
                    ws[f'C{row_index}'].border = thin_border
                    
                    ws[f'D{row_index}'] = last_seen
                    ws[f'D{row_index}'].alignment = Alignment(horizontal="center")
                    ws[f'D{row_index}'].border = thin_border
                    
                    ws[f'E{row_index}'] = last_match_day
                    ws[f'E{row_index}'].alignment = Alignment(horizontal="center")
                    ws[f'E{row_index}'].border = thin_border
                    
                    row_index += 1
            else:
                ws[f'A{row_index}'] = "No unmatched player names found"
                ws.merge_cells(f'A{row_index}:E{row_index}')
                ws[f'A{row_index}'].alignment = Alignment(horizontal="center")
                row_index += 1
            
            # Add team to summary sheet
            summary[f'A{summary_row_index}'] = team_name
            summary[f'A{summary_row_index}'].border = thin_border
            
            summary[f'B{summary_row_index}'] = player_count
            summary[f'B{summary_row_index}'].alignment = Alignment(horizontal="center")
            summary[f'B{summary_row_index}'].border = thin_border
            
            summary[f'C{summary_row_index}'] = unmatched_count
            summary[f'C{summary_row_index}'].alignment = Alignment(horizontal="center")
            summary[f'C{summary_row_index}'].border = thin_border
            
            # Count appearances for this team
            try:
                # Get appearances for all players in this team
                player_ids = [p["id"] for p in players.data] if players.data else []
                appearance_count = 0
                
                if player_ids:
                    # Get appearances for all players in this team
                    appearances_result = supabase.table("appearances").select("id").in_("player_id", player_ids).execute()
                    appearance_count = len(appearances_result.data) if appearances_result.data else 0
            except Exception as e:
                log_progress(f"Error counting appearances for team {team_name}: {str(e)}", None, None, None, "error")
                appearance_count = 0
            
            summary[f'D{summary_row_index}'] = appearance_count
            summary[f'D{summary_row_index}'].alignment = Alignment(horizontal="center")
            summary[f'D{summary_row_index}'].border = thin_border
            
            summary_row_index += 1
        
        # Add totals to summary sheet
        log_progress("Finalizing summary sheet", 90, 6, "Finalizing summary")
        
        summary[f'A{summary_row_index}'] = "TOTAL"
        summary[f'A{summary_row_index}'].font = Font(bold=True)
        summary[f'A{summary_row_index}'].border = thin_border
        
        summary[f'B{summary_row_index}'] = total_players
        summary[f'B{summary_row_index}'].font = Font(bold=True)
        summary[f'B{summary_row_index}'].alignment = Alignment(horizontal="center")
        summary[f'B{summary_row_index}'].border = thin_border
        
        summary[f'C{summary_row_index}'] = total_unmatched
        summary[f'C{summary_row_index}'].font = Font(bold=True)
        summary[f'C{summary_row_index}'].alignment = Alignment(horizontal="center")
        summary[f'C{summary_row_index}'].border = thin_border
        
        # Calculate total appearances from all teams
        try:
            # Get total appearances from database
            total_appearances_query = supabase.table("appearances").select("id").execute()
            if total_appearances_query.data:
                # Just count the total number of records returned
                total_appearances = len(total_appearances_query.data)
            else:
                # Fallback to our accumulated count
                log_progress("No appearances found in database, using accumulated count", None, None, None, "warning")
        except Exception as e:
            log_progress(f"Error counting total appearances: {str(e)}", None, None, None, "error")
            # Use the sum we've been accumulating if the query fails
        
        summary[f'D{summary_row_index}'] = total_appearances
        summary[f'D{summary_row_index}'].font = Font(bold=True)
        summary[f'D{summary_row_index}'].alignment = Alignment(horizontal="center")
        summary[f'D{summary_row_index}'].border = thin_border
        
        summary_row_index += 1
        
        # Save to a temporary file
        log_progress("Saving Excel file", 95, 7, "Saving file")
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        # Mark as complete
        with task_lock:
            task = processing_tasks[export_id]
            task['status'] = 'complete'
            task['file_path'] = temp_file.name
            task['progress'] = 100
            task['log_messages'].append({
                'message': "Excel export completed successfully!",
                'type': 'success'
            })
        
        log_progress("Excel export completed successfully!", 100, 7, "Complete", "success")
        
    except Exception as e:
        print(f"Error exporting Excel: {str(e)}")
        
        with task_lock:
            task = processing_tasks[export_id]
            task['status'] = 'error'
            task['error'] = str(e)
            task['log_messages'].append({
                'message': f"Error exporting Excel: {str(e)}",
                'type': 'error'
            })

# API routes
@app.route('/api/teams')
def get_all_teams_api():
    """Get all teams for dropdowns"""
    result = supabase.table("teams").select("id, name").order("name").execute()
    return jsonify(result.data if result.data else [])

@app.route('/api/team/<team_id>')
def get_team_api(team_id):
    """API endpoint to get team information by ID"""
    try:
        # Get team from database
        team = supabase.table("teams").select("*").eq("id", team_id).execute()
        
        if team.data and len(team.data) > 0:
            return jsonify(team.data[0])
        else:
            return jsonify({"error": "Team not found"}), 404
    except Exception as e:
        print(f"Error retrieving team: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/matches/<match_id>/update-teams', methods=['POST'])
def update_match_teams(match_id):
    """Update the home and away teams for a match"""
    try:
        # Get the request data
        data = request.json
        home_team_id = data.get('home_team_id')
        away_team_id = data.get('away_team_id')
        
        # Validate the input
        if not home_team_id or not away_team_id:
            return jsonify({"success": False, "error": "Both home and away team IDs are required"})
        
        if home_team_id == away_team_id:
            return jsonify({"success": False, "error": "Home and away teams cannot be the same"})
            
        # Get the current match data to check if teams have changed
        match_result = supabase.table("matches").select("*").eq("id", match_id).execute()
        
        if not match_result.data or len(match_result.data) == 0:
            return jsonify({"success": False, "error": "Match not found"})
            
        current_match = match_result.data[0]
        current_home_id = current_match.get("home_team_id")
        current_away_id = current_match.get("away_team_id")
        
        # If neither team has changed, return success immediately
        if current_home_id == home_team_id and current_away_id == away_team_id:
            return jsonify({"success": True, "message": "No changes needed"})
        
        # Update the match teams
        update_result = supabase.table("matches").update({
            "home_team_id": home_team_id,
            "away_team_id": away_team_id
        }).eq("id", match_id).execute()
        
        if not update_result.data or len(update_result.data) == 0:
            return jsonify({"success": False, "error": "Failed to update match teams"})
        
        # If teams have changed, we should clean up appearances
        # Players from previous teams should no longer be associated with this match
        if current_home_id != home_team_id or current_away_id != away_team_id:
            # Get all players from previous teams
            previous_players = []
            
            if current_home_id != home_team_id:
                home_players = get_players_by_team_id(current_home_id)
                previous_players.extend([p["id"] for p in home_players])
                
            if current_away_id != away_team_id:
                away_players = get_players_by_team_id(current_away_id)
                previous_players.extend([p["id"] for p in away_players])
            
            # Delete appearances for these players in this match
            if previous_players:
                for player_id in previous_players:
                    supabase.table("appearances").delete().eq("player_id", player_id).eq("match_id", match_id).execute()
        
        # Return success response
        return jsonify({
            "success": True,
            "message": "Match teams updated successfully",
            "match_id": match_id
        })
        
    except Exception as e:
        print(f"Error updating match teams: {str(e)}")
        return jsonify({"success": False, "error": f"An error occurred: {str(e)}"})

@app.route('/api/match_players/<match_id>', methods=['GET'])
def get_match_players_api(match_id):
    """API endpoint to get all players and unmatched players for match modals"""
    try:
        # Get match info to determine current teams
        match = supabase.table("matches").select("*, home_team:home_team_id(name), away_team:away_team_id(name)").eq("id", match_id).execute()
        
        if not match.data or len(match.data) == 0:
            return jsonify({"success": False, "error": "Match not found"})
            
        match_data = match.data[0]
        home_team_id = match_data['home_team_id']
        away_team_id = match_data['away_team_id']
        
        # Get all teams for grouping
        teams = supabase.table("teams").select("id, name").order("name").execute()
        teams_data = teams.data if teams.data else []
        
        # Get all players from all teams
        all_players = supabase.table("players").select("id, name, team_id").order("name").execute()
        all_players_data = all_players.data if all_players.data else []
        
        # Get all unmatched players that aren't already matched
        unmatched_players = supabase.table("unmatched_players").select("id, name, team_id, occurrence_count").eq("status", "unmatched").order("name").execute()
        unmatched_players_data = unmatched_players.data if unmatched_players.data else []
        
        # Organize players by team
        organized_players = {
            "match_teams": {
                "home": {
                    "id": home_team_id,
                    "name": match_data['home_team']['name'] if 'home_team' in match_data else "Unknown",
                    "players": []
                },
                "away": {
                    "id": away_team_id,
                    "name": match_data['away_team']['name'] if 'away_team' in match_data else "Unknown",
                    "players": []
                }
            },
            "other_teams": [],
            "unmatched_players": []
        }
        
        # Process all players
        for player in all_players_data:
            team_id = player.get("team_id")
            player_data = {
                "id": player.get("id"),
                "name": player.get("name"),
                "team_id": team_id
            }
            
            if team_id == home_team_id:
                organized_players["match_teams"]["home"]["players"].append(player_data)
            elif team_id == away_team_id:
                organized_players["match_teams"]["away"]["players"].append(player_data)
            else:
                # Find the team in other_teams or create it
                team_found = False
                for team in organized_players["other_teams"]:
                    if team["id"] == team_id:
                        team["players"].append(player_data)
                        team_found = True
                        break
                
                if not team_found:
                    # Find team name
                    team_name = "Unknown Team"
                    for team in teams_data:
                        if team["id"] == team_id:
                            team_name = team["name"]
                            break
                    
                    organized_players["other_teams"].append({
                        "id": team_id,
                        "name": team_name,
                        "players": [player_data]
                    })
        
        # Process unmatched players
        for player in unmatched_players_data:
            team_id = player.get("team_id")
            # Find team name
            team_name = "Unknown Team"
            for team in teams_data:
                if team["id"] == team_id:
                    team_name = team["name"]
                    break
                    
            organized_players["unmatched_players"].append({
                "id": player.get("id"),
                "name": player.get("name"),
                "team_id": team_id,
                "team_name": team_name,
                "occurrence_count": player.get("occurrence_count", 1)
            })
        
        return jsonify({
            "success": True,
            "players": organized_players
        })
        
    except Exception as e:
        print(f"Error retrieving players for match modal: {str(e)}")
        return jsonify({"success": False, "error": str(e)})

# Admin routes
@app.route('/admin/cleanup', methods=['GET'])
def admin_cleanup():
    """
    Endpoint to handle cleanup requests.
    """
    try:
        uploads_cleaned, frames_cleaned = file_manager.cleanup_old_files(
            app.config['UPLOAD_FOLDER'],
            app.config['FRAMES_FOLDER']
        )
        
        return jsonify({
            'success': True,
            'cleaned_uploads': uploads_cleaned,
            'cleaned_frames': frames_cleaned
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Make sure session is modified when storing results
@app.after_request
def after_request(response):
    # Mark session as modified to ensure it's saved
    if 'result' in session:
        session.modified = True
    return response

@app.route('/increment_unmatched_player/<match_day_id>', methods=['POST'])
def increment_unmatched_player(match_day_id):
    """Increment the occurrence count for an unmatched player"""
    try:
        # Get data from request
        data = request.json if request.is_json else request.form.to_dict()
        player_id = data.get('player_id')
        team = data.get('team')
        
        # Validate input
        if not player_id:
            return jsonify({"success": False, "error": "Player ID is required"})
        
        # Get match details
        match_result = supabase.table("matches").select("*").eq("id", match_day_id).execute()
        
        if not match_result.data or len(match_result.data) == 0:
            return jsonify({"success": False, "error": "Match not found"})
            
        match = match_result.data[0]
        
        # Get the player's current data
        player_result = supabase.table("unmatched_players").select("*").eq("id", player_id).execute()
        
        if not player_result.data or len(player_result.data) == 0:
            return jsonify({"success": False, "error": "Unmatched player not found"})
        
        player = player_result.data[0]
        new_occurrence_count = (player.get("occurrence_count") or 1) + 1
        
        # Update the team_id if it has changed
        team_id = None
        if team:
            team_id = match["home_team_id"] if team == "home" else match["away_team_id"]
        
        # Update unmatched player occurrence count and last_match_id
        update_data = {
            "occurrence_count": new_occurrence_count,
            "last_seen": time.strftime("%Y-%m-%d"),
            "last_match_id": match_day_id
        }
        
        # Only update team_id if provided
        if team_id:
            update_data["team_id"] = team_id
        
        result = supabase.table("unmatched_players").update(update_data).eq("id", player_id).execute()
        
        if not result.data or len(result.data) == 0:
            return jsonify({"success": False, "error": "Failed to update player"})
        
        # Check if this is the first time the player appears in this match
        is_new_for_match = player.get("last_match_id") != match_day_id
        
        return jsonify({
            "success": True,
            "player_id": player_id,
            "occurrence_count": new_occurrence_count,
            "is_new": is_new_for_match,
            "message": "Player occurrence count incremented"
        })
        
    except Exception as e:
        print(f"Error incrementing unmatched player: {str(e)}")
        return jsonify({"success": False, "error": str(e)})

@app.route('/decrement_unmatched_player/<match_day_id>', methods=['POST'])
def decrement_unmatched_player(match_day_id):
    """Decrement the occurrence count for an unmatched player"""
    try:
        # Get data from request
        data = request.json if request.is_json else request.form.to_dict()
        player_id = data.get('player_id')
        
        # Validate input
        if not player_id:
            return jsonify({"success": False, "error": "Player ID is required"})
        
        # Get the player's current data
        player_result = supabase.table("unmatched_players").select("*").eq("id", player_id).execute()
        
        if not player_result.data or len(player_result.data) == 0:
            return jsonify({"success": False, "error": "Unmatched player not found"})
        
        player = player_result.data[0]
        current_count = player.get("occurrence_count") or 1
        
        # Ensure we don't go below 1
        if current_count <= 1:
            return jsonify({"success": False, "error": "Player already has minimum occurrence count"})
        
        new_occurrence_count = current_count - 1
        
        # Update unmatched player occurrence count
        result = supabase.table("unmatched_players").update({
            "occurrence_count": new_occurrence_count
        }).eq("id", player_id).execute()
        
        if not result.data or len(result.data) == 0:
            return jsonify({"success": False, "error": "Failed to update player"})
        
        return jsonify({
            "success": True,
            "player_id": player_id,
            "occurrence_count": new_occurrence_count,
            "message": "Player occurrence count decremented"
        })
        
    except Exception as e:
        print(f"Error decrementing unmatched player: {str(e)}")
        return jsonify({"success": False, "error": str(e)})

# Application entry point
if __name__ == '__main__':
    # Schedule regular file cleanup
    file_manager.schedule_cleanup(app)
    app.run(debug=True) 