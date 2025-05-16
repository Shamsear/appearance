import openpyxl
from openpyxl import Workbook
import os

def create_sample_excel(filepath="APPEARANCE.xlsx"):
    """
    Create a sample Excel file with team sheets for testing
    """
    print(f"Creating sample Excel file at {filepath}...")
    
    # Create a new workbook
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    
    # Add team sheets
    teams = [
        "MANCHESTER UNITED", 
        "MANCHESTER CITY", 
        "LIVERPOOL", 
        "CHELSEA", 
        "ARSENAL", 
        "TOTTENHAM"
    ]
    
    # Add some players for each team
    players = {
        "MANCHESTER UNITED": [
            "RASHFORD", "FERNANDES", "SANCHO", "VARANE", "CASEMIRO",
            "MARTINEZ", "SHAW", "DE GEA", "ANTONY", "MCTOMINAY"
        ],
        "MANCHESTER CITY": [
            "HAALAND", "DE BRUYNE", "FODEN", "MAHREZ", "RODRI",
            "BERNARDO", "EDERSON", "DIAS", "WALKER", "GREALISH"
        ],
        "LIVERPOOL": [
            "SALAH", "VAN DIJK", "ALEXANDER-ARNOLD", "ALISSON", "DIAZ",
            "NUNEZ", "ROBERTSON", "FABINHO", "HENDERSON", "JOTA"
        ],
        "CHELSEA": [
            "MOUNT", "HAVERTZ", "SILVA", "KANTE", "PULISIC",
            "JAMES", "STERLING", "KEPA", "CHILWELL", "KOVACIC"
        ],
        "ARSENAL": [
            "SAKA", "ODEGAARD", "MARTINELLI", "RAMSDALE", "PARTEY",
            "JESUS", "WHITE", "GABRIEL", "XHAKA", "SMITH ROWE"
        ],
        "TOTTENHAM": [
            "KANE", "SON", "KULUSEVSKI", "RICHARLISON", "LLORIS",
            "ROMERO", "PERISIC", "DIER", "BENTANCUR", "EMERSON"
        ]
    }
    
    # Headers for each sheet
    headers = ["Player Name", "Position", "Status", "Appearances", "MD1", "MD2", "MD3", "MD4", "MD5"]
    
    # Create a sheet for each team with player data
    for team in teams:
        sheet = wb.create_sheet(team)
        
        # Add headers
        for col_idx, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_idx).value = header
        
        # Add players
        for row_idx, player in enumerate(players[team], 2):
            sheet.cell(row=row_idx, column=1).value = player
            sheet.cell(row=row_idx, column=2).value = "FWD" if row_idx % 3 == 0 else "MID" if row_idx % 3 == 1 else "DEF"
            sheet.cell(row=row_idx, column=3).value = "Active"
            sheet.cell(row=row_idx, column=4).value = 0
            
            # Add empty cells for match day columns
            for md_col in range(5, 10):
                sheet.cell(row=row_idx, column=md_col).value = ""
    
    # Save the workbook
    wb.save(filepath)
    print(f"Sample Excel file created with {len(teams)} teams.")
    return filepath

if __name__ == "__main__":
    create_sample_excel() 