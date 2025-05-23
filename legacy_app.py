import os
import time
import tempfile
import re
import json
import uuid
import threading
import concurrent.futures
import multiprocessing
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, send_file
from flask_session import Session
from werkzeug.utils import secure_filename
from dotenv import load_dotenv
from supabase import create_client, Client
import file_manager
import traceback
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

# Import the necessary functions from your existing script
from extract_and_update_player_data import (
    initialize_ocr, extract_frames, identify_player_ratings_frames,
    normalize_player_name, are_similar_names, extract_text_from_image
)

# Load environment variables
load_dotenv()

# Initialize Supabase client
supabase_url = os.environ.get("SUPABASE_URL")
supabase_key = os.environ.get("SUPABASE_KEY")
supabase: Client = create_client(supabase_url, supabase_key)

# Flask app configuration
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev_key_for_testing")

# Configure server-side sessions
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_FILE_DIR'] = os.path.join(os.getcwd(), 'flask_sessions')
app.config['SESSION_PERMANENT'] = False
app.config['SESSION_USE_SIGNER'] = True
Session(app)

app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['FRAMES_FOLDER'] = os.path.join('static', 'frames')
app.config['MAX_CONTENT_LENGTH'] = 500 * 1024 * 1024  # 500 MB max upload size
app.config['ALLOWED_EXTENSIONS'] = {'mp4', 'avi', 'mkv', 'mov'}

# Make sure folders exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['FRAMES_FOLDER'], exist_ok=True)
os.makedirs(app.config['SESSION_FILE_DIR'], exist_ok=True)

# Global variable to track processing status
processing_tasks = {}
# Lock for thread-safe access to processing_tasks
task_lock = threading.Lock()

# Get the number of CPU cores available
cpu_count = multiprocessing.cpu_count()
# Use at most CPU cores or 5 threads, whichever is smaller
max_workers = min(cpu_count, 5)
# Global thread pool executor for parallel processing
thread_pool_executor = concurrent.futures.ThreadPoolExecutor(max_workers=max_workers)

# Global variable to track batch processing
batch_processing_tasks = {}
# Lock for thread-safe access to batch processing tasks
batch_lock = threading.Lock()

# Function to clean up old tasks
def cleanup_old_tasks():
    current_time = time.time()
    expired_ids = []
    
    with task_lock:
        for task_id, task in list(processing_tasks.items()):
            # Add start_time if missing (for legacy tasks)
            if 'start_time' not in task:
                task['start_time'] = current_time
                
            # Remove tasks older than 1 hour
            if current_time - task['start_time'] > 3600:  # 1 hour
                expired_ids.append(task_id)
                print(f"Removing expired task {task_id} - {task['status']}")
        
        # Remove expired tasks
        for task_id in expired_ids:
            del processing_tasks[task_id]
    
    # Also clean up batch processing tasks
    expired_batch_ids = []
    with batch_lock:
        for batch_id, batch in list(batch_processing_tasks.items()):
            # Add start_time if missing (for legacy tasks)
            if 'start_time' not in batch:
                batch['start_time'] = current_time
                
            # For completed batches, keep them for 24 hours for user reference
            retention_time = 86400  # 24 hours for completed batches
            if batch['status'] not in ['complete', 'complete_with_errors', 'error']:
                # For non-completed batches, keep them for only 2 hours
                retention_time = 7200  # 2 hours
                
            if current_time - batch['start_time'] > retention_time:
                expired_batch_ids.append(batch_id)
                print(f"Removing expired batch {batch_id} - {batch['status']}")
        
        # Remove expired batch tasks
        for batch_id in expired_batch_ids:
            del batch_processing_tasks[batch_id]

# Clean up old tasks before each request
@app.before_request
def before_request_cleanup():
    cleanup_old_tasks()

# Helper function to check if file extension is allowed
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

# Extract teams from filename
def extract_teams_from_filename(filename):
    match = re.match(r'([a-z\s]+)\s+vs\s+([a-z\s]+)', filename.lower())
    
    if match:
        home_team = match.group(1).strip().upper()
        away_team = match.group(2).strip().upper()
        return home_team, away_team
    
    return None, None

# Get match day from filename
def extract_match_day(filename):
    match_day = "MD1"  # Default
    if "md" in filename.lower():
        md_match = re.search(r'md(\d+)', filename.lower())
        if md_match:
            match_day = f"MD{md_match.group(1)}"
    return match_day

# Find team by name
def find_team_by_name(team_name):
    result = supabase.table("teams").select("*").ilike("name", team_name).execute()
    if result.data and len(result.data) > 0:
        return result.data[0]
    return None

# Get players by team ID
def get_players_by_team_id(team_id):
    result = supabase.table("players").select("*").eq("team_id", team_id).execute()
    return result.data if result.data else []

# Update player appearances
def update_player_appearances(player_id, match_id):
    # Check if this appearance already exists
    result = supabase.table("appearances").select("*").eq("player_id", player_id).eq("match_id", match_id).execute()
    
    if result.data and len(result.data) > 0:
        # Already exists, don't create duplicate
        return False
        
    # Create new appearance record
    result = supabase.table("appearances").insert({
        "player_id": player_id,
        "match_id": match_id
    }).execute()
    
    if result.data and len(result.data) > 0:
        # Update total_appearances count in players table
        player = supabase.table("players").select("total_appearances").eq("id", player_id).execute()
        if player.data and len(player.data) > 0:
            current_count = player.data[0].get("total_appearances", 0) or 0
            supabase.table("players").update({"total_appearances": current_count + 1}).eq("id", player_id).execute()
        return True
    
    return False

# Process text to extract player names with improved filtering
def process_text_to_extract_names(ocr_results, section_type="home"):
    extracted_names = []
    
    # Known UI elements and false positives to exclude
    ui_elements = [
        "player ratings", "player rating", "ratings", "rating",
        "home", "away", "back", "next", "prev", "click",
        "menu", "main menu", "main", "page", "ack", "maia",
        "irwf", "squad", "rank", "rankings", "default", "formation", 
        "tactics", "attacking", "balanced", "defensive", "awav"
    ]
    
    # Common football positions to exclude
    positions = [
        "gk", "goalkeeper", "goalie",
        "lb", "rb", "cb", "wb", "sweeper", "libero",
        "cdm", "cm", "cam", "lm", "rm", "lwb", "rwb", 
        "lw", "rw", "cf", "st", "striker", "forward",
        "defender", "midfielder", "winger", "fullback",
        "center back", "centre back", "left back", "right back",
        "defensive mid", "central mid", "attacking mid",
        "left wing", "right wing", "center forward"
    ]
    
    for detection in ocr_results:
        text = detection[1]
        confidence = detection[2]
        
        # Skip low confidence detections
        if confidence < 0.2:
            continue
        
        # Skip ratings values (numbers)
        if re.match(r'^[\d.]+$', text):
            continue
        
        # Skip position indicators (expanded to include more variations)
        if re.match(r'^[A-Z]{1,3}$', text):
            continue
            
        # Skip longer position designations (like "ST" with a number)  
        if re.match(r'^[A-Z]{1,3}\d*$', text):
            continue
        
        # Skip UI elements and false positives
        text_lower = text.lower()
        if any(ui_text in text_lower for ui_text in ui_elements):
            continue
            
        # Skip football positions
        if text_lower in positions or any(pos == text_lower for pos in positions):
            continue
        
        # Clean the name from any trailing numbers (like "90" or "95")
        clean_name = re.sub(r'\s+\d+$', '', text)
        clean_name = clean_name.strip()
        
        # Check for fully visible names with proper formatting
        if len(clean_name) >= 3 and clean_name[0].isupper():
            # Skip partially visible/truncated names
            
            # Check for missing spaces in names that should have spaces
            if len(clean_name) > 10 and ' ' not in clean_name:
                continue
            
            # Skip names that appear to be cut off (e.g., ending with unusual characters)
            if re.search(r'[^\w\s]$', clean_name):
                continue
                
            # Skip names that are suspiciously short for a full name
            if len(clean_name.split()) > 1 and any(len(part) < 2 for part in clean_name.split()):
                continue
                
            # Skip names with unusual character patterns that suggest partial visibility
            if re.search(r'[^A-Za-z\s\'\-]', clean_name):
                continue
                
            # Add the name if it passes all checks and is not already in the list
            if clean_name not in extracted_names:
                extracted_names.append(clean_name)
    
    return extracted_names

# Function to group frames by player names
def group_frames_by_players(player_frames):
    """Group frames that contain the same set of player names, keeping only one representative frame per group"""
    print("Grouping frames with similar player names...")
    
    # Dictionary to track player sets and their representative frames
    player_sets = {}
    
    # First, extract player names from each frame
    for frame in player_frames:
        # Get all the detected player names for this frame
        names = []
        for detection in frame["ocr_results"]:
            text = detection["text"]
            confidence = detection["confidence"]
            
            # Filter out non-player data
            if confidence > 0.2 and not any(ui in text.lower() for ui in ["player ratings", "home", "away", "back"]):
                names.append(text)
        
        # Process the names to extract actual player names
        if frame["is_home"]:
            extracted_names = process_text_to_extract_names([
                [None, item["text"], item["confidence"]] 
                for item in frame["ocr_results"]
            ], "home")
        elif frame["is_away"]:
            extracted_names = process_text_to_extract_names([
                [None, item["text"], item["confidence"]] 
                for item in frame["ocr_results"]
            ], "away")
        else:
            extracted_names = []
        
        # Sort and convert to tuple to use as dictionary key
        extracted_names.sort()
        names_tuple = tuple(extracted_names)
        
        # Skip if no names found
        if not names_tuple:
            continue
        
        # Track which team this frame belongs to
        team_type = "home" if frame["is_home"] else "away"
        
        # Add to our group tracking
        key = (team_type, names_tuple)
        if key not in player_sets:
            player_sets[key] = []
        player_sets[key].append(frame)
    
    # Create a list of deduplicated frames
    deduplicated_frames = []
    
    # For each group, keep only the first frame
    for (team_type, names), frames in player_sets.items():
        print(f"Found {len(frames)} {team_type} frames with players: {', '.join(names)}")
        # Keep the first frame from each group
        deduplicated_frames.append(frames[0])
    
    print(f"Reduced from {len(player_frames)} to {len(deduplicated_frames)} frames after deduplication")
    return deduplicated_frames

# Extract frames and save for review
def extract_frames_for_review(video_path, home_team_id, away_team_id, match_day, callback=None):
    # Create a unique session ID for this processing session
    session_id = str(uuid.uuid4())
    session_dir = os.path.join(app.config['FRAMES_FOLDER'], session_id)
    os.makedirs(session_dir, exist_ok=True)
    
    print(f"Starting frame extraction for review, saving to {session_dir}")
    if callback:
        callback(f"Starting frame extraction for review, saving to session directory", 2, "Initializing")
    
    # Validate team IDs (ensure they're not None or "None")
    if not home_team_id or home_team_id == "None":
        print(f"Warning: Invalid home team ID: {home_team_id}")
        if callback:
            callback(f"Invalid home team ID: {home_team_id}", 2, "Error", "error")
        return None, "Invalid home team ID"
        
    if not away_team_id or away_team_id == "None":
        print(f"Warning: Invalid away team ID: {away_team_id}")
        if callback:
            callback(f"Invalid away team ID: {away_team_id}", 2, "Error", "error")
        return None, "Invalid away team ID"
    
    # Initialize OCR
    reader = initialize_ocr()
    print(f"OCR initialized")
    if callback:
        callback("OCR engine initialized", 2, "OCR Ready")
    
    # Extract frames
    print(f"Extracting frames from {video_path}")
    if callback:
        callback(f"Extracting frames from video: {os.path.basename(video_path)}", 2, "Extracting frames")
    
    start_time = time.time()
    frames = extract_frames(video_path, session_dir)
    extraction_time = time.time() - start_time
    
    if not frames:
        print(f"No frames could be extracted from the video at {video_path}")
        if callback:
            callback("No frames could be extracted from the video", 2, "Error", "error")
        return None, "No frames could be extracted from the video"
    
    print(f"Successfully extracted {len(frames)} frames")
    if callback:
        callback(f"Successfully extracted {len(frames)} frames in {extraction_time:.1f} seconds", 2, f"Extracted {len(frames)} frames")
        # Update the task with the total frames extracted
        with task_lock:
            for task_id, task_data in processing_tasks.items():
                if task_data.get('filepath') == video_path:
                    task_data['frames_processed'] = len(frames)
                    break
    
    # Find frames with player ratings
    player_frames = []
    
    print(f"Scanning frames for player ratings...")
    if callback:
        callback("Scanning frames for player ratings", 3, "Scanning frames")
    
    frame_count = 0
    home_frame_count = 0
    away_frame_count = 0
    
    for frame_path in frames:
        frame_count += 1
        if frame_count % 10 == 0:
            print(f"Processed {frame_count}/{len(frames)} frames")
            if callback:
                progress_percent = int((frame_count / len(frames)) * 100)
                callback(f"Processed {frame_count}/{len(frames)} frames ({progress_percent}%)", 3, f"Scanned {frame_count} frames")
                # Update the task with current progress
                with task_lock:
                    for task_id, task_data in processing_tasks.items():
                        if task_data.get('filepath') == video_path:
                            task_data['frames_processed'] = frame_count
                            break
        
        # Get OCR results for the frame
        ocr_results = extract_text_from_image(frame_path, reader)
        
        # Check if this frame contains player ratings
        is_home = False
        is_away = False
        for detection in ocr_results:
            if "Player Ratings: Home" in detection[1]:
                is_home = True
                home_frame_count += 1
                print(f"Found home player ratings frame: {frame_path}")
                if callback:
                    callback(f"Found home player ratings frame {home_frame_count}", 3, f"Found {home_frame_count} home frames", "success")
            if "Player Ratings: Away" in detection[1]:
                is_away = True
                away_frame_count += 1
                print(f"Found away player ratings frame: {frame_path}")
                if callback:
                    callback(f"Found away player ratings frame {away_frame_count}", 3, f"Found {away_frame_count} away frames", "success")
        
        # If it's a player ratings frame, save for review
        if is_home or is_away:
            # Get relative path for template
            rel_path = os.path.relpath(frame_path, app.config['FRAMES_FOLDER'])
            
            # Process text to extract player names
            extracted_text = []
            for detection in ocr_results:
                text = detection[1]
                confidence = detection[2]
                if confidence > 0.2:
                    # Skip known UI elements
                    text_lower = text.lower()
                    if text_lower == "player ratings" or text_lower == "home" or text_lower == "away" or text_lower == "back":
                        continue
                    
                    extracted_text.append({
                        "text": text,
                        "confidence": float(confidence)
                    })
            
            # Add to player frames list
            player_frames.append({
                "path": rel_path,
                "is_home": is_home,
                "is_away": is_away,
                "ocr_results": extracted_text
            })
    
    total_player_frames = len(player_frames)
    print(f"Found {total_player_frames} player rating frames ({home_frame_count} home, {away_frame_count} away)")
    if callback:
        callback(f"Found {total_player_frames} player rating frames ({home_frame_count} home, {away_frame_count} away)", 
                 4, f"Found {total_player_frames} frames", "success")
        # Update the task with player frames count
        with task_lock:
            for task_id, task_data in processing_tasks.items():
                if task_data.get('filepath') == video_path:
                    task_data['player_frames'] = total_player_frames
                    break
    
    # Store original frame count before deduplication
    original_frame_count = len(player_frames)
    
    # Deduplicate frames that have the same player names
    if callback:
        callback("Starting frame deduplication to remove duplicate player sets", 5, "Deduplicating frames")
    
    deduplicated_frames = group_frames_by_players(player_frames)
    
    # Calculate deduplication stats
    frames_removed = original_frame_count - len(deduplicated_frames)
    reduction_percent = int((frames_removed / original_frame_count) * 100) if original_frame_count > 0 else 0
    
    if frames_removed > 0:
        print(f"Removed {frames_removed} duplicate frames ({reduction_percent}% reduction)")
        if callback:
            callback(f"Removed {frames_removed} duplicate frames ({reduction_percent}% reduction)", 
                     5, f"Deduplicated {frames_removed} frames", "success")
            # Update task with deduplication stats
            with task_lock:
                for task_id, task_data in processing_tasks.items():
                    if task_data.get('filepath') == video_path:
                        task_data['original_frame_count'] = original_frame_count
                        task_data['deduplicated_frame_count'] = len(deduplicated_frames)
                        break
    
    # Create session data
    session_data = {
        "session_id": session_id,
        "home_team_id": home_team_id,
        "away_team_id": away_team_id,
        "match_day": match_day,
        "frames": deduplicated_frames,
        "video_path": video_path,
        "original_frame_count": original_frame_count
    }
    
    # Save session data
    session_file = os.path.join(session_dir, "session.json")
    with open(session_file, "w") as f:
        json.dump(session_data, f)
    
    print(f"Saved session data to {session_file}")
    if callback:
        callback("Session data saved, ready for review", 5, "Review ready", "success")
    
    return session_id, None

# Store unmatched player names in database
def store_unmatched_player(name, team_id, match_id):
    """Store unmatched player names in the database for future reference"""
    try:
        # Check if this unmatched player + team combination already exists
        result = supabase.table("unmatched_players").select("*").eq("name", name).eq("team_id", team_id).execute()
        
        if result.data and len(result.data) > 0:
            # If it exists, update the last_seen field and increment occurrence count
            player_id = result.data[0]["id"]
            occurrence_count = result.data[0].get("occurrence_count", 1) + 1
            
            supabase.table("unmatched_players").update({
                "last_seen": time.strftime("%Y-%m-%d"),
                "occurrence_count": occurrence_count,
                "last_match_id": match_id
            }).eq("id", player_id).execute()
            
            print(f"Updated unmatched player: {name} for team {team_id} (seen {occurrence_count} times)")
            return player_id
        else:
            # Create new unmatched player record
            result = supabase.table("unmatched_players").insert({
                "name": name,
                "team_id": team_id,
                "first_seen": time.strftime("%Y-%m-%d"),
                "last_seen": time.strftime("%Y-%m-%d"),
                "occurrence_count": 1,
                "last_match_id": match_id
            }).execute()
            
            if result.data and len(result.data) > 0:
                print(f"Added new unmatched player: {name} for team {team_id}")
                return result.data[0]["id"]
            
    except Exception as e:
        print(f"Error storing unmatched player {name}: {str(e)}")
    
    return None

# Process uploaded video and update database
def process_video(video_path, home_team_id, away_team_id, match_day, progress_callback=None):
    print(f"Starting automatic video processing for {video_path}")
    if progress_callback:
        progress_callback(f"Starting automatic video processing for {os.path.basename(video_path)}")
    
    # Create match record
    match_data = {
        "home_team_id": home_team_id,
        "away_team_id": away_team_id,
        "match_day": match_day,
        "date": time.strftime("%Y-%m-%d")
    }
    
    print(f"Creating match record for {match_day}: home={home_team_id}, away={away_team_id}")
    if progress_callback:
        progress_callback(f"Creating match record for {match_day}", 2, "Creating match record")
    
    match_result = supabase.table("matches").insert(match_data).execute()
    
    if not match_result.data or len(match_result.data) == 0:
        error_msg = "Failed to create match record"
        print(f"Error: {error_msg}")
        if progress_callback:
            progress_callback(f"Error: {error_msg}", 2, "Database error", "error")
        return {"error": error_msg}, None, None
    
    match_id = match_result.data[0]["id"]
    print(f"Created match record with ID: {match_id}")
    if progress_callback:
        progress_callback(f"Created match record with ID: {match_id}", 2, "Match record created", "success")
    
    # Initialize OCR
    reader = initialize_ocr()
    print(f"OCR initialized")
    if progress_callback:
        progress_callback("OCR engine initialized", 2, "OCR Ready")
    
    # Extract frames
    frames_dir = tempfile.mkdtemp()
    print(f"Extracting frames to temporary directory: {frames_dir}")
    if progress_callback:
        progress_callback(f"Extracting frames to temporary directory", 2, "Extracting frames")
    
    start_time = time.time()
    frames = extract_frames(video_path, frames_dir)
    extraction_time = time.time() - start_time
    
    if not frames:
        error_msg = "No frames extracted from video"
        print(f"Error: {error_msg}")
        if progress_callback:
            progress_callback(f"Error: {error_msg}", 2, "Extraction error", "error")
        return {"error": error_msg}, None, None
    
    print(f"Successfully extracted {len(frames)} frames")
    if progress_callback:
        progress_callback(f"Successfully extracted {len(frames)} frames in {extraction_time:.1f} seconds", 2, f"Extracted {len(frames)} frames", "success")
        # Update task with frames count
        with task_lock:
            for task_id, task_data in processing_tasks.items():
                if task_data.get('filepath') == video_path:
                    task_data['frames_processed'] = len(frames)
                    break
    
    # Identify player ratings frames and extract names
    print(f"Identifying player ratings frames...")
    if progress_callback:
        progress_callback("Identifying player ratings frames", 3, "Analyzing frames")
    
    start_time = time.time()
    home_players, away_players = identify_player_ratings_frames(frames, reader)
    identification_time = time.time() - start_time
    
    # Apply additional filtering to remove UI elements 
    home_players = [p for p in home_players if not any(ui in p.lower() for ui in [
        "player ratings", "player rating", "home", "away", "back"
    ])]
    
    away_players = [p for p in away_players if not any(ui in p.lower() for ui in [
        "player ratings", "player rating", "home", "away", "back"
    ])]
    
    print(f"Found {len(home_players)} home players and {len(away_players)} away players")
    if progress_callback:
        progress_callback(f"Found {len(home_players)} home players and {len(away_players)} away players in {identification_time:.1f} seconds", 
                         3, f"Found {len(home_players) + len(away_players)} players", "success")
        # Update task with player frames count
        with task_lock:
            for task_id, task_data in processing_tasks.items():
                if task_data.get('filepath') == video_path:
                    task_data['player_frames'] = len(home_players) + len(away_players)
                    break
    
    # Get players from database
    db_home_players = get_players_by_team_id(home_team_id)
    db_away_players = get_players_by_team_id(away_team_id)
    
    print(f"Found {len(db_home_players)} home players and {len(db_away_players)} away players in database")
    if progress_callback:
        progress_callback(f"Found {len(db_home_players)} home players and {len(db_away_players)} away players in database", 
                         4, "Matching players", "info")
    
    # Track which players appeared
    home_appearances = []
    away_appearances = []
    
    # Process home players
    print(f"Matching home players...")
    if progress_callback:
        progress_callback(f"Matching home players with database records", 4, "Matching home players")
    
    matched_home_player_names = set()
    for extracted_name in home_players:
        print(f"  Trying to match home player: {extracted_name}")
        found_match = False
        for db_player in db_home_players:
            # Check if names match
            if are_similar_names(extracted_name, db_player["name"]):
                print(f"  Matched {extracted_name} with {db_player['name']}")
                matched_home_player_names.add(extracted_name.lower())
                if progress_callback:
                    progress_callback(f"Matched home player: {extracted_name} → {db_player['name']}", 4, "Matching names", "success")
                # Update appearance
                if update_player_appearances(db_player["id"], match_id):
                    home_appearances.append({
                        "id": db_player["id"],
                        "name": db_player["name"],
                        "extracted_name": extracted_name
                    })
                    print(f"  Updated appearance for {db_player['name']}")
                    found_match = True
                break
        
        # Store unmatched player name
        if not found_match:
            store_unmatched_player(extracted_name, home_team_id, match_id)
            if progress_callback:
                progress_callback(f"Storing unmatched home player: {extracted_name}", 4, "Unmatched name", "warning")
    
    # Process away players
    print(f"Matching away players...")
    if progress_callback:
        progress_callback(f"Matching away players with database records", 4, "Matching away players")
    
    matched_away_player_names = set()
    for extracted_name in away_players:
        print(f"  Trying to match away player: {extracted_name}")
        found_match = False
        for db_player in db_away_players:
            # Check if names match
            if are_similar_names(extracted_name, db_player["name"]):
                print(f"  Matched {extracted_name} with {db_player['name']}")
                matched_away_player_names.add(extracted_name.lower())
                if progress_callback:
                    progress_callback(f"Matched away player: {extracted_name} → {db_player['name']}", 4, "Matching names", "success")
                # Update appearance
                if update_player_appearances(db_player["id"], match_id):
                    away_appearances.append({
                        "id": db_player["id"],
                        "name": db_player["name"],
                        "extracted_name": extracted_name  # Using DB name since user confirmed
                    })
                    print(f"  Updated appearance for {db_player['name']}")
                    found_match = True
                break
        
        # Store unmatched player name
        if not found_match:
            store_unmatched_player(extracted_name, away_team_id, match_id)
            if progress_callback:
                progress_callback(f"Storing unmatched away player: {extracted_name}", 4, "Unmatched name", "warning")
    
    # Cleanup
    import shutil
    print(f"Cleaning up temporary directory: {frames_dir}")
    if progress_callback:
        progress_callback("Cleaning up temporary files", 5, "Cleanup")
    shutil.rmtree(frames_dir, ignore_errors=True)
    
    result = {
        "match_id": match_id,
        "home_players_extracted": len(matched_home_player_names),
        "away_players_extracted": len(matched_away_player_names),
        "home_players_matched": len(home_appearances),
        "away_players_matched": len(away_appearances),
        "home_players_unmatched": len(matched_home_player_names) - len(home_appearances),
        "away_players_unmatched": len(matched_away_player_names) - len(away_appearances),
    }
    
    print(f"Processing complete. Result: {result}")
    if progress_callback:
        total_matched = len(home_appearances) + len(away_appearances)
        progress_callback(f"Processing complete. Updated {total_matched} player appearances.", 5, f"Updated {total_matched} players", "success")
    
    return result, home_appearances, away_appearances

@app.route('/')
def index():
    # Check for active processing tasks
    active_tasks = get_active_processing_tasks()
    
    if active_tasks:
        # Redirect to the active process view that shows ongoing processes
        return redirect(url_for('active_processes'))
    
    # Get all teams for dropdown
    teams = supabase.table("teams").select("*").order("name").execute()
    teams_data = teams.data if teams.data else []
    
    return render_template('index.html', teams=teams_data)

def get_active_processing_tasks():
    """Get all active processing tasks that are not completed or errored"""
    active_tasks = []
    
    # Get current time for comparison
    current_time = time.time()
    
    # Check single video processing tasks
    with task_lock:
        for task_id, task in processing_tasks.items():
            # Skip Excel exports or completed/errored tasks
            if task.get('task_type') == 'excel_export':
                continue
            if task.get('status') in ['complete', 'error']:
                continue
            # Only include tasks that started within the last hour (to avoid showing very old tasks)
            if 'start_time' in task and current_time - task['start_time'] < 3600:
                active_tasks.append({
                    'id': task_id,
                    'type': 'single',
                    'status': task.get('status', 'unknown'),
                    'process_type': task.get('process_type', 'auto'),
                    'current_step': task.get('current_step', 1),
                    'start_time': task.get('start_time', current_time)
                })
    
    # Check batch processing tasks
    with batch_lock:
        for batch_id, batch in batch_processing_tasks.items():
            # Skip completed/errored batches
            if batch.get('status') in ['complete', 'complete_with_errors', 'error']:
                continue
            # Only include batches that started within the last 2 hours
            if 'start_time' in batch and current_time - batch['start_time'] < 7200:
                active_tasks.append({
                    'id': batch_id,
                    'type': 'batch',
                    'status': batch.get('status', 'unknown'),
                    'process_type': batch.get('process_type', 'auto'),
                    'total_videos': len(batch.get('videos', [])),
                    'completed_videos': batch.get('completed_videos', 0),
                    'start_time': batch.get('start_time', current_time)
                })
    
    return active_tasks

# Register context processor directly
@app.context_processor
def inject_active_task_count():
    active_tasks = get_active_processing_tasks()
    return {'active_task_count': len(active_tasks)}

@app.route('/active-processes')
def active_processes():
    """Show page with all active processing tasks"""
    active_tasks = get_active_processing_tasks()
    
    if not active_tasks:
        # If no active tasks, redirect back to index
        flash('No active processing tasks found', 'info')
        return redirect(url_for('index'))
    
    return render_template('active_processes.html', active_tasks=active_tasks)

@app.route('/upload_video', methods=['POST'])
def upload_video():
    if 'video' not in request.files:
        flash('No video file uploaded', 'danger')
        return redirect(url_for('index'))
    
    file = request.files['video']
    
    if file.filename == '':
        flash('No video selected', 'danger')
        return redirect(url_for('index'))
    
    if not allowed_file(file.filename):
        flash('Invalid file type. Allowed: mp4, avi, mkv, mov', 'danger')
        return redirect(url_for('index'))
    
    # Get team IDs from form
    home_team_id = request.form.get('home_team')
    away_team_id = request.form.get('away_team')
    match_day = request.form.get('match_day', 'MD1')
    process_type = request.form.get('process_type', 'auto')
    
    # Validate team IDs
    if not home_team_id or home_team_id == "None":
        flash('Invalid home team selected', 'danger')
        return redirect(url_for('index'))
        
    if not away_team_id or away_team_id == "None":
        flash('Invalid away team selected', 'danger')
        return redirect(url_for('index'))
    
    # Save uploaded file
    filename = secure_filename(file.filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)
    
    # Generate a unique ID for this upload
    upload_id = str(uuid.uuid4())
    
    # Initialize processing status
    with task_lock:
        processing_tasks[upload_id] = {
            'status': 'uploading',
            'current_step': 1,
            'filepath': filepath,
            'home_team_id': home_team_id,
            'away_team_id': away_team_id,
            'match_day': match_day,
            'process_type': process_type,
            'result': None,
            'session_id': None,
            'error': None,
            'start_time': time.time()
        }
    
    # Start background thread for processing
    if process_type == 'review':
        thread = threading.Thread(target=process_video_for_review, args=(upload_id,))
        thread.daemon = True  # Set as daemon thread
        thread.start()
    else:
        thread = threading.Thread(target=process_video_automatically, args=(upload_id,))
        thread.daemon = True  # Set as daemon thread
        thread.start()
    
    # Redirect to processing page
    return redirect(url_for('processing_page', upload_id=upload_id))

def check_processing_timeout(task, timeout_minutes=10):
    """Check if a processing task has exceeded the timeout limit"""
    if 'start_time' not in task:
        return False
    
    elapsed_time = time.time() - task['start_time']
    timeout_seconds = timeout_minutes * 60
    
    if elapsed_time > timeout_seconds:
        return True
    return False

def process_video_automatically(upload_id):
    """Process the video automatically without user intervention"""
    with task_lock:
        task = processing_tasks[upload_id]
        # Add start time for expiration tracking
        task['start_time'] = time.time()
        # Initialize log messages array
        task['log_messages'] = []
        # Initialize step details
        task['step_details'] = {}
        # Initialize frame counters
        task['frames_processed'] = 0
        task['player_frames'] = 0
        # Update status to extracting frames
        task['status'] = 'extracting_frames'
        task['current_step'] = 2
        # Add log entry
        task['log_messages'].append({
            'message': f"Starting to process video: {os.path.basename(task['filepath'])}",
            'type': 'info'
        })
    
    try:
        # Check for timeout periodically
        if check_processing_timeout(task):
            with task_lock:
                task['status'] = 'error'
                task['error'] = 'Processing timed out after 10 minutes'
                task['log_messages'].append({
                    'message': 'Processing timed out after 10 minutes',
                    'type': 'error'
                })
            print(f"Task {upload_id} timed out")
            return
                
        # Create a progress logging function
        def log_progress(message, step=None, details=None, message_type='info'):
            with task_lock:
                task['log_messages'].append({
                    'message': message,
                    'type': message_type
                })
                if step and details:
                    task['step_details'][str(step)] = details
                print(f"Progress update: {message}")
        
        # Start processing
        log_progress("Initializing OCR engine...", 2, "Initializing")
        
        # Actually process the video
        log_progress("Starting automatic video processing...", 2, "Processing video")
        
        start_time = time.time()
        result, home_appearances, away_appearances = process_video(
            task['filepath'], task['home_team_id'], task['away_team_id'], task['match_day'],
            progress_callback=log_progress
        )
        processing_time = time.time() - start_time
        
        log_progress(f"Processing completed in {processing_time:.1f} seconds", 5, f"Processed in {processing_time:.1f}s")
        
        with task_lock:
            # Check for timeout again
            if check_processing_timeout(task):
                task['status'] = 'error'
                task['error'] = 'Processing timed out after 10 minutes'
                task['log_messages'].append({
                    'message': 'Processing timed out after 10 minutes',
                    'type': 'error'
                })
                print(f"Task {upload_id} timed out")
                return
                
            # Update status to identifying frames
            task['status'] = 'identifying_frames'
            task['current_step'] = 3
            
            # Update status to extracting text
            task['status'] = 'extracting_text'
            task['current_step'] = 4
            
            # Update status to updating database
            task['status'] = 'updating_database'
            task['current_step'] = 5
            
            if "error" in result:
                task['status'] = 'error'
                task['error'] = result['error']
                task['log_messages'].append({
                    'message': f"Error: {result['error']}",
                    'type': 'error'
                })
            else:
                # Store results in task for session retrieval
                task['result'] = result
                task['home_appearances'] = home_appearances
                task['away_appearances'] = away_appearances
                
                # Add summary log
                total_players = len(home_appearances) + len(away_appearances)
                task['log_messages'].append({
                    'message': f"Found {len(home_appearances)} home players and {len(away_appearances)} away players",
                    'type': 'success'
                })
                task['log_messages'].append({
                    'message': f"Successfully updated {total_players} player appearances in database",
                    'type': 'success'
                })
                
                # Mark as complete
                task['status'] = 'complete'
            
    except Exception as e:
        with task_lock:
            task['status'] = 'error'
            task['error'] = str(e)
            task['log_messages'].append({
                'message': f"Error in processing: {str(e)}",
                'type': 'error'
            })
            print(f"Error in processing: {str(e)}")
            print(traceback.format_exc())  # Print full traceback

def process_video_for_review(upload_id):
    """Process the video for review"""
    with task_lock:
        task = processing_tasks[upload_id]
        # Add start time for expiration tracking
        task['start_time'] = time.time()
        # Initialize log messages array
        task['log_messages'] = []
        # Initialize step details
        task['step_details'] = {}
        # Update status to extracting frames
        task['status'] = 'extracting_frames'
        task['current_step'] = 2
        # Add log entry
        task['log_messages'].append({
            'message': f"Starting to process video: {os.path.basename(task['filepath'])}",
            'type': 'info'
        })
    
    try:
        # Check for timeout periodically
        if check_processing_timeout(task):
            with task_lock:
                task['status'] = 'error'
                task['error'] = 'Processing timed out after 10 minutes'
                task['log_messages'].append({
                    'message': 'Processing timed out after 10 minutes',
                    'type': 'error'
                })
            print(f"Task {upload_id} timed out")
            return
        
        # Start frame extraction
        with task_lock:
            task['log_messages'].append({
                'message': "Initializing OCR engine...",
                'type': 'info'
            })
        
        # Create a callback function to track progress
        def progress_callback(message, step=None, details=None, message_type='info'):
            with task_lock:
                task['log_messages'].append({
                    'message': message,
                    'type': message_type
                })
                if step and details:
                    task['step_details'][str(step)] = details
                print(f"Progress update: {message}")
        
        # Actually extract frames for review - pass a context for tracking progress
        progress_callback("Starting frame extraction...", 2, "Initializing extraction")
        session_id, error = extract_frames_for_review(
            task['filepath'], 
            task['home_team_id'], 
            task['away_team_id'], 
            task['match_day'],
            callback=progress_callback
        )
        
        with task_lock:
            # Check for timeout again
            if check_processing_timeout(task):
                task['status'] = 'error'
                task['error'] = 'Processing timed out after 10 minutes'
                task['log_messages'].append({
                    'message': 'Processing timed out after 10 minutes',
                    'type': 'error'
                })
                print(f"Task {upload_id} timed out")
                return
                
            # Update status to identifying frames
            task['status'] = 'identifying_frames'
            task['current_step'] = 3
            task['log_messages'].append({
                'message': "Starting to identify player rating frames...",
                'type': 'info'
            })
            
            # Update status to extracting text
            task['status'] = 'extracting_text'
            task['current_step'] = 4
            task['log_messages'].append({
                'message': "Extracting text from identified frames...",
                'type': 'info'
            })
            
            # Update status to preparing review
            task['status'] = 'preparing_review'
            task['current_step'] = 5
            task['log_messages'].append({
                'message': "Preparing review page with extracted frames...",
                'type': 'info'
            })
            
            if error:
                task['status'] = 'error'
                task['error'] = error
                task['log_messages'].append({
                    'message': f"Error: {error}",
                    'type': 'error'
                })
            else:
                # Store session ID
                task['session_id'] = session_id
                
                # Mark as complete
                task['status'] = 'complete'
                task['log_messages'].append({
                    'message': "Processing completed successfully!",
                    'type': 'success'
                })
            
    except Exception as e:
        with task_lock:
            task['status'] = 'error'
            task['error'] = str(e)
            task['log_messages'].append({
                'message': f"Error in processing: {str(e)}",
                'type': 'error'
            })
            print(f"Error in processing: {str(e)}")
            print(traceback.format_exc())  # Print the full traceback

@app.route('/processing/<upload_id>')
def processing_page(upload_id):
    if upload_id not in processing_tasks:
        flash('Invalid upload ID', 'danger')
        return redirect(url_for('index'))
    
    task = processing_tasks[upload_id]
    
    # Check if task is already complete
    if task['status'] == 'complete':
        if task['process_type'] == 'review':
            return redirect(url_for('review_frames', session_id=task['session_id']))
        else:
            # Create a server-side session to store results
            session['result'] = task['result']
            session['home_appearances'] = task['home_appearances']
            session['away_appearances'] = task['away_appearances']
            return redirect(url_for('results'))
    
    # If there's an error, flash message and redirect
    if task['status'] == 'error':
        flash(f"Error processing video: {task['error']}", 'danger')
        return redirect(url_for('index'))
    
    # Show processing page
    return render_template(
        'processing.html',
        upload_id=upload_id,
        process_type=task['process_type']
    )

# Make sure session is modified when storing results
@app.after_request
def after_request(response):
    # Mark session as modified to ensure it's saved
    if 'result' in session:
        session.modified = True
    return response

@app.route('/upload-status/<upload_id>')
def upload_status(upload_id):
    with task_lock:
        if upload_id not in processing_tasks:
            print(f"Invalid upload ID requested: {upload_id}")
            return jsonify({'error': 'Invalid upload ID'}), 404
        
        task = processing_tasks[upload_id]
        print(f"Status for {upload_id}: {task['status']}, step: {task['current_step']}")
        
        response = {
            'status': task['status'],
            'current_step': task['current_step'],
            'process_type': task['process_type'],
            'timestamp': time.time()
        }
        
        # Add more detailed progress information
        if 'frames_processed' in task:
            response['frames_processed'] = task['frames_processed']
        
        if 'player_frames' in task:
            response['player_frames'] = task['player_frames']
        
        if 'original_frame_count' in task and 'deduplicated_frame_count' in task:
            response['original_frame_count'] = task['original_frame_count']
            response['deduplicated_frame_count'] = task['deduplicated_frame_count']
        
        # Add step-specific details if available
        if 'step_details' in task:
            response['step_details'] = task['step_details']
        
        # Add any log messages
        if 'log_messages' in task:
            response['log_messages'] = task['log_messages']
            # Reset log messages after sending (to avoid duplicates)
            task['log_messages'] = []
        
        if task['status'] == 'complete':
            if task['process_type'] == 'review':
                response['session_id'] = task['session_id']
            print(f"Task complete for {upload_id}, redirecting to {'review' if task['process_type'] == 'review' else 'results'}")
        
        if task['status'] == 'error':
            response['error'] = task['error']
            print(f"Error in task {upload_id}: {task['error']}")
    
    return jsonify(response)

@app.route('/review/<session_id>')
def review_frames(session_id):
    # Check if session exists
    session_dir = os.path.join(app.config['FRAMES_FOLDER'], session_id)
    session_file = os.path.join(session_dir, "session.json")
    
    if not os.path.exists(session_file):
        flash("Review session not found", 'danger')
        return redirect(url_for('index'))
    
    # Load session data
    with open(session_file, "r") as f:
        session_data = json.load(f)
    
    # Validate team IDs
    home_team_id = session_data.get("home_team_id")
    away_team_id = session_data.get("away_team_id")
    
    if not home_team_id or home_team_id == "None":
        flash("Invalid home team ID", 'danger')
        return redirect(url_for('index'))
        
    if not away_team_id or away_team_id == "None":
        flash("Invalid away team ID", 'danger')
        return redirect(url_for('index'))
    
    # Get team data
    try:
        home_team = supabase.table("teams").select("*").eq("id", home_team_id).execute()
        away_team = supabase.table("teams").select("*").eq("id", away_team_id).execute()
    except Exception as e:
        flash(f"Error retrieving team data: {str(e)}", 'danger')
        return redirect(url_for('index'))
    
    home_team_name = home_team.data[0]["name"] if home_team.data else "Unknown Team"
    away_team_name = away_team.data[0]["name"] if away_team.data else "Unknown Team"
    
    # Get players for both teams
    try:
        home_players = get_players_by_team_id(home_team_id)
        away_players = get_players_by_team_id(away_team_id)
    except Exception as e:
        flash(f"Error retrieving player data: {str(e)}", 'danger')
        return redirect(url_for('index'))
        
    # Process frames to better filter OCR results
    for frame in session_data["frames"]:
        # Initialize clean OCR results lists
        frame["filtered_ocr_results"] = []
        
        # Get the OCR data
        for ocr_item in frame["ocr_results"]:
            text = ocr_item["text"]
            confidence = ocr_item["confidence"]
            
            # Skip known UI elements
            text_lower = text.lower()
            if "player ratings" in text_lower or text_lower == "home" or text_lower == "away" or text_lower == "back":
                continue
                
            # Keep valid OCR results
            frame["filtered_ocr_results"].append(ocr_item)
        
        # Process names using improved filtering
        if frame["is_home"]:
            frame["suggested_names"] = process_text_to_extract_names([
                [None, item["text"], item["confidence"]] 
                for item in frame["filtered_ocr_results"]
            ], "home")
        elif frame["is_away"]:
            frame["suggested_names"] = process_text_to_extract_names([
                [None, item["text"], item["confidence"]] 
                for item in frame["filtered_ocr_results"]
            ], "away")
        else:
            frame["suggested_names"] = []
    
    # Count home and away frames
    home_frames = [f for f in session_data["frames"] if f["is_home"]]
    away_frames = [f for f in session_data["frames"] if f["is_away"]]
    
    # Get deduplication info from session data if available
    original_frame_count = session_data.get("original_frame_count", 0)
    current_frame_count = len(session_data["frames"])
    frames_removed = original_frame_count - current_frame_count if original_frame_count > current_frame_count else 0
    deduplication_percent = int((frames_removed / original_frame_count * 100)) if original_frame_count > 0 else 0
    
    # Check if this session is part of a batch
    batch_id = None
    for bid, batch in batch_processing_tasks.items():
        for video in batch.get('videos', []):
            if video.get('session_id') == session_id:
                batch_id = bid
                break
        if batch_id:
            break
    
    return render_template(
        'review_frames.html',
        session_id=session_id,
        session_data=session_data,
        home_team_name=home_team_name,
        away_team_name=away_team_name,
        home_players=home_players,
        away_players=away_players,
        home_frame_count=len(home_frames),
        away_frame_count=len(away_frames),
        original_frame_count=original_frame_count,
        frames_removed=frames_removed,
        deduplication_percent=deduplication_percent,
        batch_id=batch_id  # Pass batch_id to template if this is part of a batch
    )

@app.route('/review/<session_id>/process', methods=['POST'])
def process_review(session_id):
    # Check if session exists
    session_dir = os.path.join(app.config['FRAMES_FOLDER'], session_id)
    session_file = os.path.join(session_dir, "session.json")
    
    if not os.path.exists(session_file):
        return jsonify({"success": False, "error": "Session not found"})
    
    # Load session data
    with open(session_file, "r") as f:
        session_data = json.load(f)
    
    # Get player data from form
    player_data = request.json
    
    if not player_data or not isinstance(player_data, dict):
        return jsonify({"success": False, "error": "Invalid player data"})
    
    # Store edited player names if provided
    if 'edited_frames' in player_data and isinstance(player_data['edited_frames'], dict):
        # Save edited names to session data
        for frame_idx, edited_names in player_data['edited_frames'].items():
            frame_idx = int(frame_idx)
            if frame_idx < len(session_data['frames']):
                session_data['frames'][frame_idx]['edited_names'] = edited_names
        
        # Save updated session data
        with open(session_file, "w") as f:
            json.dump(session_data, f)
    
    # Validate team IDs
    home_team_id = session_data.get("home_team_id")
    away_team_id = session_data.get("away_team_id")
    
    if not home_team_id or home_team_id == "None":
        return jsonify({"success": False, "error": "Invalid home team ID"})
        
    if not away_team_id or away_team_id == "None":
        return jsonify({"success": False, "error": "Invalid away team ID"})
    
    # Create match record
    match_data = {
        "home_team_id": home_team_id,
        "away_team_id": away_team_id,
        "match_day": session_data["match_day"],
        "date": time.strftime("%Y-%m-%d")
    }
    
    try:
        match_result = supabase.table("matches").insert(match_data).execute()
        
        if not match_result.data or len(match_result.data) == 0:
            return jsonify({"success": False, "error": "Failed to create match record"})
        
        match_id = match_result.data[0]["id"]
    except Exception as e:
        return jsonify({"success": False, "error": f"Database error: {str(e)}"})
    
    # Update appearances based on user-confirmed players
    home_appearances = []
    away_appearances = []
    
    # Collect all edited player names for tracking which were unmatched
    all_home_player_names = set()
    all_away_player_names = set()
    
    for frame in session_data["frames"]:
        if frame.get("is_home") and "edited_names" in frame:
            all_home_player_names.update(name for name in frame["edited_names"] if name.strip())
        elif frame.get("is_away") and "edited_names" in frame:
            all_away_player_names.update(name for name in frame["edited_names"] if name.strip())
        elif frame.get("is_home") and "suggested_names" in frame:
            all_home_player_names.update(name for name in frame["suggested_names"] if name.strip())
        elif frame.get("is_away") and "suggested_names" in frame:
            all_away_player_names.update(name for name in frame["suggested_names"] if name.strip())
    
    # Process home players
    if "home_players" in player_data:
        matched_home_players = set()
        for player_id in player_data["home_players"]:
            if player_id and update_player_appearances(player_id, match_id):
                try:
                    # Get player details
                    player = supabase.table("players").select("*").eq("id", player_id).execute()
                    if player.data and len(player.data) > 0:
                        player_name = player.data[0]["name"]
                        matched_home_players.add(player_name.lower())
                        home_appearances.append({
                            "id": player_id,
                            "name": player_name,
                            "extracted_name": player_name  # Using DB name since user confirmed
                        })
                except Exception as e:
                    print(f"Error getting player details for {player_id}: {str(e)}")
                    # Continue processing other players even if one fails
        
        # Store unmatched home players
        for player_name in all_home_player_names:
            # Check if this player name wasn't matched to any database player
            if not any(player_name.lower() in name.lower() or name.lower() in player_name.lower() for name in matched_home_players):
                store_unmatched_player(player_name, home_team_id, match_id)
    
    # Process away players
    if "away_players" in player_data:
        matched_away_players = set()
        for player_id in player_data["away_players"]:
            if player_id and update_player_appearances(player_id, match_id):
                try:
                    # Get player details
                    player = supabase.table("players").select("*").eq("id", player_id).execute()
                    if player.data and len(player.data) > 0:
                        player_name = player.data[0]["name"]
                        matched_away_players.add(player_name.lower())
                        away_appearances.append({
                            "id": player_id,
                            "name": player_name,
                            "extracted_name": player_name  # Using DB name since user confirmed
                        })
                except Exception as e:
                    print(f"Error getting player details for {player_id}: {str(e)}")
                    # Continue processing other players even if one fails
        
        # Store unmatched away players
        for player_name in all_away_player_names:
            # Check if this player name wasn't matched to any database player
            if not any(player_name.lower() in name.lower() or name.lower() in player_name.lower() for name in matched_away_players):
                store_unmatched_player(player_name, away_team_id, match_id)
    
    # Store results in session for results page
    session_obj = {
        "match_id": match_id,
        "home_players_extracted": len(all_home_player_names),
        "away_players_extracted": len(all_away_player_names),
        "home_players_matched": len(home_appearances),
        "away_players_matched": len(away_appearances),
        "home_players_unmatched": len(all_home_player_names) - len(home_appearances),
        "away_players_unmatched": len(all_away_player_names) - len(away_appearances),
    }
    
    # Store results in flask session for display
    session['result'] = session_obj
    session['home_appearances'] = home_appearances
    session['away_appearances'] = away_appearances
    
    # Check if this session is part of a batch
    batch_id = None
    for bid, batch in batch_processing_tasks.items():
        for video in batch.get('videos', []):
            if video.get('session_id') == session_id:
                batch_id = bid
                # Update video status to indicate it's been reviewed
                video['reviewed'] = True
                video['match_id'] = match_id
                video['home_appearances'] = len(home_appearances)
                video['away_appearances'] = len(away_appearances)
                break
        if batch_id:
            break
    
    # Add a parameter to indicate if we should redirect to batch review
    if batch_id and player_data.get('from_batch'):
        return jsonify({
            "success": True, 
            "match_id": match_id,
            "redirect_url": url_for('batch_review', batch_id=batch_id)
        })
    else:
        return jsonify({
            "success": True, 
            "match_id": match_id,
            "redirect_url": url_for('results')
        })

@app.route('/results')
def results():
    # Retrieve results from session
    result = session.get('result', {})
    home_appearances = session.get('home_appearances', [])
    away_appearances = session.get('away_appearances', [])
    
    match_id = result.get('match_id')
    if not match_id:
        flash('No result data found', 'danger')
        return redirect(url_for('index'))
    
    # Get match data
    match = supabase.table("matches").select("*, home_team:home_team_id(name), away_team:away_team_id(name)").eq("id", match_id).execute()
    if not match.data or len(match.data) == 0:
        flash('Match data not found', 'danger')
        return redirect(url_for('index'))
    
    # Get team names
    home_team_name = match.data[0]['home_team']['name'] if 'home_team' in match.data[0] else "Unknown Team"
    away_team_name = match.data[0]['away_team']['name'] if 'away_team' in match.data[0] else "Unknown Team"
    
    # Get complete player data with total appearances
    home_player_ids = [p['id'] for p in home_appearances]
    away_player_ids = [p['id'] for p in away_appearances]
    
    home_players = []
    away_players = []
    
    if home_player_ids:
        home_players_data = supabase.table("players").select("*").in_("id", home_player_ids).execute()
        home_players = home_players_data.data if home_players_data.data else []
    
    if away_player_ids:
        away_players_data = supabase.table("players").select("*").in_("id", away_player_ids).execute()
        away_players = away_players_data.data if away_players_data.data else []
    
    return render_template(
        'results.html', 
        match_id=match_id,
        home_team_name=home_team_name,
        away_team_name=away_team_name,
        match_date=match.data[0].get('date', 'Unknown'),
        match_day=match.data[0].get('match_day', 'Unknown'),
        home_players=home_players,
        away_players=away_players,
        home_player_count=len(home_players),
        away_player_count=len(away_players),
        total_appearances_added=len(home_players) + len(away_players)
    )

@app.route('/players')
def players():
    team_id = request.args.get('team_id')
    if not team_id:
        players_data = supabase.table("players").select("*, team:team_id(name)").order("name").execute()
    else:
        players_data = supabase.table("players").select("*, team:team_id(name)").eq("team_id", team_id).order("name").execute()
    
    teams = supabase.table("teams").select("*").order("name").execute()
    
    return render_template(
        'players.html', 
        players=players_data.data if players_data.data else [],
        teams=teams.data if teams.data else []
    )

@app.route('/matches')
def matches():
    # Get all matches with team information
    matches_data = supabase.table("matches").select("*, home_team:home_team_id(name), away_team:away_team_id(name)").order("date", desc=True).execute()
    
    # Get all teams for the create match day form
    all_teams = supabase.table("teams").select("id, name").order("name").execute()
    
    # If there are matches, get additional information for each
    if matches_data.data:
        for match in matches_data.data:
            # Get appearance counts
            try:
                appearances = supabase.table("appearances").select("*, player:player_id(team_id)").eq("match_id", match["id"]).execute()
                
                # Count by team
                home_count = 0
                away_count = 0
                
                if appearances.data:
                    for appearance in appearances.data:
                        if 'player' in appearance and appearance['player']:
                            team_id = appearance['player'].get('team_id')
                            if team_id == match['home_team_id']:
                                home_count += 1
                            elif team_id == match['away_team_id']:
                                away_count += 1
                
                match["home_appearances"] = home_count
                match["away_appearances"] = away_count
                
                # Get unmatched player counts
                unmatched = supabase.table("unmatched_players").select("team_id").eq("last_match_id", match["id"]).execute()
                
                # Count by team
                home_unmatched = 0
                away_unmatched = 0
                
                if unmatched.data:
                    for player in unmatched.data:
                        team_id = player.get('team_id')
                        if team_id == match['home_team_id']:
                            home_unmatched += 1
                        elif team_id == match['away_team_id']:
                            away_unmatched += 1
                
                match["home_unmatched"] = home_unmatched
                match["away_unmatched"] = away_unmatched
                
            except Exception as e:
                print(f"Error getting stats for match {match['id']}: {str(e)}")
                match["home_appearances"] = 0
                match["away_appearances"] = 0
                match["home_unmatched"] = 0
                match["away_unmatched"] = 0
    
    return render_template('matches.html', 
                           matches=matches_data.data if matches_data.data else [],
                           all_teams=all_teams.data if all_teams.data else [])

@app.route('/match/<match_id>')
def match_details(match_id):
    # Get match info
    match = supabase.table("matches").select("*, home_team:home_team_id(name), away_team:away_team_id(name)").eq("id", match_id).execute()
    
    if not match.data or len(match.data) == 0:
        flash('Match not found', 'danger')
        return redirect(url_for('matches'))
    
    # Get appearances for this match
    appearances = supabase.table("appearances").select("*, player:player_id(name, team_id)").eq("match_id", match_id).execute()
    
    # Get unmatched players for this match
    unmatched_players = supabase.table("unmatched_players").select("*").eq("last_match_id", match_id).execute()
    
    # Separate by team
    home_appearances = []
    away_appearances = []
    home_unmatched = []
    away_unmatched = []
    
    # Process regular appearances
    if appearances.data:
        for appearance in appearances.data:
            if 'player' in appearance and appearance['player']:
                team_id = appearance['player'].get('team_id')
                if team_id == match.data[0]['home_team_id']:
                    home_appearances.append(appearance)
                elif team_id == match.data[0]['away_team_id']:
                    away_appearances.append(appearance)
    
    # Process unmatched players
    if unmatched_players.data:
        for player in unmatched_players.data:
            team_id = player.get('team_id')
            if team_id == match.data[0]['home_team_id']:
                home_unmatched.append(player)
            elif team_id == match.data[0]['away_team_id']:
                away_unmatched.append(player)
    
    return render_template(
        'match_details.html', 
        match=match.data[0], 
        home_appearances=home_appearances,
        away_appearances=away_appearances,
        home_unmatched=home_unmatched,
        away_unmatched=away_unmatched
    )

# Add a manual cleanup route for admins
@app.route('/admin/cleanup', methods=['GET'])
def admin_cleanup():
    uploads_cleaned, frames_cleaned = file_manager.cleanup_old_files(
        app.config['UPLOAD_FOLDER'],
        app.config['FRAMES_FOLDER']
    )
    
    return jsonify({
        'success': True,
        'cleaned_uploads': uploads_cleaned,
        'cleaned_frames': frames_cleaned
    })

# Add a debug route to help diagnose issues
@app.route('/debug/session/<session_id>')
def debug_session(session_id):
    if not app.debug:
        return jsonify({"error": "Debug routes only available in debug mode"}), 403
        
    # Check if session exists
    session_dir = os.path.join(app.config['FRAMES_FOLDER'], session_id)
    session_file = os.path.join(session_dir, "session.json")
    
    if not os.path.exists(session_file):
        return jsonify({"error": "Session not found"}), 404
    
    # Load session data
    with open(session_file, "r") as f:
        session_data = json.load(f)
    
    # Get frame paths
    frames_info = []
    for i, frame in enumerate(session_data.get("frames", [])):
        frames_info.append({
            "index": i,
            "path": frame.get("path", "None"),
            "is_home": frame.get("is_home", False),
            "is_away": frame.get("is_away", False),
            "ocr_count": len(frame.get("ocr_results", [])),
            "suggested_names": frame.get("suggested_names", [])
        })
    
    # Return diagnostic info
    return jsonify({
        "session_id": session_id,
        "home_team_id": session_data.get("home_team_id"),
        "away_team_id": session_data.get("away_team_id"),
        "match_day": session_data.get("match_day"),
        "video_path": session_data.get("video_path"),
        "frame_count": len(session_data.get("frames", [])),
        "frames_info": frames_info
    })

@app.route('/export/teams/excel')
def export_teams_excel():
    """Initiate Excel export process"""
    # Generate a unique ID for this export task
    export_id = str(uuid.uuid4())
    
    # Initialize export status
    with task_lock:
        processing_tasks[export_id] = {
            'status': 'starting',
            'current_step': 1,
            'step_details': {},
            'log_messages': [],
            'start_time': time.time(),
            'task_type': 'excel_export',
            'progress': 0,
            'teams_processed': 0,
            'total_teams': 0
        }
    
    # Start background thread for processing
    thread = threading.Thread(target=generate_excel_export, args=(export_id,))
    thread.daemon = True
    thread.start()
    
    # Redirect to processing page
    return redirect(url_for('export_progress', export_id=export_id))

@app.route('/export/progress/<export_id>')
def export_progress(export_id):
    """Show progress page for Excel export"""
    if export_id not in processing_tasks:
        flash('Invalid export ID', 'danger')
        return redirect(url_for('index'))
    
    task = processing_tasks[export_id]
    
    # Check if task is already complete
    if task['status'] == 'complete':
        return redirect(url_for('download_excel', export_id=export_id))
    
    # If there's an error, flash message and redirect
    if task['status'] == 'error':
        flash(f"Error generating Excel file: {task.get('error', 'Unknown error')}", 'danger')
        return redirect(url_for('index'))
    
    # Show processing page
    return render_template(
        'export_progress.html',
        export_id=export_id
    )

@app.route('/export/status/<export_id>')
def export_status(export_id):
    """API endpoint to get export status"""
    with task_lock:
        if export_id not in processing_tasks:
            return jsonify({'error': 'Invalid export ID'}), 404
        
        task = processing_tasks[export_id]
        
        response = {
            'status': task['status'],
            'progress': task.get('progress', 0),
            'teams_processed': task.get('teams_processed', 0),
            'total_teams': task.get('total_teams', 0),
            'timestamp': time.time()
        }
        
        # Add step-specific details if available
        if 'step_details' in task:
            response['step_details'] = task['step_details']
        
        # Add any log messages
        if 'log_messages' in task:
            response['log_messages'] = task['log_messages']
            # Reset log messages after sending (to avoid duplicates)
            task['log_messages'] = []
        
        if task['status'] == 'complete':
            response['download_url'] = url_for('download_excel', export_id=export_id)
        
        if task['status'] == 'error':
            response['error'] = task.get('error', 'Unknown error')
    
    return jsonify(response)

@app.route('/export/download/<export_id>')
def download_excel(export_id):
    """Download the generated Excel file"""
    if export_id not in processing_tasks:
        flash('Invalid export ID', 'danger')
        return redirect(url_for('index'))
    
    task = processing_tasks[export_id]
    
    if task['status'] != 'complete' or 'file_path' not in task:
        flash('Excel file not ready for download', 'warning')
        return redirect(url_for('export_progress', export_id=export_id))
    
    # Generate a filename with date
    current_date = time.strftime("%Y%m%d")
    filename = f"player_appearances_{current_date}.xlsx"
    
    # Send the file
    return send_file(
        task['file_path'],
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

def generate_excel_export(export_id):
    """Background task to generate Excel export"""
    with task_lock:
        task = processing_tasks[export_id]
        # Add log entry
        task['log_messages'].append({
            'message': "Starting Excel export process",
            'type': 'info'
        })
    
    try:
        # Create a progress logging function
        def log_progress(message, progress=None, step=None, details=None, message_type='info'):
            with task_lock:
                task = processing_tasks[export_id]
                task['log_messages'].append({
                    'message': message,
                    'type': message_type
                })
                if progress is not None:
                    task['progress'] = progress
                if step and details:
                    task['step_details'][str(step)] = details
                print(f"Export progress: {message}")
        
        # Create a new workbook
        wb = Workbook()
        
        log_progress("Fetching team data from database", 5, 1, "Fetching teams")
        
        # Get all teams
        teams = supabase.table("teams").select("*").order("name").execute()
        
        if not teams.data:
            with task_lock:
                task['status'] = 'error'
                task['error'] = 'No teams found to export'
            return
        
        # Set total teams count
        with task_lock:
            task['total_teams'] = len(teams.data)
        
        log_progress(f"Found {len(teams.data)} teams", 10, 1, f"Found {len(teams.data)} teams")
        
        # Get all matches
        log_progress("Fetching match data", 15, 2, "Fetching matches")
        matches = supabase.table("matches").select("id, match_day, date").order("date").execute()
        match_days = []
        match_map = {}
        
        if matches.data:
            # Create a mapping of match IDs to match days
            for match in matches.data:
                match_id = match["id"]
                match_day = match["match_day"]
                match_map[match_id] = match_day
                if match_day not in match_days:
                    match_days.append(match_day)
            
            log_progress(f"Found {len(matches.data)} matches across {len(match_days)} match days", 
                        20, 2, f"Found {len(matches.data)} matches")
        else:
            log_progress("No matches found", 20, 2, "No matches found", "warning")
        
        # Sort match days
        match_days.sort()
        
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        
        # Define styles
        log_progress("Setting up Excel styles and formats", 25, 3, "Setting up styles")
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        subheader_font = Font(bold=True, size=11)
        subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        unmatched_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # Create a summary sheet first
        log_progress("Creating summary sheet", 30, 4, "Creating summary")
        summary = wb.create_sheet(title="Summary", index=0)
        summary['A1'] = "Teams and Players Summary"
        summary['A1'].font = Font(bold=True, size=14)
        summary.merge_cells('A1:D1')
        
        summary['A3'] = "Team Name"
        summary['A3'].font = header_font
        summary['A3'].fill = header_fill
        summary['A3'].alignment = header_alignment
        summary['A3'].border = thin_border
        
        summary['B3'] = "Players"
        summary['B3'].font = header_font
        summary['B3'].fill = header_fill
        summary['B3'].alignment = header_alignment
        summary['B3'].border = thin_border
        
        summary['C3'] = "Unmatched Names"
        summary['C3'].font = header_font
        summary['C3'].fill = header_fill
        summary['C3'].alignment = header_alignment
        summary['C3'].border = thin_border
        
        summary['D3'] = "Total Appearances"
        summary['D3'].font = header_font
        summary['D3'].fill = header_fill
        summary['D3'].alignment = header_alignment
        summary['D3'].border = thin_border
        
        # Set column widths for summary
        summary.column_dimensions['A'].width = 30
        summary.column_dimensions['B'].width = 15
        summary.column_dimensions['C'].width = 20
        summary.column_dimensions['D'].width = 20
        
        # Process each team
        log_progress("Starting to process individual team data", 35, 5, "Processing teams")
        
        total_players = 0
        total_unmatched = 0
        total_appearances = 0
        summary_row_index = 4
        
        # Calculate base progress percentage and increment per team
        base_progress = 35
        progress_per_team = 55 / len(teams.data)  # 55% of progress bar allocated to team processing
        
        for team_index, team in enumerate(teams.data):
            team_id = team["id"]
            team_name = team["name"]
            
            current_progress = base_progress + (team_index * progress_per_team)
            log_progress(f"Processing team {team_index+1}/{len(teams.data)}: {team_name}", 
                        int(current_progress), 5, f"Team {team_index+1}/{len(teams.data)}")
            
            # Update teams processed counter
            with task_lock:
                task['teams_processed'] = team_index + 1
            
            # Create a sheet for this team
            # Ensure sheet name is valid (max 31 chars, no special chars)
            sheet_name = team_name[:31].replace('/', '_').replace('\\', '_').replace('?', '_').replace('*', '_').replace('[', '_').replace(']', '_').replace(':', '_')
            ws = wb.create_sheet(title=sheet_name)
            
            # Set column widths
            ws.column_dimensions['A'].width = 30  # Player name
            ws.column_dimensions['B'].width = 15  # Total appearances
            
            # Add headers
            ws['A1'] = team_name
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:E1')
            
            ws['A3'] = "Player Name"
            ws['A3'].font = header_font
            ws['A3'].fill = header_fill
            ws['A3'].alignment = header_alignment
            ws['A3'].border = thin_border
            
            ws['B3'] = "Total Appearances"
            ws['B3'].font = header_font
            ws['B3'].fill = header_fill
            ws['B3'].alignment = header_alignment
            ws['B3'].border = thin_border
            
            # Add match day columns
            col_index = 3  # Start from column C
            for match_day in match_days:
                col_letter = get_column_letter(col_index)
                ws[f'{col_letter}3'] = match_day
                ws[f'{col_letter}3'].font = header_font
                ws[f'{col_letter}3'].fill = header_fill
                ws[f'{col_letter}3'].alignment = header_alignment
                ws[f'{col_letter}3'].border = thin_border
                ws.column_dimensions[col_letter].width = 12
                col_index += 1
            
            # Get players for this team
            players = supabase.table("players").select("*").eq("team_id", team_id).order("name").execute()
            player_count = len(players.data) if players.data else 0
            total_players += player_count
            
            row_index = 4  # Start from row 4 for player data
            
            # Add player data
            if players.data:
                for player in players.data:
                    player_id = player["id"]
                    player_name = player["name"]
                    
                    # Add player name
                    ws[f'A{row_index}'] = player_name
                    ws[f'A{row_index}'].border = thin_border
                    
                    # Get player appearances
                    appearances = supabase.table("appearances").select("match_id").eq("player_id", player_id).execute()
                    
                    # Count total appearances
                    total_appearances_for_player = len(appearances.data) if appearances.data else 0
                    total_appearances += total_appearances_for_player
                    
                    ws[f'B{row_index}'] = total_appearances_for_player
                    ws[f'B{row_index}'].alignment = Alignment(horizontal="center")
                    ws[f'B{row_index}'].border = thin_border
                    
                    # Mark appearances by match day
                    if appearances.data:
                        col_index = 3  # Start from column C
                        for match_day in match_days:
                            col_letter = get_column_letter(col_index)
                            
                            # Check if player appeared in any match with this match day
                            appeared = False
                            for appearance in appearances.data:
                                match_id = appearance["match_id"]
                                if match_id in match_map and match_map[match_id] == match_day:
                                    appeared = True
                                    break
                            
                            if appeared:
                                ws[f'{col_letter}{row_index}'] = 1
                                ws[f'{col_letter}{row_index}'].alignment = Alignment(horizontal="center")
                            else:
                                ws[f'{col_letter}{row_index}'] = 0
                                ws[f'{col_letter}{row_index}'].alignment = Alignment(horizontal="center")
                            
                            ws[f'{col_letter}{row_index}'].border = thin_border
                            col_index += 1
                    
                    row_index += 1
            
            # Add a separator
            row_index += 1
            ws[f'A{row_index}'] = "Unmatched Player Names"
            ws[f'A{row_index}'].font = subheader_font
            ws[f'A{row_index}'].fill = subheader_fill
            ws.merge_cells(f'A{row_index}:E{row_index}')
            
            # Get unmatched players for this team
            unmatched_players = supabase.table("unmatched_players").select("*").eq("team_id", team_id).order("occurrence_count", desc=True).execute()
            unmatched_count = len(unmatched_players.data) if unmatched_players.data else 0
            total_unmatched += unmatched_count
            
            row_index += 1
            ws[f'A{row_index}'] = "Player Name"
            ws[f'A{row_index}'].font = subheader_font
            ws[f'A{row_index}'].fill = unmatched_fill
            ws[f'A{row_index}'].border = thin_border
            
            ws[f'B{row_index}'] = "Occurrences"
            ws[f'B{row_index}'].font = subheader_font
            ws[f'B{row_index}'].fill = unmatched_fill
            ws[f'B{row_index}'].alignment = Alignment(horizontal="center")
            ws[f'B{row_index}'].border = thin_border
            
            ws[f'C{row_index}'] = "First Seen"
            ws[f'C{row_index}'].font = subheader_font
            ws[f'C{row_index}'].fill = unmatched_fill
            ws[f'C{row_index}'].alignment = Alignment(horizontal="center")
            ws[f'C{row_index}'].border = thin_border
            
            ws[f'D{row_index}'] = "Last Seen"
            ws[f'D{row_index}'].font = subheader_font
            ws[f'D{row_index}'].fill = unmatched_fill
            ws[f'D{row_index}'].alignment = Alignment(horizontal="center")
            ws[f'D{row_index}'].border = thin_border
            
            ws[f'E{row_index}'] = "Last Match"
            ws[f'E{row_index}'].font = subheader_font
            ws[f'E{row_index}'].fill = unmatched_fill
            ws[f'E{row_index}'].alignment = Alignment(horizontal="center")
            ws[f'E{row_index}'].border = thin_border
            
            row_index += 1
            
            # Add unmatched player data
            if unmatched_players.data:
                for player in unmatched_players.data:
                    player_name = player["name"]
                    occurrences = player.get("occurrence_count", 1)
                    first_seen = player.get("first_seen", "")
                    last_seen = player.get("last_seen", "")
                    last_match_id = player.get("last_match_id", "")
                    last_match_day = match_map.get(last_match_id, "") if last_match_id else ""
                    
                    ws[f'A{row_index}'] = player_name
                    ws[f'A{row_index}'].border = thin_border
                    
                    ws[f'B{row_index}'] = occurrences
                    ws[f'B{row_index}'].alignment = Alignment(horizontal="center")
                    ws[f'B{row_index}'].border = thin_border
                    
                    ws[f'C{row_index}'] = first_seen
                    ws[f'C{row_index}'].alignment = Alignment(horizontal="center")
                    ws[f'C{row_index}'].border = thin_border
                    
                    ws[f'D{row_index}'] = last_seen
                    ws[f'D{row_index}'].alignment = Alignment(horizontal="center")
                    ws[f'D{row_index}'].border = thin_border
                    
                    ws[f'E{row_index}'] = last_match_day
                    ws[f'E{row_index}'].alignment = Alignment(horizontal="center")
                    ws[f'E{row_index}'].border = thin_border
                    
                    row_index += 1
            else:
                ws[f'A{row_index}'] = "No unmatched player names found"
                ws.merge_cells(f'A{row_index}:E{row_index}')
                ws[f'A{row_index}'].alignment = Alignment(horizontal="center")
                row_index += 1
            
            # Add team to summary sheet
            summary[f'A{summary_row_index}'] = team_name
            summary[f'A{summary_row_index}'].border = thin_border
            
            summary[f'B{summary_row_index}'] = player_count
            summary[f'B{summary_row_index}'].alignment = Alignment(horizontal="center")
            summary[f'B{summary_row_index}'].border = thin_border
            
            summary[f'C{summary_row_index}'] = unmatched_count
            summary[f'C{summary_row_index}'].alignment = Alignment(horizontal="center")
            summary[f'C{summary_row_index}'].border = thin_border
            
            # Count appearances for this team
            try:
                # The previous query was incorrect: appearances.player:player_id(team_id) doesn't exist
                # We need to get appearances via player IDs for this team
                player_ids = [p["id"] for p in players.data] if players.data else []
                appearance_count = 0
                
                if player_ids:
                    # Get appearances for all players in this team
                    appearances_result = supabase.table("appearances").select("id").in_("player_id", player_ids).execute()
                    appearance_count = len(appearances_result.data) if appearances_result.data else 0
            except Exception as e:
                log_progress(f"Error counting appearances for team {team_name}: {str(e)}", None, None, None, "error")
                appearance_count = 0
            
            summary[f'D{summary_row_index}'] = appearance_count
            summary[f'D{summary_row_index}'].alignment = Alignment(horizontal="center")
            summary[f'D{summary_row_index}'].border = thin_border
            
            summary_row_index += 1
        
        # Add totals to summary sheet
        log_progress("Finalizing summary sheet", 90, 6, "Finalizing summary")
        
        summary[f'A{summary_row_index}'] = "TOTAL"
        summary[f'A{summary_row_index}'].font = Font(bold=True)
        summary[f'A{summary_row_index}'].border = thin_border
        
        summary[f'B{summary_row_index}'] = total_players
        summary[f'B{summary_row_index}'].font = Font(bold=True)
        summary[f'B{summary_row_index}'].alignment = Alignment(horizontal="center")
        summary[f'B{summary_row_index}'].border = thin_border
        
        summary[f'C{summary_row_index}'] = total_unmatched
        summary[f'C{summary_row_index}'].font = Font(bold=True)
        summary[f'C{summary_row_index}'].alignment = Alignment(horizontal="center")
        summary[f'C{summary_row_index}'].border = thin_border
        
        # Calculate total appearances from all teams
        try:
            # Get total appearances from database - use a different approach
            # The count property might not be accessible as expected
            total_appearances_query = supabase.table("appearances").select("id").execute()
            if total_appearances_query.data:
                # Just count the total number of records returned
                total_appearances = len(total_appearances_query.data)
            else:
                # Fallback to our accumulated count
                log_progress("No appearances found in database, using accumulated count", None, None, None, "warning")
        except Exception as e:
            log_progress(f"Error counting total appearances: {str(e)}", None, None, None, "error")
            # Use the sum we've been accumulating if the query fails
        
        summary[f'D{summary_row_index}'] = total_appearances
        summary[f'D{summary_row_index}'].font = Font(bold=True)
        summary[f'D{summary_row_index}'].alignment = Alignment(horizontal="center")
        summary[f'D{summary_row_index}'].border = thin_border
        
        summary_row_index += 1
        
        # Save to a temporary file
        log_progress("Saving Excel file", 95, 7, "Saving file")
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        # Mark as complete
        with task_lock:
            task = processing_tasks[export_id]
            task['status'] = 'complete'
            task['file_path'] = temp_file.name
            task['progress'] = 100
            task['log_messages'].append({
                'message': "Excel export completed successfully!",
                'type': 'success'
            })
        
        log_progress("Excel export completed successfully!", 100, 7, "Complete", "success")
        
    except Exception as e:
        print(f"Error exporting Excel: {str(e)}")
        print(traceback.format_exc())
        
        with task_lock:
            task = processing_tasks[export_id]
            task['status'] = 'error'
            task['error'] = str(e)
            task['log_messages'].append({
                'message': f"Error exporting Excel: {str(e)}",
                'type': 'error'
            })

def process_video_batch(batch_id):
    """Process a batch of videos in parallel"""
    with batch_lock:
        batch_info = batch_processing_tasks[batch_id]
        # Initialize the batch
        batch_info['start_time'] = time.time()
        batch_info['status'] = 'processing'
        batch_info['completed_videos'] = 0
        batch_info['total_videos'] = len(batch_info['videos'])
        batch_info['results'] = {}
        batch_info['log_messages'] = [{
            'message': f"Starting batch processing of {len(batch_info['videos'])} videos",
            'type': 'info',
            'timestamp': time.time()
        }]
    
    try:
        # Create a list to track futures
        futures = []
        
        # Submit each video for processing
        for video_info in batch_info['videos']:
            # Create a future for this video
            future = thread_pool_executor.submit(
                process_single_video_in_batch,
                batch_id,
                video_info['id'],
                video_info['filepath'],
                video_info['home_team_id'],
                video_info['away_team_id'],
                video_info['match_day']
            )
            futures.append(future)
        
        # Wait for all futures to complete
        for future in concurrent.futures.as_completed(futures):
            try:
                # Get the result
                result = future.result()
                print(f"Completed processing video: {result}")
                
                # Update batch status
                with batch_lock:
                    batch_info['completed_videos'] += 1
                    progress = (batch_info['completed_videos'] / batch_info['total_videos']) * 100
                    
                    # Add a log message
                    batch_info['log_messages'].append({
                        'message': f"Completed {batch_info['completed_videos']} of {batch_info['total_videos']} videos ({progress:.1f}%)",
                        'type': 'info',
                        'timestamp': time.time()
                    })
                    
                    # Check if all videos are processed
                    if batch_info['completed_videos'] >= batch_info['total_videos']:
                        batch_info['status'] = 'complete'
                        batch_info['end_time'] = time.time()
                        processing_time = batch_info['end_time'] - batch_info['start_time']
                        batch_info['log_messages'].append({
                            'message': f"Batch processing completed in {processing_time:.1f} seconds",
                            'type': 'success',
                            'timestamp': time.time()
                        })
            except Exception as e:
                print(f"Error processing video in batch: {str(e)}")
                print(traceback.format_exc())
                
                # Update batch with error
                with batch_lock:
                    batch_info['completed_videos'] += 1
                    batch_info['log_messages'].append({
                        'message': f"Error processing a video: {str(e)}",
                        'type': 'error',
                        'timestamp': time.time()
                    })
                    
                    # Check if all videos are processed despite errors
                    if batch_info['completed_videos'] >= batch_info['total_videos']:
                        batch_info['status'] = 'complete_with_errors'
                        batch_info['end_time'] = time.time()
    
    except Exception as e:
        print(f"Error in batch processing: {str(e)}")
        print(traceback.format_exc())
        
        # Update batch with error
        with batch_lock:
            batch_info['status'] = 'error'
            batch_info['error'] = str(e)
            batch_info['log_messages'].append({
                'message': f"Fatal error in batch processing: {str(e)}",
                'type': 'error',
                'timestamp': time.time()
            })

def check_batch_video_timeout(video_info, timeout_minutes=20):
    """Check if a video in a batch has timed out"""
    if 'start_time' not in video_info:
        video_info['start_time'] = time.time()
        
    elapsed_time = time.time() - video_info['start_time']
    timeout_seconds = timeout_minutes * 60
    
    if elapsed_time > timeout_seconds:
        return True
    return False

def process_single_video_in_batch(batch_id, video_id, filepath, home_team_id, away_team_id, match_day):
    """Process a single video as part of a batch"""
    try:
        # Log start of processing
        with batch_lock:
            batch_info = batch_processing_tasks[batch_id]
            video_info = next((v for v in batch_info['videos'] if v['id'] == video_id), None)
            if video_info:
                video_info['status'] = 'processing'
                video_info['start_time'] = time.time()
                batch_info['log_messages'].append({
                    'message': f"Processing video: {os.path.basename(filepath)}",
                    'type': 'info',
                    'timestamp': time.time(),
                    'video_id': video_id
                })
        
        # Check for timeout periodically
        if check_batch_video_timeout(video_info, timeout_minutes=20):
            with batch_lock:
                video_info['status'] = 'error'
                video_info['error'] = 'Processing timed out after 20 minutes'
                batch_info['log_messages'].append({
                    'message': f"Processing of {os.path.basename(filepath)} timed out after 20 minutes",
                    'type': 'error',
                    'timestamp': time.time(),
                    'video_id': video_id
                })
            print(f"Video {video_id} in batch {batch_id} timed out")
            return {
                'video_id': video_id,
                'status': 'error',
                'filepath': filepath,
                'error': 'Timeout'
            }
        
        # Process the video
        def progress_callback(message, step=None, details=None, message_type='info'):
            with batch_lock:
                batch_info = batch_processing_tasks[batch_id]
                batch_info['log_messages'].append({
                    'message': message,
                    'type': message_type,
                    'timestamp': time.time(),
                    'video_id': video_id,
                    'step': step,
                    'details': details
                })
        
        # Actually process the video
        start_time = time.time()
        result, home_appearances, away_appearances = process_video(
            filepath, home_team_id, away_team_id, match_day,
            progress_callback=progress_callback
        )
        processing_time = time.time() - start_time
        
        # Update batch info with results
        with batch_lock:
            batch_info = batch_processing_tasks[batch_id]
            video_info = next((v for v in batch_info['videos'] if v['id'] == video_id), None)
            
            if video_info:
                video_info['end_time'] = time.time()
                video_info['processing_time'] = processing_time
                
                if "error" in result:
                    video_info['status'] = 'error'
                    video_info['error'] = result['error']
                    batch_info['log_messages'].append({
                        'message': f"Error processing {os.path.basename(filepath)}: {result['error']}",
                        'type': 'error',
                        'timestamp': time.time(),
                        'video_id': video_id
                    })
                else:
                    video_info['status'] = 'complete'
                    video_info['result'] = result
                    video_info['home_appearances'] = home_appearances
                    video_info['away_appearances'] = away_appearances
                    
                    total_players = len(home_appearances) + len(away_appearances)
                    batch_info['log_messages'].append({
                        'message': f"Successfully processed {os.path.basename(filepath)} in {processing_time:.1f}s. Found {total_players} players.",
                        'type': 'success',
                        'timestamp': time.time(),
                        'video_id': video_id
                    })
        
        return {
            'video_id': video_id,
            'status': 'complete' if "error" not in result else 'error',
            'filepath': filepath
        }
        
    except Exception as e:
        print(f"Error processing video {video_id} in batch {batch_id}: {str(e)}")
        print(traceback.format_exc())
        
        # Update batch info with error
        with batch_lock:
            batch_info = batch_processing_tasks[batch_id]
            video_info = next((v for v in batch_info['videos'] if v['id'] == video_id), None)
            
            if video_info:
                video_info['status'] = 'error'
                video_info['error'] = str(e)
                video_info['end_time'] = time.time()
                batch_info['log_messages'].append({
                    'message': f"Error processing {os.path.basename(filepath)}: {str(e)}",
                    'type': 'error',
                    'timestamp': time.time(),
                    'video_id': video_id
                })
        
        # Re-raise to be caught by the batch processor
        raise

@app.route('/upload_multiple_videos', methods=['POST'])
def upload_multiple_videos():
    if 'videos[]' not in request.files:
        flash('No video files uploaded', 'danger')
        return redirect(url_for('index'))
    
    files = request.files.getlist('videos[]')
    
    if not files or len(files) == 0:
        flash('No videos selected', 'danger')
        return redirect(url_for('index'))
    
    # Get team IDs for each video
    team_data = json.loads(request.form.get('team_data', '[]'))
    
    # Get processing type (auto or review)
    process_type = request.form.get('process_type', 'auto')
    
    # Validate team data format
    if not team_data or len(team_data) != len(files):
        flash('Invalid team data provided. Each video must have team information.', 'danger')
        return redirect(url_for('index'))
    
    # Create batch ID
    batch_id = str(uuid.uuid4())
    
    # Initialize batch processing info
    batch_processing_tasks[batch_id] = {
        'id': batch_id,
        'status': 'uploading',
        'start_time': time.time(),
        'process_type': process_type,
        'videos': [],
        'log_messages': [{
            'message': f"Uploading {len(files)} videos",
            'type': 'info',
            'timestamp': time.time()
        }]
    }
    
    # Process each file
    for i, file in enumerate(files):
        if file.filename == '':
            continue
            
        if not allowed_file(file.filename):
            # Log error but continue with other files
            with batch_lock:
                batch_processing_tasks[batch_id]['log_messages'].append({
                    'message': f"Invalid file type for {file.filename}. Skipping.",
                    'type': 'warning',
                    'timestamp': time.time()
                })
            continue
        
        # Get team IDs from team_data
        video_data = team_data[i]
        home_team_id = video_data.get('home_team')
        away_team_id = video_data.get('away_team')
        match_day = video_data.get('match_day', 'MD1')
        
        # Validate team IDs
        if not home_team_id or home_team_id == "None" or not away_team_id or away_team_id == "None":
            # Log error but continue with other files
            with batch_lock:
                batch_processing_tasks[batch_id]['log_messages'].append({
                    'message': f"Invalid team data for {file.filename}. Skipping.",
                    'type': 'warning',
                    'timestamp': time.time()
                })
            continue
        
        # Save uploaded file
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # Generate a unique ID for this video
        video_id = str(uuid.uuid4())
        
        # Add to batch videos
        with batch_lock:
            batch_processing_tasks[batch_id]['videos'].append({
                'id': video_id,
                'filename': filename,
                'filepath': filepath,
                'home_team_id': home_team_id,
                'away_team_id': away_team_id,
                'match_day': match_day,
                'status': 'uploaded',
                'session_id': None  # Will store session ID for review mode
            })
            
            batch_processing_tasks[batch_id]['log_messages'].append({
                'message': f"Uploaded video: {filename}",
                'type': 'info',
                'timestamp': time.time(),
                'video_id': video_id
            })
    
    # Start batch processing in a separate thread based on process type
    if process_type == 'review':
        threading.Thread(target=process_video_batch_for_review, args=(batch_id,), daemon=True).start()
    else:
        threading.Thread(target=process_video_batch, args=(batch_id,), daemon=True).start()
    
    # Redirect to batch processing page
    return redirect(url_for('batch_processing_page', batch_id=batch_id))

@app.route('/batch-processing/<batch_id>')
def batch_processing_page(batch_id):
    # Check if batch exists
    if batch_id not in batch_processing_tasks:
        flash('Batch processing not found', 'danger')
        return redirect(url_for('index'))
        
    return render_template('batch_processing.html', batch_id=batch_id)

@app.route('/batch-status/<batch_id>')
def batch_status(batch_id):
    """Get the status of a batch processing task"""
    if batch_id not in batch_processing_tasks:
        return jsonify({'error': 'Batch not found'}), 404
    
    with batch_lock:
        batch_info = batch_processing_tasks[batch_id]
        
        # Calculate progress
        total_videos = len(batch_info['videos'])
        completed_videos = sum(1 for video in batch_info['videos'] if video['status'] in ['complete', 'error'])
        progress = (completed_videos / total_videos * 100) if total_videos > 0 else 0
        
        # Use log_messages instead of logs (which doesn't exist)
        return jsonify({
            'id': batch_id,
            'status': batch_info['status'],
            'process_type': batch_info.get('process_type', 'auto'),  # Include process_type in response
            'progress': progress,
            'total_videos': total_videos,
            'completed_videos': completed_videos,
            'videos': batch_info['videos'],
            'logs': batch_info['log_messages'][-50:] if 'log_messages' in batch_info else [],  # Return only the last 50 log entries
            'start_time': batch_info['start_time']
        })

@app.route('/batch-results/<batch_id>')
def batch_results(batch_id):
    # Check if batch exists
    if batch_id not in batch_processing_tasks:
        flash('Batch processing not found', 'danger')
        return redirect(url_for('index'))
    
    # Get batch data
    with batch_lock:
        batch_info = batch_processing_tasks[batch_id]
        
        # Check if processing is complete
        if batch_info['status'] not in ['complete', 'complete_with_errors']:
            flash('Batch processing is still in progress', 'warning')
            return redirect(url_for('batch_processing_page', batch_id=batch_id))
        
        # Check if this is a review batch
        if batch_info.get('process_type') == 'review':
            return redirect(url_for('batch_review', batch_id=batch_id))
        
        # Collect results for display
        results = []
        for video in batch_info['videos']:
            if video['status'] == 'complete':
                # Get team names
                home_team = supabase.table("teams").select("name").eq("id", video['home_team_id']).execute().data[0]['name'] if supabase.table("teams").select("name").eq("id", video['home_team_id']).execute().data else "Unknown"
                away_team = supabase.table("teams").select("name").eq("id", video['away_team_id']).execute().data[0]['name'] if supabase.table("teams").select("name").eq("id", video['away_team_id']).execute().data else "Unknown"
                
                results.append({
                    'filename': video['filename'],
                    'home_team': home_team,
                    'away_team': away_team,
                    'match_day': video['match_day'],
                    'processing_time': video.get('processing_time', 0),
                    'home_players_matched': len(video.get('home_appearances', [])),
                    'away_players_matched': len(video.get('away_appearances', [])),
                    'home_appearances': video.get('home_appearances', []),
                    'away_appearances': video.get('away_appearances', []),
                    'match_id': video.get('result', {}).get('match_id', None)
                })
    
    return render_template('batch_results.html', batch_id=batch_id, results=results)

@app.route('/batch-review/<batch_id>')
def batch_review(batch_id):
    # Find the batch with the given ID
    batch = None
    for task_id, task in batch_processing_tasks.items():
        if task_id == batch_id:
            batch = task
            break
    
    if not batch:
        flash('Batch not found', 'error')
        return redirect(url_for('index'))
    
    # Get all sessions associated with this batch
    batch_sessions = []
    for video_info in batch.get('videos', []):
        session_id = video_info.get('session_id')
        if session_id:
            try:
                # Check if session exists by looking for session file
                session_dir = os.path.join(app.config['FRAMES_FOLDER'], session_id)
                session_file = os.path.join(session_dir, "session.json")
                
                if os.path.exists(session_file):
                    # Load session data
                    with open(session_file, "r") as f:
                        session_data = json.load(f)
                    
                    # Get team names from Supabase
                    home_team = supabase.table("teams").select("name").eq("id", video_info.get('home_team_id')).execute()
                    away_team = supabase.table("teams").select("name").eq("id", video_info.get('away_team_id')).execute()
                    
                    home_team_name = home_team.data[0]['name'] if home_team.data and len(home_team.data) > 0 else 'Unknown'
                    away_team_name = away_team.data[0]['name'] if away_team.data and len(away_team.data) > 0 else 'Unknown'
                    
                    # Determine video status
                    status = "Processed" if video_info.get('status') == 'complete' else \
                             "Failed" if video_info.get('status') == 'error' else "Pending"
                    
                    # Create session data object
                    session_data = {
                        'id': session_id,
                        'video_id': video_info.get('id'),
                        'filename': os.path.basename(video_info.get('filepath', '')),
                        'home_team': home_team_name,
                        'away_team': away_team_name,
                        'match_day': video_info.get('match_day', 'Unknown'),
                        'status': status,
                        'teams': f"{home_team_name} {away_team_name}" # For search functionality
                    }
                    batch_sessions.append(session_data)
            except Exception as e:
                print(f"Error loading session {session_id}: {str(e)}")
                # Continue with other sessions
    
    # Calculate counts for status summary
    total_videos = len(batch_sessions)
    processed_videos = sum(1 for video in batch_sessions if video['status'] == 'Processed')
    pending_videos = sum(1 for video in batch_sessions if video['status'] == 'Pending')
    failed_videos = sum(1 for video in batch_sessions if video['status'] == 'Failed')
    
    return render_template(
        'batch_review.html',
        batch=batch,
        batch_sessions=batch_sessions,
        videos=batch_sessions,  # Pass batch_sessions as videos for iteration
        total_videos=total_videos,
        processed_videos=processed_videos,
        pending_videos=pending_videos,
        failed_videos=failed_videos
    )

@app.route('/api/team/<team_id>')
def get_team_api(team_id):
    """API endpoint to get team information by ID"""
    try:
        # Get team from database
        team = supabase.table("teams").select("*").eq("id", team_id).execute()
        
        if team.data and len(team.data) > 0:
            return jsonify(team.data[0])
        else:
            return jsonify({"error": "Team not found"}), 404
    except Exception as e:
        print(f"Error retrieving team: {str(e)}")
        return jsonify({"error": str(e)}), 500

def process_video_batch_for_review(batch_id):
    """Process a batch of videos for review"""
    with batch_lock:
        batch_info = batch_processing_tasks[batch_id]
        # Initialize the batch
        batch_info['start_time'] = time.time()
        batch_info['status'] = 'processing'
        batch_info['completed_videos'] = 0
        batch_info['total_videos'] = len(batch_info['videos'])
        batch_info['results'] = {}
        batch_info['log_messages'] = [{
            'message': f"Starting batch processing of {len(batch_info['videos'])} videos for review",
            'type': 'info',
            'timestamp': time.time()
        }]
    
    try:
        # Create a list to track futures
        futures = []
        
        # Submit each video for processing
        for video_info in batch_info['videos']:
            # Create a future for this video
            future = thread_pool_executor.submit(
                process_single_video_for_review_in_batch,
                batch_id,
                video_info['id'],
                video_info['filepath'],
                video_info['home_team_id'],
                video_info['away_team_id'],
                video_info['match_day']
            )
            futures.append(future)
        
        # Wait for all futures to complete
        for future in concurrent.futures.as_completed(futures):
            try:
                # Get the result
                result = future.result()
                print(f"Completed processing video for review: {result}")
                
                # Update batch status
                with batch_lock:
                    batch_info['completed_videos'] += 1
                    progress = (batch_info['completed_videos'] / batch_info['total_videos']) * 100
                    
                    # Add a log message
                    batch_info['log_messages'].append({
                        'message': f"Completed {batch_info['completed_videos']} of {batch_info['total_videos']} videos ({progress:.1f}%)",
                        'type': 'info',
                        'timestamp': time.time()
                    })
                    
                    # Check if all videos are processed
                    if batch_info['completed_videos'] >= batch_info['total_videos']:
                        batch_info['status'] = 'complete'
                        batch_info['end_time'] = time.time()
                        processing_time = batch_info['end_time'] - batch_info['start_time']
                        batch_info['log_messages'].append({
                            'message': f"Batch processing for review completed in {processing_time:.1f} seconds",
                            'type': 'success',
                            'timestamp': time.time()
                        })
            except Exception as e:
                print(f"Error processing video for review in batch: {str(e)}")
                print(traceback.format_exc())
                
                # Update batch with error
                with batch_lock:
                    batch_info['completed_videos'] += 1
                    batch_info['log_messages'].append({
                        'message': f"Error processing a video for review: {str(e)}",
                        'type': 'error',
                        'timestamp': time.time()
                    })
                    
                    # Check if all videos are processed despite errors
                    if batch_info['completed_videos'] >= batch_info['total_videos']:
                        batch_info['status'] = 'complete_with_errors'
                        batch_info['end_time'] = time.time()
    
    except Exception as e:
        print(f"Error in batch processing for review: {str(e)}")
        print(traceback.format_exc())
        
        # Update batch with error
        with batch_lock:
            batch_info['status'] = 'error'
            batch_info['error'] = str(e)
            batch_info['log_messages'].append({
                'message': f"Fatal error in batch processing for review: {str(e)}",
                'type': 'error',
                'timestamp': time.time()
            })

def process_single_video_for_review_in_batch(batch_id, video_id, filepath, home_team_id, away_team_id, match_day):
    """Process a single video for review as part of a batch"""
    try:
        # Log start of processing
        with batch_lock:
            batch_info = batch_processing_tasks[batch_id]
            video_info = next((v for v in batch_info['videos'] if v['id'] == video_id), None)
            if video_info:
                video_info['status'] = 'processing'
                video_info['start_time'] = time.time()
                batch_info['log_messages'].append({
                    'message': f"Processing video for review: {os.path.basename(filepath)}",
                    'type': 'info',
                    'timestamp': time.time(),
                    'video_id': video_id
                })
        
        # Check for timeout periodically
        if check_batch_video_timeout(video_info, timeout_minutes=20):
            with batch_lock:
                video_info['status'] = 'error'
                video_info['error'] = 'Processing timed out after 20 minutes'
                batch_info['log_messages'].append({
                    'message': f"Processing of {os.path.basename(filepath)} for review timed out after 20 minutes",
                    'type': 'error',
                    'timestamp': time.time(),
                    'video_id': video_id
                })
            print(f"Video {video_id} in batch {batch_id} timed out")
            return {
                'video_id': video_id,
                'status': 'error',
                'filepath': filepath,
                'error': 'Timeout'
            }
        
        # Process the video for review
        def progress_callback(message, step=None, details=None, message_type='info'):
            with batch_lock:
                batch_info = batch_processing_tasks[batch_id]
                batch_info['log_messages'].append({
                    'message': message,
                    'type': message_type,
                    'timestamp': time.time(),
                    'video_id': video_id,
                    'step': step,
                    'details': details
                })
        
        # Extract frames for review
        start_time = time.time()
        session_id, error = extract_frames_for_review(
            filepath, home_team_id, away_team_id, match_day,
            callback=progress_callback
        )
        processing_time = time.time() - start_time
        
        # Update batch info with results
        with batch_lock:
            batch_info = batch_processing_tasks[batch_id]
            video_info = next((v for v in batch_info['videos'] if v['id'] == video_id), None)
            
            if video_info:
                video_info['end_time'] = time.time()
                video_info['processing_time'] = processing_time
                
                if error:
                    video_info['status'] = 'error'
                    video_info['error'] = error
                    batch_info['log_messages'].append({
                        'message': f"Error processing {os.path.basename(filepath)} for review: {error}",
                        'type': 'error',
                        'timestamp': time.time(),
                        'video_id': video_id
                    })
                else:
                    video_info['status'] = 'complete'
                    video_info['session_id'] = session_id
                    
                    batch_info['log_messages'].append({
                        'message': f"Successfully processed {os.path.basename(filepath)} for review in {processing_time:.1f}s.",
                        'type': 'success',
                        'timestamp': time.time(),
                        'video_id': video_id
                    })
        
        return {
            'video_id': video_id,
            'status': 'complete' if not error else 'error',
            'filepath': filepath,
            'session_id': session_id
        }
        
    except Exception as e:
        print(f"Error processing video {video_id} for review in batch {batch_id}: {str(e)}")
        print(traceback.format_exc())
        
        # Update batch info with error
        with batch_lock:
            batch_info = batch_processing_tasks[batch_id]
            video_info = next((v for v in batch_info['videos'] if v['id'] == video_id), None)
            
            if video_info:
                video_info['status'] = 'error'
                video_info['error'] = str(e)
                video_info['end_time'] = time.time()
                batch_info['log_messages'].append({
                    'message': f"Error processing {os.path.basename(filepath)} for review: {str(e)}",
                    'type': 'error',
                    'timestamp': time.time(),
                    'video_id': video_id
                })
        
        # Re-raise to be caught by the batch processor
        raise

@app.route('/match/<match_id>/edit')
def edit_match_appearances(match_id):
    """Edit player appearances for a specific match"""
    # Get match data
    match = supabase.table("matches").select("*, home_team:home_team_id(name), away_team:away_team_id(name)").eq("id", match_id).execute()
    
    if not match.data or len(match.data) == 0:
        flash('Match not found', 'danger')
        return redirect(url_for('matches'))
    
    # Get all players for both teams
    home_team_id = match.data[0]['home_team_id']
    away_team_id = match.data[0]['away_team_id']
    
    home_players = supabase.table("players").select("*").eq("team_id", home_team_id).order("name").execute()
    away_players = supabase.table("players").select("*").eq("team_id", away_team_id).order("name").execute()
    
    # Get current appearances for this match
    appearances = supabase.table("appearances").select("*").eq("match_id", match_id).execute()
    
    # Create a set of player IDs who appeared in this match for quick lookup
    appeared_player_ids = set()
    if appearances.data:
        for appearance in appearances.data:
            appeared_player_ids.add(appearance['player_id'])
    
    # Get unmatched players for this match
    unmatched_players_result = supabase.table("unmatched_players").select("*").eq("match_id", match_id).eq("status", "unmatched").execute()
    
    # Format unmatched players data for the template
    unmatched_players = []
    if unmatched_players_result.data and len(unmatched_players_result.data) > 0:
        for player in unmatched_players_result.data:
            team = "home" if player["team_id"] == home_team_id else "away"
            unmatched_players.append({
                "id": player["id"],
                "name": player["name"],
                "team": team
            })
    
    # Format team data
    home_team_name = match.data[0]['home_team']['name'] if 'home_team' in match.data[0] else "Unknown Team"
    away_team_name = match.data[0]['away_team']['name'] if 'away_team' in match.data[0] else "Unknown Team"
    match_day = match.data[0].get('match_day', 'Unknown')
    match_date = match.data[0].get('date', 'Unknown')
    
    return render_template(
        'edit_match_appearances.html',
        match_id=match_id,
        match_day_id=match_id,  # Used by unmatched players forms
        home_team_name=home_team_name,
        away_team_name=away_team_name,
        match_day=match_day,
        match_date=match_date,
        home_players=home_players.data if home_players.data else [],
        away_players=away_players.data if away_players.data else [],
        appeared_player_ids=appeared_player_ids,
        unmatched_players=unmatched_players
    )

@app.route('/match/<match_id>/update_appearances', methods=['POST'])
def update_match_appearances(match_id):
    """Update player appearances for a match"""
    # Get match data to verify it exists
    match = supabase.table("matches").select("id").eq("id", match_id).execute()
    
    if not match.data or len(match.data) == 0:
        return jsonify({"success": False, "error": "Match not found"})
    
    # Get player IDs from form
    player_data = request.json
    
    if not player_data or not isinstance(player_data, dict):
        return jsonify({"success": False, "error": "Invalid player data"})
    
    selected_players = player_data.get('player_ids', [])
    
    try:
        # First, get all current appearances for this match
        current_appearances = supabase.table("appearances").select("id, player_id").eq("match_id", match_id).execute()
        
        # Create sets for easier comparison
        current_player_ids = set()
        if current_appearances.data:
            for appearance in current_appearances.data:
                current_player_ids.add(appearance['player_id'])
        
        selected_player_ids = set(selected_players)
        
        # Players to add (in selected but not in current)
        players_to_add = selected_player_ids - current_player_ids
        
        # Players to remove (in current but not in selected)
        players_to_remove = current_player_ids - selected_player_ids
        
        # Add new appearances
        for player_id in players_to_add:
            supabase.table("appearances").insert({
                "player_id": player_id,
                "match_id": match_id
            }).execute()
            
            # Update total_appearances count in players table
            player = supabase.table("players").select("total_appearances").eq("id", player_id).execute()
            if player.data and len(player.data) > 0:
                current_count = player.data[0].get("total_appearances", 0) or 0
                supabase.table("players").update({"total_appearances": current_count + 1}).eq("id", player_id).execute()
        
        # Remove appearances that are no longer selected
        for player_id in players_to_remove:
            # Find the appearance ID to delete
            appearance_to_remove = next((a for a in current_appearances.data if a['player_id'] == player_id), None)
            
            if appearance_to_remove:
                # Delete the appearance
                supabase.table("appearances").delete().eq("id", appearance_to_remove['id']).execute()
                
                # Update total_appearances count in players table
                player = supabase.table("players").select("total_appearances").eq("id", player_id).execute()
                if player.data and len(player.data) > 0:
                    current_count = player.data[0].get("total_appearances", 0) or 0
                    if current_count > 0:  # Ensure we don't go negative
                        supabase.table("players").update({"total_appearances": current_count - 1}).eq("id", player_id).execute()
        
        return jsonify({
            "success": True,
            "added": len(players_to_add),
            "removed": len(players_to_remove)
        })
        
    except Exception as e:
        print(f"Error updating appearances: {str(e)}")
        return jsonify({"success": False, "error": str(e)})

@app.route('/add_unmatched_player/<match_day_id>', methods=['POST'])
def add_unmatched_player(match_day_id):
    """Add an unmatched player for a specific match day"""
    try:
        # Get data from request
        data = request.json if request.is_json else request.form.to_dict()
        name = data.get('name')
        team = data.get('team')
        
        # Validate input
        if not name or not team:
            return jsonify({"success": False, "error": "Name and team are required"})
        
        # Get match details
        match_result = supabase.table("matches").select("*").eq("id", match_day_id).execute()
        
        if not match_result.data or len(match_result.data) == 0:
            return jsonify({"success": False, "error": "Match not found"})
            
        match = match_result.data[0]
        team_id = match["home_team_id"] if team == "home" else match["away_team_id"]
        
        # Insert the unmatched player
        result = supabase.table("unmatched_players").insert({
            "name": name,
            "team_id": team_id,
            "match_id": match_day_id,
            "status": "unmatched"
        }).execute()
        
        if not result.data or len(result.data) == 0:
            return jsonify({"success": False, "error": "Failed to add player"})
            
        return jsonify({
            "success": True, 
            "player_id": result.data[0]["id"],
            "message": "Player added successfully"
        })
        
    except Exception as e:
        print(f"Error adding unmatched player: {str(e)}")
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})

@app.route('/edit_unmatched_player/<match_day_id>', methods=['POST'])
def edit_unmatched_player(match_day_id):
    """Update an unmatched player for a specific match day"""
    try:
        # Get data from request
        data = request.json if request.is_json else request.form.to_dict()
        player_id = data.get('player_id')
        name = data.get('name')
        team = data.get('team')
        
        # Validate input
        if not player_id or not name or not team:
            return jsonify({"success": False, "error": "Player ID, name, and team are required"})
        
        # Get match details
        match_result = supabase.table("matches").select("*").eq("id", match_day_id).execute()
        
        if not match_result.data or len(match_result.data) == 0:
            return jsonify({"success": False, "error": "Match not found"})
            
        match = match_result.data[0]
        team_id = match["home_team_id"] if team == "home" else match["away_team_id"]
        
        # Update the unmatched player
        result = supabase.table("unmatched_players").update({
            "name": name,
            "team_id": team_id
        }).eq("id", player_id).eq("match_id", match_day_id).execute()
        
        if not result.data or len(result.data) == 0:
            return jsonify({"success": False, "error": "Failed to update player"})
            
        return jsonify({
            "success": True,
            "message": "Player updated successfully"
        })
        
    except Exception as e:
        print(f"Error updating unmatched player: {str(e)}")
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})

@app.route('/delete_unmatched_player/<match_day_id>', methods=['POST'])
def delete_unmatched_player(match_day_id):
    """Delete an unmatched player for a specific match day"""
    try:
        # Get data from request
        data = request.json if request.is_json else request.form.to_dict()
        player_id = data.get('player_id')
        
        # Validate input
        if not player_id:
            return jsonify({"success": False, "error": "Player ID is required"})
        
        # Delete the unmatched player
        result = supabase.table("unmatched_players").delete().eq("id", player_id).eq("match_id", match_day_id).execute()
        
        if not result.data or len(result.data) == 0:
            return jsonify({"success": False, "error": "Failed to delete player"})
            
        return jsonify({
            "success": True,
            "message": "Player deleted successfully"
        })
        
    except Exception as e:
        print(f"Error deleting unmatched player: {str(e)}")
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})

@app.route('/match_player/<match_day_id>', methods=['POST'])
def match_player(match_day_id):
    """Match an unmatched player to an existing player"""
    try:
        # Get data from request
        data = request.json if request.is_json else request.form.to_dict()
        unmatched_player_id = data.get('unmatched_player_id')
        existing_player_id = data.get('existing_player_id')
        
        # Validate input
        if not unmatched_player_id or not existing_player_id:
            return jsonify({"success": False, "error": "Unmatched player ID and existing player ID are required"})
        
        # Get the unmatched player
        unmatched_result = supabase.table("unmatched_players").select("*").eq("id", unmatched_player_id).execute()
        
        if not unmatched_result.data or len(unmatched_result.data) == 0:
            return jsonify({"success": False, "error": "Unmatched player not found"})
            
        unmatched_player = unmatched_result.data[0]
        
        # Update the status of the unmatched player
        status_result = supabase.table("unmatched_players").update({
            "status": "matched",
            "matched_player_id": existing_player_id
        }).eq("id", unmatched_player_id).execute()
        
        if not status_result.data or len(status_result.data) == 0:
            return jsonify({"success": False, "error": "Failed to update unmatched player status"})
        
        # Create an appearance for the existing player if not already present
        appearance_result = supabase.table("appearances").select("*").eq("player_id", existing_player_id).eq("match_id", match_day_id).execute()
        
        if not appearance_result.data or len(appearance_result.data) == 0:
            # Create new appearance
            supabase.table("appearances").insert({
                "player_id": existing_player_id,
                "match_id": match_day_id
            }).execute()
            
            # Update total appearances for the player
            player = supabase.table("players").select("total_appearances").eq("id", existing_player_id).execute()
            if player.data and len(player.data) > 0:
                current_count = player.data[0].get("total_appearances", 0) or 0
                supabase.table("players").update({"total_appearances": current_count + 1}).eq("id", existing_player_id).execute()
            
        return jsonify({
            "success": True,
            "message": "Player matched successfully"
        })
        
    except Exception as e:
        print(f"Error matching player: {str(e)}")
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})

@app.route('/create_match_day', methods=['POST'])
def create_match_day():
    """Create a new match day"""
    try:
        # Get data from request
        data = request.json if request.is_json else request.form.to_dict()
        match_day = data.get('match_day')
        match_date = data.get('match_date')
        home_team_id = data.get('home_team_id')
        away_team_id = data.get('away_team_id')
        
        # Set default date to today if not provided
        if not match_date:
            from datetime import date
            match_date = date.today().isoformat()
        
        # Validate input
        if not match_day or not home_team_id or not away_team_id:
            flash("Match day and team selection are required", "danger")
            return redirect(url_for('matches'))
            
        # Insert the new match day
        result = supabase.table("matches").insert({
            "match_day": match_day,
            "date": match_date,
            "home_team_id": home_team_id,
            "away_team_id": away_team_id,
            "status": "pending"
        }).execute()
        
        if not result.data or len(result.data) == 0:
            flash("Failed to create match day", "danger")
            return redirect(url_for('matches'))
            
        # Redirect to edit appearances page for the new match
        match_id = result.data[0]["id"]
        flash("Match day created successfully", "success")
        return redirect(url_for('edit_match_appearances', match_id=match_id))
        
    except Exception as e:
        print(f"Error creating match day: {str(e)}")
        traceback.print_exc()
        flash(f"Error: {str(e)}", "danger")
        return redirect(url_for('matches'))

@app.route('/reconnect_terminal', methods=['POST'])
def reconnect_terminal():
    """
    Endpoint to handle terminal reconnection requests.
    Handles socket-related issues gracefully.
    """
    try:
        # Get input parameters
        data = request.json
        if not data:
            return jsonify({
                "success": False,
                "error": "No data provided in request"
            })
        
        # Extract either session_id or batch_id (depending on which page is requesting)
        session_id = data.get('session_id')
        batch_id = data.get('batch_id')
        
        # Log the reconnection attempt
        print(f"Terminal reconnection requested: session_id={session_id}, batch_id={batch_id}")
        
        # Initialize success messages
        success_message = "Terminal connection restored"
        log_message = None
        
        # For individual processing - handle reconnection for a specific session
        if session_id:
            with task_lock:
                if session_id in processing_tasks:
                    # Add a log entry to the processing task
                    processing_tasks[session_id]['log_messages'].append({
                        'message': "Terminal reconnection requested",
                        'type': 'info',
                        'timestamp': time.time()
                    })
                    
                    # Set flag that may be useful for backend processes
                    processing_tasks[session_id]['reconnected'] = True
                    
                    success_message = f"Terminal reconnected for session {session_id}"
                    log_message = f"Terminal reconnection successful for session {session_id}"
                else:
                    return jsonify({
                        "success": False,
                        "error": f"Session {session_id} not found or no longer active"
                    })
        
        # For batch processing - handle reconnection for a batch
        elif batch_id:
            with batch_lock:
                if batch_id in batch_processing_tasks:
                    # Add a log entry to the batch processing task
                    batch_processing_tasks[batch_id]['log_messages'].append({
                        'message': "Terminal reconnection requested",
                        'type': 'info',
                        'timestamp': time.time()
                    })
                    
                    # Set flag that may be useful for backend processes
                    batch_processing_tasks[batch_id]['reconnected'] = True
                    
                    success_message = f"Terminal reconnected for batch {batch_id}"
                    log_message = f"Terminal reconnection successful for batch {batch_id}"
                else:
                    return jsonify({
                        "success": False,
                        "error": f"Batch {batch_id} not found or no longer active"
                    })
        else:
            return jsonify({
                "success": False,
                "error": "Neither session_id nor batch_id provided"
            })
                
        # Optional: Perform additional backend operations to ensure connectivity
        # For example, test a socket connection, restart a subprocess, etc.
        try:
            # Socket-related operations would go here - but wrapped in their own try/except
            # This ensures that even if socket operations fail, we still return a valid response
            pass
        except OSError as socket_error:
            # Log the error but continue - don't let socket errors crash the endpoint
            print(f"Socket operation failed but continuing: {str(socket_error)}")
        
        # Log the successful reconnection
        if log_message:
            print(log_message)
        
        # Return success response
        return jsonify({
            "success": True,
            "message": success_message
        })
    
    except OSError as e:
        # Specifically handle Windows socket errors (WinError 10038)
        print(f"Socket error in reconnect_terminal: {str(e)}")
        # Return a partial success - the frontend should still function even if sockets fail
        return jsonify({
            "success": True,
            "message": "Reconnection processed with socket warnings",
            "warning": str(e)
        })
    except Exception as e:
        # Handle any other unexpected errors
        print(f"Error in reconnect_terminal: {str(e)}")
        print(traceback.format_exc())
        return jsonify({
            "success": False,
            "error": str(e)
        })

@app.route('/api/teams')
def get_all_teams_api():
    """Get all teams for dropdowns"""
    result = supabase.table("teams").select("id, name").order("name").execute()
    return jsonify(result.data if result.data else [])

if __name__ == '__main__':
    # Schedule regular file cleanup
    file_manager.schedule_cleanup(app)
    app.run(debug=True) 